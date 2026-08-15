from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, or_
from openpyxl import load_workbook
from typing import List, Optional
import io
import re

from app.database import engine, get_db, Base
from app.models.models import (
    Sede, Cuatrimestre, Catedra, Docente, DocenteSede,
    Alumno, Curso, CatedraCurso, Asignacion, Inscripcion
)

Base.metadata.create_all(bind=engine)

app = FastAPI(title="Sistema Horarios IEA", version="16.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ==================== HELPERS ====================

# v17.0: Áreas de especialidad predefinidas (selector múltiple, sin escritura libre)
AREAS_ESPECIALIDAD = [
    {"id": "contable", "nombre": "Contable y Administrativa", "color": "#2E75B6"},
    {"id": "economicas", "nombre": "Económicas y Finanzas", "color": "#1E8449"},
    {"id": "comercio", "nombre": "Comercio Exterior y Aduana", "color": "#B7950B"},
    {"id": "juridica", "nombre": "Jurídica y Legal", "color": "#6C3483"},
    {"id": "salud", "nombre": "Salud y Acompañamiento", "color": "#C0392B"},
    {"id": "pedagogica", "nombre": "Pedagógica y Educación", "color": "#D35400"},
    {"id": "psicosocial", "nombre": "Psicología y Ciencias Sociales", "color": "#AF7AC5"},
    {"id": "duras", "nombre": "Ciencias Duras (Mat/Fís/Quím)", "color": "#34495E"},
    {"id": "tecnologia", "nombre": "Tecnología, Datos e IA", "color": "#16A085"},
    {"id": "idiomas", "nombre": "Idiomas", "color": "#E67E22"},
    {"id": "comunicacion", "nombre": "Comunicación, Marketing y Publicidad", "color": "#CB4335"},
    {"id": "turismo", "nombre": "Turismo y Hotelería", "color": "#2980B9"},
    {"id": "eventos", "nombre": "Eventos y Producción", "color": "#8E44AD"},
    {"id": "logistica", "nombre": "Logística y Transporte", "color": "#7D6608"},
    {"id": "higiene", "nombre": "Higiene y Seguridad", "color": "#943126"},
    {"id": "agropecuaria", "nombre": "Agropecuaria y Ambiente", "color": "#196F3D"},
    {"id": "rrhh", "nombre": "Recursos Humanos", "color": "#5D6D7E"},
]
AREAS_VALIDAS = {a["id"] for a in AREAS_ESPECIALIDAD}

def sql_set(db, tabla, campo, valor, id_registro):
    """Ejecuta un UPDATE aislado. Si falla hace rollback para no abortar la transacción.
    Devuelve True si guardó. Reemplaza el patrón try/except:pass que rompía todo lo siguiente."""
    from sqlalchemy import text
    try:
        db.execute(text(f"UPDATE {tabla} SET {campo} = :val WHERE id = :id"),
                   {"val": valor, "id": id_registro})
        db.commit()
        return True
    except Exception:
        db.rollback()
        return False

def hora_a_minutos(h):
    """'09:30' -> 570. Devuelve None si no parsea."""
    if not h: return None
    try:
        partes = str(h).strip().replace('.', ':').split(':')
        return int(partes[0]) * 60 + (int(partes[1][:2]) if len(partes) > 1 else 0)
    except Exception:
        return None

def minutos_a_hora(m):
    if m is None: return None
    return f"{int(m)//60:02d}:{int(m)%60:02d}"

def rangos_se_pisan(ini_a, fin_a, ini_b, fin_b, dur_default=90):
    """Compara dos franjas horarias reales. Si no hay hora_fin usa duración por defecto."""
    a1 = hora_a_minutos(ini_a); b1 = hora_a_minutos(ini_b)
    if a1 is None or b1 is None: return False
    a2 = hora_a_minutos(fin_a) or (a1 + dur_default)
    b2 = hora_a_minutos(fin_b) or (b1 + dur_default)
    return a1 < b2 and b1 < a2


def sort_key_codigo(codigo):
    m = re.match(r'c\.(\d+)', codigo or '', re.IGNORECASE)
    return int(m.group(1)) if m else 9999

SEDES_NORMALIZADAS = {
    'avellaneda': 'Avellaneda',
    'caballito': 'Caballito',
    'vicente lopez': 'Vicente López',
    'vicente lópez': 'Vicente López',
    'online interior': 'Online - Interior',
    'online - interior': 'Online - Interior',
    'online- interior': 'Online - Interior',
    'online': 'Online - Interior',
    'liniers': 'Liniers',
}

def normalizar_sede(texto):
    if not texto:
        return None
    t = texto.strip().lower()
    return SEDES_NORMALIZADAS.get(t, texto.strip())

def clasificar_alumno_curso(curso_texto):
    """
    Clasifica un alumno según su CURSO:
    - es_cied: True si dice CIED o es Online/Interior
    - sede_referencia: sede normalizada extraída del paréntesis
    - modalidad_alumno: 'virtual' si CIED/Online, 'presencial' si no
    """
    if not curso_texto:
        return 'virtual', None, True
    curso = curso_texto.strip()
    es_cied = 'CIED' in curso.upper()
    # Extraer sede del paréntesis
    m_sede = re.search(r'\(([^)]+)\)', curso)
    sede_raw = m_sede.group(1).strip() if m_sede else None
    sede = normalizar_sede(sede_raw) if sede_raw else None
    # Online-Interior siempre es virtual (igual que CIED)
    es_online = sede in ['Online - Interior'] if sede else False
    if es_cied or es_online:
        modalidad = 'virtual'
    else:
        modalidad = 'presencial'
    return modalidad, sede, es_cied

def extraer_turno_materia(materia_texto):
    """Extrae el turno de la columna MATERIA: Mañana, Noche, Virtual"""
    if not materia_texto:
        return None
    m = re.search(r'-\s*(Mañana|Noche|Virtual|Tarde)', materia_texto, re.IGNORECASE)
    return m.group(1).capitalize() if m else None


# ==================== MIGRACIÓN ====================

def run_migration(db):
    from sqlalchemy import text, inspect
    resultado = []
    try:
        inspector = inspect(engine)
        tables = inspector.get_table_names()
        resultado.append(f"Tablas: {tables}")
        Base.metadata.create_all(bind=engine)
        resultado.append("✅ create_all")
        inspector = inspect(engine)
        tables = inspector.get_table_names()
        # --- Migrar catedras ---
        if 'catedras' in tables:
            cols = [c['name'] for c in inspector.get_columns('catedras')]
            if 'link_meet' not in cols:
                db.execute(text("ALTER TABLE catedras ADD COLUMN link_meet VARCHAR"))
                db.commit()
                resultado.append("✅ catedras.link_meet")
            for col_c in ['notas', 'decision_apertura']:
                if col_c not in cols:
                    try:
                        db.execute(text(f"ALTER TABLE catedras ADD COLUMN {col_c} VARCHAR"))
                        db.commit()
                    except Exception: db.rollback()
        # --- Migrar asignaciones ---
        if 'asignaciones' in tables:
            cols = [c['name'] for c in inspector.get_columns('asignaciones')]
            if 'hora' in cols and 'hora_inicio' not in cols:
                try:
                    db.execute(text("ALTER TABLE asignaciones RENAME COLUMN hora TO hora_inicio"))
                    db.commit()
                except Exception as e:
                    db.rollback()
            cols = [c['name'] for c in inspector.get_columns('asignaciones')]
            for col, tipo in [
                ('hora_inicio', 'VARCHAR'), ('hora_fin', 'VARCHAR'),
                ('modalidad', "VARCHAR DEFAULT 'virtual_tm'"),
                ('sede_id', 'INTEGER REFERENCES sedes(id)'),
                ('recibe_alumnos_presenciales', 'BOOLEAN DEFAULT FALSE'),
                ('modificada', 'BOOLEAN DEFAULT FALSE'),
            ]:
                if col not in cols:
                    try:
                        db.execute(text(f"ALTER TABLE asignaciones ADD COLUMN {col} {tipo}"))
                        db.commit()
                        resultado.append(f"✅ asignaciones.{col}")
                    except Exception as e:
                        db.rollback()
        # --- Migrar inscripciones (v5.0) ---
        if 'inscripciones' in tables:
            cols = [c['name'] for c in inspector.get_columns('inscripciones')]
            for col, tipo in [
                ('turno', 'VARCHAR'),
                ('modalidad_alumno', 'VARCHAR'),
                ('sede_referencia', 'VARCHAR'),
                ('curso_nombre', 'VARCHAR'),
                ('es_edi', 'BOOLEAN DEFAULT FALSE'),
                ('edi_materia', 'VARCHAR'),
            ]:
                if col not in cols:
                    try:
                        db.execute(text(f"ALTER TABLE inscripciones ADD COLUMN {col} {tipo}"))
                        db.commit()
                        resultado.append(f"✅ inscripciones.{col}")
                    except Exception as e:
                        db.rollback()
        # --- Migrar docentes (v5.0: horas) ---
        if 'docentes' in tables:
            cols = [c['name'] for c in inspector.get_columns('docentes')]
            if 'horas_asignadas' not in cols:
                try:
                    db.execute(text("ALTER TABLE docentes ADD COLUMN horas_asignadas INTEGER DEFAULT 0"))
                    db.commit()
                    resultado.append("✅ docentes.horas_asignadas")
                except Exception as e:
                    db.rollback()
            for col_s in ['sociedad_cfpea', 'sociedad_isftea']:
                if col_s not in cols:
                    try:
                        db.execute(text(f"ALTER TABLE docentes ADD COLUMN {col_s} BOOLEAN DEFAULT FALSE"))
                        db.commit()
                        resultado.append(f"✅ docentes.{col_s}")
                    except Exception as e:
                        db.rollback()
            for col_s in ['materias_av', 'materias_cab', 'materias_vl']:
                if col_s not in cols:
                    try:
                        db.execute(text(f"ALTER TABLE docentes ADD COLUMN {col_s} INTEGER DEFAULT 0"))
                        db.commit()
                    except Exception: db.rollback()
            if 'notas' not in cols:
                try:
                    db.execute(text("ALTER TABLE docentes ADD COLUMN notas VARCHAR"))
                    db.commit()
                except Exception: db.rollback()
            # v16.0: especialidad y catedras de referencia
            for col_new in ['especialidad', 'catedras_referencia']:
                if col_new not in cols:
                    try:
                        db.execute(text(f"ALTER TABLE docentes ADD COLUMN {col_new} VARCHAR"))
                        db.commit()
                        resultado.append(f"✅ Columna {col_new} en docentes")
                    except Exception: db.rollback()
            # v17.0: especialidades (multi-select) + activo
            if 'especialidades' not in cols:
                try:
                    db.execute(text("ALTER TABLE docentes ADD COLUMN especialidades VARCHAR"))
                    db.commit(); resultado.append("✅ docentes.especialidades")
                except Exception: db.rollback()
            if 'activo' not in cols:
                try:
                    db.execute(text("ALTER TABLE docentes ADD COLUMN activo BOOLEAN DEFAULT TRUE"))
                    db.commit(); resultado.append("✅ docentes.activo")
                except Exception: db.rollback()
            # v17.0: DNI deja de ser obligatorio
            try:
                db.execute(text("ALTER TABLE docentes ALTER COLUMN dni DROP NOT NULL"))
                db.commit(); resultado.append("✅ docentes.dni ahora opcional")
            except Exception: db.rollback()
        # --- v17.0: tabla catedra_dictado (se dicta != se abre) ---
        if 'catedra_dictado' not in tables:
            try:
                db.execute(text("""
                    CREATE TABLE IF NOT EXISTS catedra_dictado (
                        id SERIAL PRIMARY KEY,
                        catedra_id INTEGER REFERENCES catedras(id) ON DELETE CASCADE,
                        cuatrimestre_id INTEGER NOT NULL,
                        se_dicta BOOLEAN DEFAULT TRUE,
                        decision_forzada VARCHAR,
                        motivo_forzado VARCHAR,
                        UNIQUE (catedra_id, cuatrimestre_id)
                    )
                """))
                db.commit(); resultado.append("✅ tabla catedra_dictado")
            except Exception: db.rollback()
        else:
            cols_cd = [c['name'] for c in inspector.get_columns('catedra_dictado')]
            for col_cd, tipo_cd in [('decision_forzada', 'VARCHAR'), ('motivo_forzado', 'VARCHAR')]:
                if col_cd not in cols_cd:
                    try:
                        db.execute(text(f"ALTER TABLE catedra_dictado ADD COLUMN {col_cd} {tipo_cd}"))
                        db.commit()
                    except Exception: db.rollback()
        # --- Crear tabla docente_disponibilidad (v5.0) ---
        if 'docente_disponibilidad' not in tables:
            try:
                db.execute(text("""
                    CREATE TABLE IF NOT EXISTS docente_disponibilidad (
                        id SERIAL PRIMARY KEY,
                        docente_id INTEGER REFERENCES docentes(id) ON DELETE CASCADE,
                        dia VARCHAR NOT NULL,
                        hora VARCHAR NOT NULL,
                        disponible BOOLEAN DEFAULT TRUE
                    )
                """))
                db.commit()
                resultado.append("✅ Tabla docente_disponibilidad creada")
            except Exception as e:
                db.rollback()
        # --- Crear tablas auxiliares ---
        for tbl, ddl in [
            ('docente_sede', """CREATE TABLE IF NOT EXISTS docente_sede (
                id SERIAL PRIMARY KEY, docente_id INTEGER REFERENCES docentes(id) ON DELETE CASCADE,
                sede_id INTEGER REFERENCES sedes(id) ON DELETE CASCADE)"""),
            ('catedra_curso', """CREATE TABLE IF NOT EXISTS catedra_curso (
                id SERIAL PRIMARY KEY, catedra_id INTEGER REFERENCES catedras(id) ON DELETE CASCADE,
                curso_id INTEGER REFERENCES cursos(id) ON DELETE CASCADE, turno VARCHAR,
                sede_id INTEGER REFERENCES sedes(id))"""),
            ('plan_carrera', """CREATE TABLE IF NOT EXISTS plan_carrera (
                id SERIAL PRIMARY KEY, sede VARCHAR NOT NULL, carrera VARCHAR NOT NULL,
                anno VARCHAR, codigo_catedra VARCHAR NOT NULL, nombre_catedra VARCHAR,
                dia_tm VARCHAR, hora_tm VARCHAR, dia_tn VARCHAR, hora_tn VARCHAR)"""),
        ]:
            if tbl not in tables:
                try:
                    db.execute(text(ddl))
                    db.commit()
                    resultado.append(f"✅ Tabla {tbl}")
                except Exception as e:
                    db.rollback()
        # --- Limpiar sedes duplicadas ---
        try:
            sede_sin = db.query(Sede).filter(Sede.nombre == "Vicente Lopez").first()
            sede_con = db.query(Sede).filter(Sede.nombre == "Vicente López").first()
            if sede_sin and sede_con:
                for tbl in ['cursos', 'asignaciones']:
                    db.execute(text(f"UPDATE {tbl} SET sede_id = {sede_con.id} WHERE sede_id = {sede_sin.id}"))
                db.execute(text(f"DELETE FROM docente_sede WHERE sede_id = {sede_sin.id} AND docente_id IN (SELECT docente_id FROM docente_sede WHERE sede_id = {sede_con.id})"))
                db.execute(text(f"UPDATE docente_sede SET sede_id = {sede_con.id} WHERE sede_id = {sede_sin.id}"))
                if 'catedra_curso' in tables:
                    db.execute(text(f"UPDATE catedra_curso SET sede_id = {sede_con.id} WHERE sede_id = {sede_sin.id}"))
                db.execute(text(f"DELETE FROM sedes WHERE id = {sede_sin.id}"))
                db.commit()
            elif sede_sin and not sede_con:
                sede_sin.nombre = "Vicente López"
                db.commit()
        except Exception as e:
            db.rollback()
    except Exception as e:
        resultado.append(f"❌ {e}")
    return resultado


@app.on_event("startup")
async def startup():
    db = next(get_db())
    migr = run_migration(db)
    for m in migr:
        print(f"  [MIG] {m}")
    try:
        from app.seed_data import SEDES, CATEDRAS, CURSOS, DOCENTES
        for nombre, color in SEDES:
            if not db.query(Sede).filter(Sede.nombre == nombre).first():
                db.add(Sede(nombre=nombre, color=color))
        db.commit()
        if not db.query(Cuatrimestre).first():
            for anio in range(2026, 2031):
                db.add(Cuatrimestre(nombre=f"1er Cuatrimestre {anio}", anio=anio, numero=1, activo=(anio == 2026)))
                db.add(Cuatrimestre(nombre=f"2do Cuatrimestre {anio}", anio=anio, numero=2, activo=False))
            db.commit()
        else:
            for anio in range(2026, 2031):
                for num in [1, 2]:
                    nombre_c = f"{'1er' if num == 1 else '2do'} Cuatrimestre {anio}"
                    if not db.query(Cuatrimestre).filter(Cuatrimestre.anio == anio, Cuatrimestre.numero == num).first():
                        db.add(Cuatrimestre(nombre=nombre_c, anio=anio, numero=num, activo=False))
            db.commit()
        if db.query(Catedra).count() == 0:
            for codigo, nombre in CATEDRAS:
                db.add(Catedra(codigo=codigo, nombre=nombre))
            db.commit()
        if db.query(Curso).count() == 0:
            for sede_nombre, nombre in CURSOS:
                sede = db.query(Sede).filter(Sede.nombre == sede_nombre).first()
                db.add(Curso(nombre=nombre, sede_id=sede.id if sede else None))
            db.commit()
        if db.query(Docente).count() == 0:
            for dni, nombre, apellido in DOCENTES:
                db.add(Docente(dni=dni, nombre=nombre, apellido=apellido))
            db.commit()
    except ImportError:
        pass
    except Exception as e:
        print(f"  ⚠️ Seed: {e}")
    db.close()
    print("🚀 IEA Horarios v11.0 iniciado")

CLAVE_ACCESO = "IEA2026"

@app.get("/")
def root():
    return {"status": "ok", "sistema": "IEA Horarios v6.0"}

@app.get("/api/reparar")
def reparar_bd(db: Session = Depends(get_db)):
    resultado = run_migration(db)
    try:
        from app.seed_data import SEDES, CATEDRAS, CURSOS, DOCENTES
        for nombre, color in SEDES:
            if not db.query(Sede).filter(Sede.nombre == nombre).first():
                db.add(Sede(nombre=nombre, color=color))
        db.commit()
        for anio in range(2026, 2031):
            for num in [1, 2]:
                nombre_c = f"{'1er' if num == 1 else '2do'} Cuatrimestre {anio}"
                if not db.query(Cuatrimestre).filter(Cuatrimestre.anio == anio, Cuatrimestre.numero == num).first():
                    db.add(Cuatrimestre(nombre=nombre_c, anio=anio, numero=num, activo=False))
        db.commit()
    except Exception as e:
        resultado.append(f"⚠️ {e}")
    return {"resultado": resultado}

@app.get("/api/diagnostico")
def diagnostico_bd(db: Session = Depends(get_db)):
    from sqlalchemy import text, inspect
    info = {}
    try:
        inspector = inspect(engine)
        for t in inspector.get_table_names():
            cols = [c['name'] for c in inspector.get_columns(t)]
            count = db.execute(text(f"SELECT COUNT(*) FROM {t}")).scalar()
            info[t] = {"columnas": cols, "registros": count}
    except Exception as e:
        info["error"] = str(e)
    return info

@app.post("/api/login")
def login(data: dict):
    if data.get("clave", "") == CLAVE_ACCESO:
        return {"ok": True}
    raise HTTPException(status_code=401, detail="Contraseña incorrecta")

@app.get("/api/sedes")
def get_sedes(db: Session = Depends(get_db)):
    return [{"id": s.id, "nombre": s.nombre, "color": s.color} for s in db.query(Sede).all()]

@app.get("/api/cuatrimestres")
def get_cuatrimestres(db: Session = Depends(get_db)):
    return [{"id": c.id, "nombre": c.nombre, "anio": c.anio, "numero": c.numero, "activo": c.activo}
            for c in db.query(Cuatrimestre).order_by(Cuatrimestre.anio, Cuatrimestre.numero).all()]


# ==================== CÁTEDRAS ====================

@app.get("/api/catedras")
def get_catedras(cuatrimestre_id: int = None, db: Session = Depends(get_db)):
    from sqlalchemy import text
    catedras = db.query(Catedra).all()
    catedras_sorted = sorted(catedras, key=lambda c: sort_key_codigo(c.codigo))
    # Pre-cargar desglose: turno × (sede presencial o CIED)
    # IMPORTANTE: solo contar inscripciones que YA fueron clasificadas (tienen turno y modalidad_alumno)
    desglose_query = """
        SELECT catedra_id,
            turno,
            modalidad_alumno,
            sede_referencia,
            COUNT(*) as cnt
        FROM inscripciones
        WHERE modalidad_alumno IS NOT NULL
    """
    if cuatrimestre_id:
        desglose_query += f" AND cuatrimestre_id = {cuatrimestre_id}"
    desglose_query += " GROUP BY catedra_id, turno, modalidad_alumno, sede_referencia"
    # Contar sin clasificar aparte
    sin_clasif_query = "SELECT catedra_id, COUNT(*) FROM inscripciones WHERE modalidad_alumno IS NULL"
    if cuatrimestre_id:
        sin_clasif_query += f" AND cuatrimestre_id = {cuatrimestre_id}"
    sin_clasif_query += " GROUP BY catedra_id"
    try:
        desglose_rows = db.execute(text(desglose_query)).fetchall()
    except Exception:
        desglose_rows = []
    sin_clasif_map = {}
    try:
        for r in db.execute(text(sin_clasif_query)).fetchall():
            sin_clasif_map[r[0]] = r[1]
    except Exception:
        pass
    # Organizar: por cátedra → {turno → {sede → count}}
    desglose_map = {}
    for row in desglose_rows:
        cat_id, turno, mod_alumno, sede, cnt = row[0], row[1], row[2], row[3], row[4]
        if cat_id not in desglose_map:
            desglose_map[cat_id] = {
                'total': 0,
                'tm_av': 0, 'tm_cab': 0, 'tm_vl': 0, 'tm_cied': 0,
                'tn_av': 0, 'tn_cab': 0, 'tn_vl': 0, 'tn_cied': 0,
                'virt_cied': 0, 'sin_clasificar': 0,
            }
        d = desglose_map[cat_id]
        d['total'] += cnt
        es_cied = (mod_alumno == 'virtual')
        sede_norm = (sede or '').strip()
        sede_key = None
        if not es_cied:
            if 'avellaneda' in sede_norm.lower(): sede_key = 'av'
            elif 'caballito' in sede_norm.lower(): sede_key = 'cab'
            elif 'vicente' in sede_norm.lower(): sede_key = 'vl'
        if turno == 'Mañana':
            if es_cied or not sede_key:
                d['tm_cied'] += cnt
            else:
                d[f'tm_{sede_key}'] += cnt
        elif turno == 'Noche':
            if es_cied or not sede_key:
                d['tn_cied'] += cnt
            else:
                d[f'tn_{sede_key}'] += cnt
        else:
            d['virt_cied'] += cnt
    # Agregar sin clasificar
    for cat_id, cnt in sin_clasif_map.items():
        if cat_id not in desglose_map:
            desglose_map[cat_id] = {
                'total': 0, 'tm_av': 0, 'tm_cab': 0, 'tm_vl': 0, 'tm_cied': 0,
                'tn_av': 0, 'tn_cab': 0, 'tn_vl': 0, 'tn_cied': 0,
                'virt_cied': 0, 'sin_clasificar': 0,
            }
        desglose_map[cat_id]['sin_clasificar'] = cnt
        desglose_map[cat_id]['total'] += cnt
    result = []
    for cat in catedras_sorted:
        try:
            asigs = []
            try:
                all_asigs = cat.asignaciones or []
                if cuatrimestre_id:
                    all_asigs = [a for a in all_asigs if a.cuatrimestre_id == cuatrimestre_id]
                for a in all_asigs:
                    asigs.append({
                        "id": a.id, "modalidad": a.modalidad, "dia": a.dia,
                        "hora_inicio": a.hora_inicio, "hora_fin": a.hora_fin,
                        "sede_id": a.sede_id,
                        "sede_nombre": a.sede.nombre if a.sede else None,
                        "recibe_alumnos_presenciales": getattr(a, 'recibe_alumnos_presenciales', False),
                        "docente_id": a.docente_id,
                        "docente": {"id": a.docente.id, "nombre": f"{a.docente.nombre} {a.docente.apellido}"} if a.docente else None,
                    })
            except Exception:
                pass
            desg = desglose_map.get(cat.id, {
                'total': 0, 'tm_av': 0, 'tm_cab': 0, 'tm_vl': 0, 'tm_cied': 0,
                'tn_av': 0, 'tn_cab': 0, 'tn_vl': 0, 'tn_cied': 0, 'virt_cied': 0, 'sin_clasificar': 0,
            })
            inscriptos = desg['total']
            tm_total = desg['tm_av'] + desg['tm_cab'] + desg['tm_vl'] + desg['tm_cied']
            tn_total = desg['tn_av'] + desg['tn_cab'] + desg['tn_vl'] + desg['tn_cied']
            clasificados = tm_total + tn_total + desg['virt_cied']
            # v7.0: Sede totals
            sede_av = desg['tm_av'] + desg['tn_av']
            sede_cab = desg['tm_cab'] + desg['tn_cab']
            sede_vl = desg['tm_vl'] + desg['tn_vl']
            sede_cied = desg['tm_cied'] + desg['tn_cied'] + desg['virt_cied']
            docentes_sugeridos = (1 if inscriptos <= 100 else (1 + -(-max(0, inscriptos - 100) // 100))) if inscriptos >= 10 else 0
            cursos_vinc = []
            try:
                for cc in (cat.cursos or []):
                    cursos_vinc.append({"id": cc.id, "curso_id": cc.curso_id, "curso_nombre": cc.curso.nombre if cc.curso else None, "turno": cc.turno, "sede_nombre": cc.sede.nombre if cc.sede else None})
            except Exception:
                pass
            result.append({
                "id": cat.id, "codigo": cat.codigo, "nombre": cat.nombre,
                "link_meet": getattr(cat, 'link_meet', None),
                "notas": getattr(cat, 'notas', None),
                "decision_apertura": getattr(cat, 'decision_apertura', None),
                "inscriptos": inscriptos,
                "tm_av": desg['tm_av'], "tm_cab": desg['tm_cab'], "tm_vl": desg['tm_vl'], "tm_cied": desg['tm_cied'], "tm_total": tm_total,
                "tn_av": desg['tn_av'], "tn_cab": desg['tn_cab'], "tn_vl": desg['tn_vl'], "tn_cied": desg['tn_cied'], "tn_total": tn_total,
                "virt_cied": desg['virt_cied'],
                "sede_av": sede_av, "sede_cab": sede_cab, "sede_vl": sede_vl, "sede_cied": sede_cied,
                "sin_clasificar": desg['sin_clasificar'],
                "docentes_sugeridos": docentes_sugeridos,
                "cursos_vinculados": cursos_vinc,
                "asignaciones": asigs,
            })
        except Exception:
            result.append({"id": cat.id, "codigo": cat.codigo, "nombre": cat.nombre,
                "link_meet": None, "inscriptos": 0,
                "tm_av": 0, "tm_cab": 0, "tm_vl": 0, "tm_cied": 0, "tm_total": 0,
                "tn_av": 0, "tn_cab": 0, "tn_vl": 0, "tn_cied": 0, "tn_total": 0,
                "virt_cied": 0, "sede_av": 0, "sede_cab": 0, "sede_vl": 0, "sede_cied": 0,
                "sin_clasificar": 0, "docentes_sugeridos": 0,
                "cursos_vinculados": [], "asignaciones": []})
    return result

@app.get("/api/catedras/stats")
def get_catedras_stats(cuatrimestre_id: int = None, db: Session = Depends(get_db)):
    from sqlalchemy import text
    total = db.query(Catedra).count()
    q = db.query(Asignacion)
    if cuatrimestre_id:
        q = q.filter(Asignacion.cuatrimestre_id == cuatrimestre_id)
    asigs = q.all()
    q_insc = db.query(Inscripcion)
    if cuatrimestre_id:
        q_insc = q_insc.filter(Inscripcion.cuatrimestre_id == cuatrimestre_id)
    total_inscripciones = q_insc.count()
    alumnos_unicos = db.query(func.count(func.distinct(Inscripcion.alumno_id)))
    if cuatrimestre_id:
        alumnos_unicos = alumnos_unicos.filter(Inscripcion.cuatrimestre_id == cuatrimestre_id)
    alumnos_unicos = alumnos_unicos.scalar() or 0
    materias_con_inscriptos = db.query(func.count(func.distinct(Inscripcion.catedra_id)))
    if cuatrimestre_id:
        materias_con_inscriptos = materias_con_inscriptos.filter(Inscripcion.cuatrimestre_id == cuatrimestre_id)
    materias_con_inscriptos = materias_con_inscriptos.scalar() or 0
    # Desglose virtual/presencial
    try:
        virt_q = "SELECT COUNT(*) FROM inscripciones WHERE modalidad_alumno = 'virtual'"
        pres_q = "SELECT COUNT(*) FROM inscripciones WHERE modalidad_alumno = 'presencial'"
        if cuatrimestre_id:
            virt_q += f" AND cuatrimestre_id = {cuatrimestre_id}"
            pres_q += f" AND cuatrimestre_id = {cuatrimestre_id}"
        virtuales = db.execute(text(virt_q)).scalar() or 0
        presenciales = db.execute(text(pres_q)).scalar() or 0
    except Exception:
        virtuales = 0
        presenciales = 0
    return {
        "total_catedras": total,
        "catedras_abiertas": len(set(a.catedra_id for a in asigs)),
        "total_asignaciones": len(asigs),
        "con_docente": len([a for a in asigs if a.docente_id and a.modalidad != 'asincronica']),
        "sin_docente": len([a for a in asigs if not a.docente_id and a.modalidad != 'asincronica']),
        "total_inscripciones": total_inscripciones,
        "alumnos_unicos": alumnos_unicos,
        "materias_con_inscriptos": materias_con_inscriptos,
        "inscriptos_virtuales": virtuales,
        "inscriptos_presenciales": presenciales,
    }

@app.put("/api/catedras/{catedra_id}")
def actualizar_catedra(catedra_id: int, data: dict, db: Session = Depends(get_db)):
    cat = db.query(Catedra).filter(Catedra.id == catedra_id).first()
    if not cat:
        raise HTTPException(status_code=404, detail="No encontrada")
    if "nombre" in data: cat.nombre = data["nombre"]
    if "link_meet" in data: cat.link_meet = data["link_meet"]
    from sqlalchemy import text as sql_text
    if "notas" in data:
        try: db.execute(sql_text(f"UPDATE catedras SET notas = :val WHERE id = :id"), {"val": data["notas"], "id": catedra_id})
        except: pass
    if "decision_apertura" in data:
        try: db.execute(sql_text(f"UPDATE catedras SET decision_apertura = :val WHERE id = :id"), {"val": data["decision_apertura"], "id": catedra_id})
        except: pass
    db.commit()
    return {"ok": True}

@app.get("/api/catedras/necesitan-docente")
def get_catedras_necesitan_docente(cuatrimestre_id: int = None, db: Session = Depends(get_db)):
    from sqlalchemy import text
    # Total inscriptos por cátedra
    total_q = "SELECT catedra_id, COUNT(*) FROM inscripciones"
    if cuatrimestre_id: total_q += f" WHERE cuatrimestre_id = {cuatrimestre_id}"
    total_q += " GROUP BY catedra_id"
    total_map = {}
    try:
        for r in db.execute(text(total_q)).fetchall(): total_map[r[0]] = r[1]
    except: pass
    # Desglose por turno+sede para info
    q_txt = """SELECT catedra_id, COALESCE(turno,'Virtual') as turno, COALESCE(modalidad_alumno,'virtual') as mod,
        COALESCE(sede_referencia,'') as sede, COUNT(*) as cnt
        FROM inscripciones WHERE modalidad_alumno IS NOT NULL"""
    if cuatrimestre_id: q_txt += f" AND cuatrimestre_id = {cuatrimestre_id}"
    q_txt += " GROUP BY catedra_id, turno, modalidad_alumno, sede_referencia"
    cat_combos = {}
    try:
        for r in db.execute(text(q_txt)).fetchall():
            cid, turno, mod, sede, cnt = r
            if cid not in cat_combos: cat_combos[cid] = {}
            es_cied = (mod == 'virtual')
            if es_cied: sede_tipo = 'CIED'
            else:
                sn = (sede or '').lower()
                if 'avellaneda' in sn: sede_tipo = 'Avellaneda'
                elif 'caballito' in sn: sede_tipo = 'Caballito'
                elif 'vicente' in sn: sede_tipo = 'Vicente López'
                else: sede_tipo = 'CIED'
            key = f"{turno}|{sede_tipo}"
            cat_combos[cid][key] = cat_combos[cid].get(key, 0) + cnt
    except: pass
    result = []
    all_cats = db.query(Catedra).all()
    for cat in all_cats:
        total = total_map.get(cat.id, 0)
        if total < 10: continue
        # Docentes necesarios vs asignados
        docs_necesarios = 1 if total <= 100 else (1 + -(-max(0, total - 100) // 100))
        asigs = db.query(Asignacion).filter(Asignacion.catedra_id == cat.id)
        if cuatrimestre_id: asigs = asigs.filter(Asignacion.cuatrimestre_id == cuatrimestre_id)
        asigs_list = asigs.all()
        docs_actuales = len([a for a in asigs_list if a.docente_id])
        faltan = max(0, docs_necesarios - docs_actuales)
        if faltan == 0: continue  # Cátedra cubierta, no la mostramos
        # Build desglose for display
        combos = cat_combos.get(cat.id, {})
        aperturas_info = []
        for key, cnt in sorted(combos.items(), key=lambda x: -x[1]):
            turno, sede_tipo = key.split('|')
            aperturas_info.append({'turno': turno, 'sede': sede_tipo, 'inscriptos': cnt})
        # Sedes ya asignadas
        sedes_asignadas = []
        for a in asigs_list:
            if a.docente_id:
                m = a.modalidad or ''
                s = a.sede.nombre if a.sede else 'CIED'
                turno_doc = 'Mañana' if 'tm' in m else ('Noche' if 'tn' in m or m == 'presencial' else 'Virtual')
                sedes_asignadas.append({'turno': turno_doc, 'sede': s,
                    'docente': f"{a.docente.nombre} {a.docente.apellido}" if a.docente else ''})
        result.append({
            "catedra_id": cat.id, "codigo": cat.codigo, "nombre": cat.nombre,
            "inscriptos_total": total, "docs_necesarios": docs_necesarios,
            "docentes_asignados": docs_actuales, "faltan": faltan,
            "docentes_nombres": [f"{a.docente.nombre} {a.docente.apellido}" for a in asigs_list if a.docente],
            "aperturas_info": aperturas_info,
            "sedes_asignadas": sedes_asignadas,
        })
    result.sort(key=lambda x: sort_key_codigo(x['codigo']))
    return result

# ===== v17.0: DICTADO DE CÁTEDRAS (se dicta != se abre) =====
# "Se dicta"  = la cátedra funciona este cuatrimestre (puede ser con video pregrabado).
# "Se abre"   = además tiene docente asignado en vivo.
# Toda cátedra abierta se dicta, pero no toda cátedra dictada se abre.

def _mapa_dictado(db, cuatrimestre_id):
    from sqlalchemy import text
    mapa = {}
    try:
        rows = db.execute(text(
            "SELECT catedra_id, se_dicta, decision_forzada, motivo_forzado "
            "FROM catedra_dictado WHERE cuatrimestre_id = :c"), {"c": cuatrimestre_id}).fetchall()
        for r in rows:
            mapa[r[0]] = {"se_dicta": bool(r[1]), "decision_forzada": r[2], "motivo_forzado": r[3]}
    except Exception:
        db.rollback()
    return mapa

def _inscriptos_por_catedra(db, cuatrimestre_id):
    from sqlalchemy import text
    q = "SELECT catedra_id, COUNT(*) FROM inscripciones"
    if cuatrimestre_id: q += f" WHERE cuatrimestre_id = {int(cuatrimestre_id)}"
    q += " GROUP BY catedra_id"
    mapa = {}
    try:
        for r in db.execute(text(q)).fetchall(): mapa[r[0]] = r[1]
    except Exception:
        db.rollback()
    return mapa

def _estado_catedra(total, tiene_docente, se_dicta, forzada):
    """Devuelve (estado, sugerido). Estado real considera el forzado manual."""
    if total >= 10: sugerido = "ABRIR"
    elif total >= 1: sugerido = "ASINCRONICA"
    else: sugerido = "SIN_ALUMNOS"
    if not se_dicta: return "NO_SE_DICTA", sugerido
    if forzada: return forzada, sugerido
    if tiene_docente: return "ABIERTA", sugerido
    return "ASINCRONICA" if total >= 1 else "SIN_ALUMNOS", sugerido

@app.get("/api/catedras/dictado")
def get_catedras_dictado(cuatrimestre_id: int = None, solo_dictadas: bool = False,
                         buscar: str = None, db: Session = Depends(get_db)):
    """Listado para el selector masivo del Paso 1."""
    if not cuatrimestre_id:
        cu = db.query(Cuatrimestre).first()
        cuatrimestre_id = cu.id if cu else 1
    dictado = _mapa_dictado(db, cuatrimestre_id)
    inscriptos = _inscriptos_por_catedra(db, cuatrimestre_id)
    asigs = db.query(Asignacion).filter(Asignacion.cuatrimestre_id == cuatrimestre_id).all()
    docentes_por_cat = {}
    for a in asigs:
        if a.docente_id:
            docentes_por_cat.setdefault(a.catedra_id, []).append(
                f"{a.docente.nombre} {a.docente.apellido}" if a.docente else "")
    cats = db.query(Catedra).all()
    if buscar:
        b = buscar.lower().strip()
        cats = [c for c in cats if b in (c.codigo or '').lower() or b in (c.nombre or '').lower()]
    resultado = []
    for cat in sorted(cats, key=lambda c: sort_key_codigo(c.codigo)):
        info = dictado.get(cat.id, {})
        se_dicta = info.get("se_dicta", False)
        if solo_dictadas and not se_dicta: continue
        total = inscriptos.get(cat.id, 0)
        docs = docentes_por_cat.get(cat.id, [])
        estado, sugerido = _estado_catedra(total, len(docs) > 0, se_dicta, info.get("decision_forzada"))
        resultado.append({
            "catedra_id": cat.id, "codigo": cat.codigo, "nombre": cat.nombre,
            "se_dicta": se_dicta, "inscriptos": total,
            "docentes": docs, "tiene_docente": len(docs) > 0,
            "estado": estado, "sugerido": sugerido,
            "decision_forzada": info.get("decision_forzada"),
            "motivo_forzado": info.get("motivo_forzado"),
        })
    return {
        "cuatrimestre_id": cuatrimestre_id,
        "total_catedras": len(resultado),
        "se_dictan": len([r for r in resultado if r["se_dicta"]]),
        "abiertas": len([r for r in resultado if r["se_dicta"] and r["tiene_docente"]]),
        "asincronicas": len([r for r in resultado if r["se_dicta"] and not r["tiene_docente"]]),
        "catedras": resultado,
    }

@app.post("/api/catedras/dictado")
def set_catedras_dictado(data: dict, db: Session = Depends(get_db)):
    """Selector masivo. data: {cuatrimestre_id, catedra_ids: [], se_dicta: bool}
    Si viene 'reemplazar_todo': las que no estén en la lista quedan en NO se dicta."""
    from sqlalchemy import text
    cuatrimestre_id = data.get("cuatrimestre_id")
    if not cuatrimestre_id: raise HTTPException(status_code=400, detail="Falta cuatrimestre_id")
    ids = data.get("catedra_ids") or []
    se_dicta = bool(data.get("se_dicta", True))
    reemplazar = bool(data.get("reemplazar_todo", False))
    afectadas = 0
    try:
        if reemplazar:
            db.execute(text("UPDATE catedra_dictado SET se_dicta = FALSE WHERE cuatrimestre_id = :c"),
                       {"c": cuatrimestre_id})
        for cid in ids:
            db.execute(text("""
                INSERT INTO catedra_dictado (catedra_id, cuatrimestre_id, se_dicta)
                VALUES (:cat, :cuat, :sd)
                ON CONFLICT (catedra_id, cuatrimestre_id)
                DO UPDATE SET se_dicta = EXCLUDED.se_dicta
            """), {"cat": cid, "cuat": cuatrimestre_id, "sd": se_dicta})
            afectadas += 1
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Error guardando dictado: {str(e)[:200]}")
    return {"ok": True, "afectadas": afectadas, "se_dicta": se_dicta}

@app.post("/api/catedras/dictado/forzar")
def forzar_decision_catedra(data: dict, db: Session = Depends(get_db)):
    """Permite ir en contra de lo que sugiere el criterio automático.
    decision: 'ABRIR' | 'ASINCRONICA' | 'SIN_ALUMNOS' | null (volver a automático)."""
    from sqlalchemy import text
    cuatrimestre_id = data.get("cuatrimestre_id")
    catedra_id = data.get("catedra_id")
    decision = data.get("decision")
    motivo = data.get("motivo")
    if not cuatrimestre_id or not catedra_id:
        raise HTTPException(status_code=400, detail="Faltan cuatrimestre_id o catedra_id")
    if decision not in (None, "", "ABRIR", "ASINCRONICA", "SIN_ALUMNOS"):
        raise HTTPException(status_code=400, detail="Decisión inválida")
    try:
        db.execute(text("""
            INSERT INTO catedra_dictado (catedra_id, cuatrimestre_id, se_dicta, decision_forzada, motivo_forzado)
            VALUES (:cat, :cuat, TRUE, :dec, :mot)
            ON CONFLICT (catedra_id, cuatrimestre_id)
            DO UPDATE SET decision_forzada = EXCLUDED.decision_forzada,
                          motivo_forzado = EXCLUDED.motivo_forzado
        """), {"cat": catedra_id, "cuat": cuatrimestre_id,
               "dec": decision or None, "mot": motivo or None})
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Error: {str(e)[:200]}")
    return {"ok": True, "decision": decision or "automatico"}

@app.post("/api/catedras/dictado/desde-inscripciones")
def dictado_desde_inscripciones(data: dict, db: Session = Depends(get_db)):
    """Atajo: marca como 'se dicta' toda cátedra que tenga al menos 1 inscripto."""
    from sqlalchemy import text
    cuatrimestre_id = data.get("cuatrimestre_id")
    if not cuatrimestre_id: raise HTTPException(status_code=400, detail="Falta cuatrimestre_id")
    minimo = int(data.get("minimo", 1))
    inscriptos = _inscriptos_por_catedra(db, cuatrimestre_id)
    ids = [cid for cid, n in inscriptos.items() if n >= minimo]
    try:
        for cid in ids:
            db.execute(text("""
                INSERT INTO catedra_dictado (catedra_id, cuatrimestre_id, se_dicta)
                VALUES (:cat, :cuat, TRUE)
                ON CONFLICT (catedra_id, cuatrimestre_id) DO UPDATE SET se_dicta = TRUE
            """), {"cat": cid, "cuat": cuatrimestre_id})
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Error: {str(e)[:200]}")
    return {"ok": True, "marcadas": len(ids), "minimo": minimo}

# ===== v17.0: Administrar el molde de horarios por carrera =====
@app.get("/api/plan-carrera/resumen")
def resumen_plan_carrera(db: Session = Depends(get_db)):
    """Lista las carreras cargadas en el molde, para poder borrarlas selectivamente."""
    from sqlalchemy import text
    try:
        rows = db.execute(text(
            "SELECT sede, carrera, COUNT(*) FROM plan_carrera GROUP BY sede, carrera ORDER BY sede, carrera"
        )).fetchall()
    except Exception:
        db.rollback()
        return {"plan_importado": False, "entradas": [], "total": 0}
    entradas = [{"sede": r[0], "carrera": r[1], "catedras": r[2]} for r in rows]
    return {"plan_importado": len(entradas) > 0, "entradas": entradas,
            "total": sum(e["catedras"] for e in entradas)}

@app.delete("/api/plan-carrera")
def borrar_plan_carrera(sede: str = None, carrera: str = None, todo: bool = False,
                        db: Session = Depends(get_db)):
    """Borra el molde completo (todo=true) o sólo una carrera/sede puntual."""
    from sqlalchemy import text
    try:
        if todo:
            res = db.execute(text("DELETE FROM plan_carrera"))
        elif sede and carrera:
            res = db.execute(text("DELETE FROM plan_carrera WHERE sede = :s AND carrera = :c"),
                             {"s": sede, "c": carrera})
        elif sede:
            res = db.execute(text("DELETE FROM plan_carrera WHERE sede = :s"), {"s": sede})
        elif carrera:
            res = db.execute(text("DELETE FROM plan_carrera WHERE carrera = :c"), {"c": carrera})
        else:
            raise HTTPException(status_code=400, detail="Indicá qué borrar (sede, carrera o todo=true)")
        db.commit()
        return {"ok": True, "borradas": res.rowcount if res.rowcount is not None else 0}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Error borrando molde: {str(e)[:200]}")

# ===== v17.0: Asignar docente directo desde "Necesitan Docente" =====
@app.post("/api/catedras/{catedra_id}/asignar-docente")
def asignar_docente_rapido(catedra_id: int, data: dict, db: Session = Depends(get_db)):
    """Asigna un docente a la cátedra sin salir de la pantalla.
    Si hay una asignación con horario y sin docente, la completa. Si no, crea una nueva."""
    from sqlalchemy import text
    docente_id = data.get("docente_id")
    cuatrimestre_id = data.get("cuatrimestre_id")
    if not cuatrimestre_id: raise HTTPException(status_code=400, detail="Falta cuatrimestre_id")
    cat = db.query(Catedra).filter(Catedra.id == catedra_id).first()
    if not cat: raise HTTPException(status_code=404, detail="Cátedra no encontrada")
    doc = db.query(Docente).filter(Docente.id == docente_id).first() if docente_id else None
    if docente_id and not doc: raise HTTPException(status_code=404, detail="Docente no encontrado")

    asignacion_id = data.get("asignacion_id")
    if asignacion_id:
        asig = db.query(Asignacion).filter(Asignacion.id == asignacion_id).first()
    else:
        asig = db.query(Asignacion).filter(
            Asignacion.catedra_id == catedra_id,
            Asignacion.cuatrimestre_id == cuatrimestre_id,
            Asignacion.docente_id.is_(None)).first()
    creada = False
    if not asig:
        asig = Asignacion(catedra_id=catedra_id, cuatrimestre_id=cuatrimestre_id,
                          modalidad='presencial_virtual')
        db.add(asig); creada = True
    asig.docente_id = docente_id
    for campo in ("dia", "hora_inicio", "sede_id"):
        if data.get(campo) is not None: setattr(asig, campo, data[campo])
    try:
        db.commit(); db.refresh(asig)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Error: {str(e)[:200]}")
    if data.get("hora_fin"):
        sql_set(db, "asignaciones", "hora_fin", data["hora_fin"], asig.id)
    # Al tener docente, la cátedra se dicta y queda abierta
    try:
        db.execute(text("""
            INSERT INTO catedra_dictado (catedra_id, cuatrimestre_id, se_dicta)
            VALUES (:cat, :cuat, TRUE)
            ON CONFLICT (catedra_id, cuatrimestre_id) DO UPDATE SET se_dicta = TRUE
        """), {"cat": catedra_id, "cuat": cuatrimestre_id})
        db.commit()
    except Exception: db.rollback()
    if docente_id: sql_set(db, "catedras", "decision_apertura", "ABRIR", catedra_id)
    return {"ok": True, "asignacion_id": asig.id, "creada": creada,
            "docente": f"{doc.nombre} {doc.apellido}" if doc else None}

# ===== v8.0: Criterio de apertura simplificado =====
@app.get("/api/catedras/criterio-apertura")
def get_criterio_apertura(cuatrimestre_id: int = None, db: Session = Depends(get_db)):
    """
    >=10 total → ABRIR. 1 doc hasta 100, luego +1 cada 100 adicionales.
    1-9 total → ASINCRÓNICA
    0 → SIN ALUMNOS
    """
    from sqlalchemy import text
    total_q = "SELECT catedra_id, COUNT(*) FROM inscripciones"
    if cuatrimestre_id: total_q += f" WHERE cuatrimestre_id = {cuatrimestre_id}"
    total_q += " GROUP BY catedra_id"
    total_map = {}
    try:
        for r in db.execute(text(total_q)).fetchall(): total_map[r[0]] = r[1]
    except: pass
    all_cats = sorted(db.query(Catedra).all(), key=lambda c: sort_key_codigo(c.codigo))
    abrir = []; asincronica = []; sin_alumnos = []
    for cat in all_cats:
        total = total_map.get(cat.id, 0)
        if total == 0:
            sin_alumnos.append({"codigo": cat.codigo, "nombre": cat.nombre, "total": 0})
        elif total < 10:
            asincronica.append({"codigo": cat.codigo, "nombre": cat.nombre, "total": total})
        else:
            docs = 1 if total <= 100 else (1 + -(-max(0, total - 100) // 100))
            # Check if already has asignacion
            tiene = db.query(Asignacion).filter(Asignacion.catedra_id == cat.id)
            if cuatrimestre_id: tiene = tiene.filter(Asignacion.cuatrimestre_id == cuatrimestre_id)
            tiene_asig = tiene.count() > 0
            docs_actuales = tiene.filter(Asignacion.docente_id.isnot(None)).count() if tiene_asig else 0
            abrir.append({"codigo": cat.codigo, "nombre": cat.nombre, "total": total,
                "docentes_sugeridos": docs, "docentes_actuales": docs_actuales,
                "faltan": max(0, docs - docs_actuales), "tiene_asignacion": tiene_asig})
    return {"abrir": abrir, "asincronica": asincronica, "sin_alumnos": sin_alumnos,
        "stats": {"total_abrir": len(abrir), "total_asincronica": len(asincronica),
            "total_sin_alumnos": len(sin_alumnos),
            "total_docentes_sugeridos": sum(a["docentes_sugeridos"] for a in abrir)}}

# ===== v12.0: Auto-marcar asincrónicas como "No abrir" =====
@app.post("/api/catedras/auto-decision-asincronicas")
def auto_decision_asincronicas(cuatrimestre_id: int = None, db: Session = Depends(get_db)):
    from sqlalchemy import text
    criterio = get_criterio_apertura(cuatrimestre_id, db)
    count = 0
    for item in criterio['asincronica']:
        cat = db.query(Catedra).filter(Catedra.codigo == item['codigo']).first()
        if cat:
            try:
                db.execute(text("UPDATE catedras SET decision_apertura = :val WHERE id = :id"),
                    {"val": "Asincrónica", "id": cat.id})
                count += 1
            except: pass
    for item in criterio['sin_alumnos']:
        cat = db.query(Catedra).filter(Catedra.codigo == item['codigo']).first()
        if cat:
            try:
                db.execute(text("UPDATE catedras SET decision_apertura = :val WHERE id = :id"),
                    {"val": "No abrir", "id": cat.id})
                count += 1
            except: pass
    db.commit()
    return {"marcadas": count, "asincronicas": len(criterio['asincronica']), "sin_alumnos": len(criterio['sin_alumnos'])}


@app.get("/api/inscriptos/por-curso")
def get_inscriptos_por_curso(cuatrimestre_id: int = None, db: Session = Depends(get_db)):
    from sqlalchemy import text
    q = """SELECT curso_nombre, 
        COUNT(*) as total_inscripciones,
        COUNT(DISTINCT alumno_id) as alumnos_unicos
        FROM inscripciones WHERE curso_nombre IS NOT NULL AND curso_nombre != ''"""
    if cuatrimestre_id:
        q += f" AND cuatrimestre_id = {cuatrimestre_id}"
    q += " GROUP BY curso_nombre ORDER BY total_inscripciones DESC"
    try:
        rows = db.execute(text(q)).fetchall()
        result = []
        for r in rows:
            curso_raw = r[0]; total_insc = r[1]; alumnos = r[2]
            nombre_limpio = re.split(r'\s*[-–]\s*(?:CIED|Cursada)', curso_raw, maxsplit=1)[0].strip()
            nombre_limpio = re.split(r'\s*\(', nombre_limpio, maxsplit=1)[0].strip()
            es_cied = 'CIED' in curso_raw.upper()
            m_sede = re.search(r'\(([^)]+)\)', curso_raw)
            sede_raw = m_sede.group(1).strip() if m_sede else ''
            sede = normalizar_sede(sede_raw) if sede_raw else 'Sin sede'
            es_online = sede in ['Online - Interior']
            modalidad = 'CIED' if (es_cied or es_online) else 'Presencial'
            es_bce = 'BCE' in curso_raw.upper() or 'SECUNDARIO' in curso_raw.upper()
            es_bea = 'BEA' in curso_raw.upper()
            tipo_curso = 'BCE' if es_bce else ('BEA' if es_bea else 'Superior')
            result.append({
                "curso_completo": curso_raw, "curso_nombre": nombre_limpio,
                "sede": sede, "modalidad": modalidad, "tipo_curso": tipo_curso,
                "inscripciones": total_insc, "alumnos_unicos": alumnos,
            })
        return result
    except Exception as e:
        return []


# ==================== ASIGNACIONES ====================

@app.post("/api/asignaciones")
def crear_asignacion(data: dict, db: Session = Depends(get_db)):
    try:
        cat_id = data.get("catedra_id")
        cuat_id = data.get("cuatrimestre_id", 1)
        modalidad = data.get("modalidad", "virtual_tm")
        dia = data.get("dia")
        hora = data.get("hora_inicio")
        sede_id = data.get("sede_id")
        docente_id = data.get("docente_id")
        if dia and hora and modalidad != 'asincronica':
            conflict = verificar_solapamiento_catedra(cat_id, dia, hora, cuat_id, None, db)
            if conflict:
                raise HTTPException(status_code=400, detail=conflict)
        # v9.0: Verificar disponibilidad del docente
        if docente_id and dia and hora:
            from sqlalchemy import text
            try:
                disp_rows = db.execute(text(f"SELECT COUNT(*) FROM docente_disponibilidad WHERE docente_id = {docente_id}")).scalar()
                if disp_rows and disp_rows > 0:
                    tiene_disp = db.execute(text(
                        f"SELECT disponible FROM docente_disponibilidad WHERE docente_id = {docente_id} AND dia = '{dia}' AND hora = '{hora}'"
                    )).fetchone()
                    if not tiene_disp or not tiene_disp[0]:
                        raise HTTPException(status_code=400, detail=f"⛔ El docente no tiene disponibilidad el {dia} a las {hora}. Revisá su disponibilidad horaria antes de asignarle.")
            except HTTPException: raise
            except Exception: pass
        asig = Asignacion(
            catedra_id=cat_id, cuatrimestre_id=cuat_id,
            docente_id=docente_id if docente_id else None,
            modalidad=modalidad,
            dia=dia if dia else None, hora_inicio=hora if hora else None,
            hora_fin=data.get("hora_fin") if data.get("hora_fin") else None,
            sede_id=sede_id if sede_id else None,
            recibe_alumnos_presenciales=data.get("recibe_alumnos_presenciales", False),
        )
        db.add(asig); db.commit()
        return {"id": asig.id, "ok": True}
    except HTTPException: raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

@app.put("/api/asignaciones/{asignacion_id}")
def actualizar_asignacion(asignacion_id: int, data: dict, db: Session = Depends(get_db)):
    try:
        asig = db.query(Asignacion).filter(Asignacion.id == asignacion_id).first()
        if not asig: raise HTTPException(status_code=404, detail="No encontrada")
        dia = data.get("dia", asig.dia)
        hora = data.get("hora_inicio", asig.hora_inicio)
        if dia and hora and data.get("modalidad", asig.modalidad) != 'asincronica':
            conflict = verificar_solapamiento_catedra(asig.catedra_id, dia, hora, asig.cuatrimestre_id, asig.id, db)
            if conflict: raise HTTPException(status_code=400, detail=conflict)
        for field in ["docente_id", "modalidad", "dia", "hora_inicio", "hora_fin", "sede_id", "recibe_alumnos_presenciales"]:
            if field in data:
                val = data[field]
                if field in ["docente_id", "sede_id"] and (val == "" or val == "null" or not val):
                    val = None
                setattr(asig, field, val if val else None)
        asig.modificada = True
        db.commit()
        return {"ok": True}
    except HTTPException: raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

@app.delete("/api/asignaciones/{asignacion_id}")
def eliminar_asignacion(asignacion_id: int, db: Session = Depends(get_db)):
    asig = db.query(Asignacion).filter(Asignacion.id == asignacion_id).first()
    if not asig: raise HTTPException(status_code=404, detail="No encontrada")
    db.delete(asig); db.commit()
    return {"ok": True}


# ==================== DOCENTES ====================

def calcular_tipo_modalidad(docente, db):
    asigs = db.query(Asignacion).filter(Asignacion.docente_id == docente.id).all()
    if not asigs: return "SIN_ASIGNACIONES"
    if any(a.recibe_alumnos_presenciales for a in asigs): return "PRESENCIAL_VIRTUAL"
    if any(a.sede_id for a in asigs): return "SEDE_VIRTUAL"
    return "REMOTO"

@app.get("/api/docentes")
def get_docentes(cuatrimestre_id: int = None, orden: str = "apellido",
                 solo_activos_cuatri: bool = False, db: Session = Depends(get_db)):
    """orden: 'apellido' (Apellido, Nombre) o 'nombre'.
    solo_activos_cuatri: sólo docentes con asignaciones en el cuatrimestre indicado."""
    from sqlalchemy import text
    if orden == "nombre":
        docentes = db.query(Docente).order_by(Docente.nombre, Docente.apellido).all()
    else:
        docentes = db.query(Docente).order_by(Docente.apellido, Docente.nombre).all()
    # Set de docentes con actividad en el cuatrimestre
    activos_cuatri = set()
    if cuatrimestre_id:
        for (did,) in db.query(Asignacion.docente_id).filter(
                Asignacion.cuatrimestre_id == cuatrimestre_id,
                Asignacion.docente_id.isnot(None)).distinct().all():
            activos_cuatri.add(did)
    if solo_activos_cuatri and cuatrimestre_id:
        docentes = [d for d in docentes if d.id in activos_cuatri]
    # Pre-load disponibilidad
    disp_map = {}
    try:
        for r in db.execute(text("SELECT docente_id, dia, hora, disponible FROM docente_disponibilidad WHERE disponible = TRUE")).fetchall():
            if r[0] not in disp_map: disp_map[r[0]] = []
            disp_map[r[0]].append(f"{r[1]} {r[2]}")
    except: pass

    # v17.2 FIX DE RAÍZ: las columnas agregadas con ALTER TABLE (especialidades, notas,
    # catedras_referencia, activo, horas...) NO están declaradas en el modelo ORM de Docente.
    # Por eso getattr(d, 'especialidades') devolvía siempre None: el dato SÍ se guardaba,
    # pero nunca se leía de vuelta. Se leen acá con SQL directo.
    extra = {}
    for col in ['especialidades', 'especialidad', 'catedras_referencia', 'notas',
                'activo', 'horas_asignadas', 'sociedad_cfpea', 'sociedad_isftea',
                'materias_av', 'materias_cab', 'materias_vl']:
        try:
            for did, val in db.execute(text(f"SELECT id, {col} FROM docentes")).fetchall():
                extra.setdefault(did, {})[col] = val
        except Exception:
            db.rollback()
    result = []
    for d in docentes:
        try:
            asigs_data = []
            all_asigs = d.asignaciones or []
            if cuatrimestre_id:
                all_asigs = [a for a in all_asigs if a.cuatrimestre_id == cuatrimestre_id]
            for a in all_asigs:
                asigs_data.append({
                    "id": a.id, "modalidad": a.modalidad, "dia": a.dia,
                    "hora_inicio": a.hora_inicio,
                    "catedra_codigo": a.catedra.codigo if a.catedra else None,
                    "catedra_nombre": a.catedra.nombre if a.catedra else None,
                    "sede_nombre": a.sede.nombre if a.sede else None,
                    "recibe_alumnos_presenciales": getattr(a, 'recibe_alumnos_presenciales', False),
                })
            sedes_data = [{"id": ds.sede.id, "nombre": ds.sede.nombre} for ds in (d.sedes or []) if ds.sede]
            tipo = calcular_tipo_modalidad(d, db)
            ex = extra.get(d.id, {})   # valores leídos por SQL directo
            horas = ex.get('horas_asignadas') or 0
            cfpea = bool(ex.get('sociedad_cfpea'))
            isftea = bool(ex.get('sociedad_isftea'))
            mat_av = ex.get('materias_av') or 0
            mat_cab = ex.get('materias_cab') or 0
            mat_vl = ex.get('materias_vl') or 0
            # v16.0: availability summary
            disp_list = disp_map.get(d.id, [])
            disp_resumen = f"{len(disp_list)} franjas" if disp_list else "Sin asignar"
            result.append({
                "id": d.id, "dni": d.dni, "nombre": d.nombre, "apellido": d.apellido,
                "email": d.email, "tipo_modalidad": tipo,
                "horas_asignadas": horas, "notas": ex.get('notas'),
                "sociedad_cfpea": cfpea, "sociedad_isftea": isftea,
                "materias_av": mat_av, "materias_cab": mat_cab, "materias_vl": mat_vl,
                "especialidad": ex.get('especialidad'),
                "catedras_referencia": ex.get('catedras_referencia'),
                "especialidades": [x for x in (ex.get('especialidades') or '').split(',') if x],
                "activo": True if ex.get('activo') is None else bool(ex.get('activo')),
                "activo_cuatrimestre": d.id in activos_cuatri,
                "nombre_completo": f"{d.apellido or ''}, {d.nombre or ''}".strip(' ,'),
                "disponibilidad_resumen": disp_resumen,
                "disponibilidad_franjas": disp_list[:6],
                "sedes": sedes_data, "asignaciones": asigs_data,
            })
        except Exception:
            result.append({"id": d.id, "dni": d.dni, "nombre": d.nombre, "apellido": d.apellido,
                "email": d.email, "tipo_modalidad": "SIN_ASIGNACIONES",
                "horas_asignadas": 0, "sociedad_cfpea": False, "sociedad_isftea": False,
                "materias_av": 0, "materias_cab": 0, "materias_vl": 0,
                "especialidad": None, "catedras_referencia": None,
                "especialidades": [], "activo": True, "activo_cuatrimestre": d.id in activos_cuatri,
                "nombre_completo": f"{d.apellido or ''}, {d.nombre or ''}".strip(' ,'),
                "disponibilidad_resumen": "Sin asignar", "disponibilidad_franjas": [],
                "sedes": [], "asignaciones": []})
    return result

@app.post("/api/docentes")
def crear_docente(data: dict, db: Session = Depends(get_db)):
    """v17.0: el DNI dejó de ser obligatorio. Sólo se exige nombre o apellido."""
    dni = (data.get("dni") or "").strip() or None
    nombre = (data.get("nombre") or "").strip()
    apellido = (data.get("apellido") or "").strip()
    if not nombre and not apellido:
        raise HTTPException(status_code=400, detail="Indicá al menos nombre o apellido")
    if dni and db.query(Docente).filter(Docente.dni == dni).first():
        raise HTTPException(status_code=400, detail="Ese DNI ya está cargado en otro docente")
    d = Docente(dni=dni, nombre=nombre, apellido=apellido, email=data.get("email"))
    db.add(d); db.commit()
    if data.get("especialidades"):
        crudo = data["especialidades"]
        if isinstance(crudo, str): crudo = [x.strip() for x in crudo.split(',')]
        elegidas = [x for x in crudo if x in AREAS_VALIDAS]
        if elegidas: sql_set(db, "docentes", "especialidades", ','.join(elegidas), d.id)
    return {"id": d.id, "ok": True}

@app.get("/api/docentes/{docente_id}/ficha")
def get_ficha_docente(docente_id: int, db: Session = Depends(get_db)):
    """Datos permanentes del docente (no dependen del cuatrimestre).
    Lee las columnas con SQL directo porque no están declaradas en el modelo ORM."""
    from sqlalchemy import text
    d = db.query(Docente).filter(Docente.id == docente_id).first()
    if not d: raise HTTPException(status_code=404, detail="No encontrado")
    ficha = {"id": d.id, "nombre": d.nombre or "", "apellido": d.apellido or "",
             "dni": d.dni or "", "email": d.email or "",
             "especialidades": [], "catedras_referencia": "", "notas": ""}
    try:
        row = db.execute(text(
            "SELECT especialidades, catedras_referencia, notas FROM docentes WHERE id = :id"
        ), {"id": docente_id}).fetchone()
        if row:
            ficha["especialidades"] = [x for x in (row[0] or '').split(',') if x]
            ficha["catedras_referencia"] = row[1] or ""
            ficha["notas"] = row[2] or ""
    except Exception:
        db.rollback()
    return ficha


@app.put("/api/docentes/{docente_id}/ficha")
def guardar_ficha_docente(docente_id: int, data: dict, db: Session = Depends(get_db)):
    """Guarda los datos permanentes del docente y devuelve lo que quedó realmente grabado,
    releyéndolo de la base. Si lo devuelto coincide con lo enviado, el guardado es real."""
    from sqlalchemy import text
    d = db.query(Docente).filter(Docente.id == docente_id).first()
    if not d: raise HTTPException(status_code=404, detail="No encontrado")

    for campo in ["nombre", "apellido", "email"]:
        if campo in data: setattr(d, campo, (data[campo] or "").strip() or None)
    if "dni" in data:
        nuevo_dni = str(data["dni"] or "").strip() or None
        if nuevo_dni and nuevo_dni != d.dni:
            otro = db.query(Docente).filter(Docente.dni == nuevo_dni, Docente.id != docente_id).first()
            if otro: raise HTTPException(status_code=400, detail="Ese DNI ya pertenece a otro docente")
        d.dni = nuevo_dni
    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Error guardando datos básicos: {str(e)[:150]}")

    if "especialidades" in data:
        crudo = data["especialidades"]
        if isinstance(crudo, str): crudo = [x.strip() for x in crudo.split(',') if x.strip()]
        elegidas = [x for x in (crudo or []) if x in AREAS_VALIDAS]
        sql_set(db, "docentes", "especialidades", ','.join(elegidas) if elegidas else None, docente_id)
    if "catedras_referencia" in data:
        sql_set(db, "docentes", "catedras_referencia", (data["catedras_referencia"] or "").strip() or None, docente_id)
    if "notas" in data:
        sql_set(db, "docentes", "notas", (data["notas"] or "").strip() or None, docente_id)

    return get_ficha_docente(docente_id, db)


@app.get("/api/areas-especialidad")
def get_areas_especialidad():
    """Catálogo fijo de áreas. El frontend arma el selector múltiple con esto."""
    return AREAS_ESPECIALIDAD

@app.put("/api/docentes/{docente_id}")
def actualizar_docente(docente_id: int, data: dict, db: Session = Depends(get_db)):
    """v17.0: reescrito. Cada campo se guarda de forma aislada con commit propio,
    así un fallo en uno no aborta la transacción y anula todos los demás (bug histórico)."""
    d = db.query(Docente).filter(Docente.id == docente_id).first()
    if not d: raise HTTPException(status_code=404, detail="No encontrado")
    guardados = []; fallidos = []

    # Campos del modelo ORM
    for field in ["nombre", "apellido", "email", "dni"]:
        if field in data:
            valor = data[field]
            if field == "dni":
                valor = (str(valor).strip() or None) if valor is not None else None
            setattr(d, field, valor)
    try:
        db.commit(); guardados.append("datos_basicos")
    except Exception:
        db.rollback(); fallidos.append("datos_basicos")

    # Numéricos: tolera "" y None sin romper
    for fld in ['horas_asignadas', 'materias_av', 'materias_cab', 'materias_vl']:
        if fld in data:
            crudo = data[fld]
            try:
                num = 0 if crudo in (None, "", "null") else int(float(crudo))
            except Exception:
                num = 0
            (guardados if sql_set(db, "docentes", fld, num, docente_id) else fallidos).append(fld)

    # Booleanos
    for fld in ['sociedad_cfpea', 'sociedad_isftea', 'activo']:
        if fld in data:
            (guardados if sql_set(db, "docentes", fld, bool(data[fld]), docente_id) else fallidos).append(fld)

    # Texto libre
    for fld in ['notas', 'especialidad', 'catedras_referencia']:
        if fld in data:
            valor = data[fld] if data[fld] not in (None, "") else None
            (guardados if sql_set(db, "docentes", fld, valor, docente_id) else fallidos).append(fld)

    # v17.0: especialidades como lista de IDs validados contra el catálogo
    if 'especialidades' in data:
        crudo = data['especialidades']
        if isinstance(crudo, str):
            crudo = [x.strip() for x in crudo.split(',') if x.strip()]
        elegidas = [x for x in (crudo or []) if x in AREAS_VALIDAS]
        valor = ','.join(elegidas) if elegidas else None
        (guardados if sql_set(db, "docentes", "especialidades", valor, docente_id) else fallidos).append("especialidades")

    return {"ok": len(fallidos) == 0, "guardados": guardados, "fallidos": fallidos}

@app.delete("/api/docentes/{docente_id}")
def eliminar_docente(docente_id: int, db: Session = Depends(get_db)):
    d = db.query(Docente).filter(Docente.id == docente_id).first()
    if not d: raise HTTPException(status_code=404, detail="No encontrado")
    db.query(Asignacion).filter(Asignacion.docente_id == docente_id).update({"docente_id": None})
    db.query(DocenteSede).filter(DocenteSede.docente_id == docente_id).delete()
    from sqlalchemy import text
    try: db.execute(text(f"DELETE FROM docente_disponibilidad WHERE docente_id = {docente_id}"))
    except: pass
    db.delete(d); db.commit()
    return {"ok": True}

@app.put("/api/docentes/{docente_id}/sedes")
def actualizar_sedes_docente(docente_id: int, data: dict, db: Session = Depends(get_db)):
    d = db.query(Docente).filter(Docente.id == docente_id).first()
    if not d: raise HTTPException(status_code=404, detail="No encontrado")
    db.query(DocenteSede).filter(DocenteSede.docente_id == docente_id).delete()
    for sid in data.get("sede_ids", []):
        db.add(DocenteSede(docente_id=docente_id, sede_id=sid))
    db.commit()
    return {"ok": True}

# ===== v5.0: Disponibilidad horaria =====
@app.get("/api/docentes/{docente_id}/disponibilidad")
def get_disponibilidad(docente_id: int, db: Session = Depends(get_db)):
    from sqlalchemy import text
    try:
        rows = db.execute(text(f"SELECT dia, hora, disponible FROM docente_disponibilidad WHERE docente_id = {docente_id}")).fetchall()
        return [{"dia": r[0], "hora": r[1], "disponible": r[2]} for r in rows]
    except Exception:
        return []

@app.put("/api/docentes/{docente_id}/disponibilidad")
def set_disponibilidad(docente_id: int, data: dict, db: Session = Depends(get_db)):
    from sqlalchemy import text
    try:
        db.execute(text(f"DELETE FROM docente_disponibilidad WHERE docente_id = {docente_id}"))
        for item in data.get("disponibilidad", []):
            dia = item.get("dia")
            hora = item.get("hora")
            disponible = item.get("disponible", True)
            if dia and hora:
                db.execute(text(
                    f"INSERT INTO docente_disponibilidad (docente_id, dia, hora, disponible) VALUES ({docente_id}, '{dia}', '{hora}', {disponible})"
                ))
        db.commit()
        return {"ok": True}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


# ==================== CURSOS ====================

@app.get("/api/cursos")
def get_cursos(sede_id: int = None, db: Session = Depends(get_db)):
    q = db.query(Curso)
    if sede_id: q = q.filter(Curso.sede_id == sede_id)
    result = []
    for c in q.order_by(Curso.nombre).all():
        catedras_vinc = []
        try:
            for cc in (c.catedras or []):
                catedras_vinc.append({"id": cc.id, "catedra_id": cc.catedra_id, "catedra_codigo": cc.catedra.codigo if cc.catedra else None, "catedra_nombre": cc.catedra.nombre if cc.catedra else None, "turno": cc.turno})
        except Exception: pass
        result.append({"id": c.id, "nombre": c.nombre, "sede_id": c.sede_id, "sede_nombre": c.sede.nombre if c.sede else None, "cant_catedras": len(catedras_vinc), "catedras": catedras_vinc})
    return result


# ==================== IMPORTACIONES ====================

@app.post("/api/importar/catedras")
async def importar_catedras(file: UploadFile = File(...), db: Session = Depends(get_db)):
    try:
        content = await file.read()
        wb = load_workbook(filename=io.BytesIO(content), read_only=True)
        ws = None
        for name in wb.sheetnames:
            if "catedr" in name.lower() or "cátedr" in name.lower(): ws = wb[name]; break
        if ws is None: ws = wb[wb.sheetnames[0]]
        creadas = actualizadas = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            vals = [str(c).strip() if c is not None else "" for c in row]
            codigo = nombre = None
            if len(vals) >= 2 and vals[1]:
                m = re.match(r'^(c\.\d+)\s+(.+)', vals[1], re.IGNORECASE)
                if m: codigo, nombre = m.group(1), m.group(2).strip()
            if not codigo and len(vals) >= 2:
                try:
                    num = int(float(vals[0]))
                    if num > 0 and vals[1]:
                        m = re.match(r'^(c\.\d+)\s+(.+)', vals[1], re.IGNORECASE)
                        if m: codigo, nombre = m.group(1), m.group(2).strip()
                        else: codigo, nombre = f"c.{num}", vals[1]
                except: pass
            if codigo:
                ex = db.query(Catedra).filter(Catedra.codigo == codigo).first()
                if ex:
                    if nombre and nombre != ex.nombre: ex.nombre = nombre; actualizadas += 1
                else:
                    db.add(Catedra(codigo=codigo, nombre=nombre or f"Cátedra {codigo}")); creadas += 1
        db.commit(); wb.close()
        return {"creadas": creadas, "actualizadas": actualizadas}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error: {str(e)}")

@app.post("/api/importar/apertura-catedras")
async def importar_apertura_catedras(file: UploadFile = File(...), cuatrimestre_id: int = 1, db: Session = Depends(get_db)):
    try:
        content = await file.read()
        wb = load_workbook(filename=io.BytesIO(content), read_only=True)
        ws = wb[wb.sheetnames[0]]
        abiertas = ya_existentes = 0
        errores = []
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            vals = [str(c).strip() if c is not None else "" for c in row]
            if not any(vals): continue
            codigo = None
            nombre = None
            for v in vals:
                m = re.match(r'^(c\.\d+)\s*(.*)', v, re.IGNORECASE)
                if m: codigo = m.group(1); nombre = m.group(2).strip() if m.group(2) else None; break
            if not codigo and vals[0]:
                try:
                    num = int(float(vals[0]))
                    if num > 0: codigo = f"c.{num}"; nombre = vals[1] if len(vals) > 1 else None
                except: pass
            if not codigo: continue
            catedra = db.query(Catedra).filter(Catedra.codigo == codigo).first()
            if not catedra:
                if nombre: catedra = Catedra(codigo=codigo, nombre=nombre); db.add(catedra); db.flush()
                else: continue
            if db.query(Asignacion).filter(Asignacion.catedra_id == catedra.id, Asignacion.cuatrimestre_id == cuatrimestre_id).first():
                ya_existentes += 1; continue
            db.add(Asignacion(catedra_id=catedra.id, cuatrimestre_id=cuatrimestre_id, modalidad='virtual_tm'))
            abiertas += 1
        db.commit(); wb.close()
        return {"abiertas": abiertas, "ya_existentes": ya_existentes, "errores": errores[:20]}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Error: {str(e)}")

@app.post("/api/importar/cursos")
async def importar_cursos(file: UploadFile = File(...), db: Session = Depends(get_db)):
    try:
        content = await file.read()
        wb = load_workbook(filename=io.BytesIO(content), read_only=True)
        ws = wb[wb.sheetnames[0]]
        creados = omitidos = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            vals = [str(c).strip() if c is not None else "" for c in row]
            sede_texto = vals[0] if len(vals) > 0 else ""
            nombre = vals[1] if len(vals) > 1 else ""
            if not nombre or len(nombre) < 3: continue
            if any(x in nombre.lower() for x in ['no disponible', '//bajas//', 'test ']): omitidos += 1; continue
            sede = db.query(Sede).filter(Sede.nombre == sede_texto).first() if sede_texto else None
            if not sede and sede_texto and len(sede_texto) > 2:
                sede = Sede(nombre=sede_texto, color="bg-gray-500"); db.add(sede); db.flush()
            if not db.query(Curso).filter(Curso.nombre == nombre).first():
                db.add(Curso(nombre=nombre, sede_id=sede.id if sede else None)); creados += 1
        db.commit(); wb.close()
        return {"creados": creados, "omitidos": omitidos}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error: {str(e)}")

@app.post("/api/importar/docentes")
async def importar_docentes(file: UploadFile = File(...), db: Session = Depends(get_db)):
    try:
        content = await file.read()
        wb = load_workbook(filename=io.BytesIO(content), read_only=True)
        ws = wb[wb.sheetnames[0]]
        headers = [str(c.value).lower().strip() if c.value else "" for c in ws[1]]
        col_map = {"dni": -1, "nombre": -1, "apellido": -1, "email": -1}
        for i, h in enumerate(headers):
            if any(x in h for x in ["dni", "documento"]): col_map["dni"] = i
            elif h in ["nombre", "nombres"]: col_map["nombre"] = i
            elif "apellido" in h: col_map["apellido"] = i
            elif any(x in h for x in ["mail", "email", "correo"]): col_map["email"] = i
        es_combinado = any("apellido y nombre" in h or "apellido, nombre" in h for h in headers)
        creados = actualizados = 0; errores = []
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            vals = [str(c).strip() if c is not None else "" for c in row]
            if es_combinado:
                dni = vals[0].replace(".", "").replace("-", "").replace(" ", "") if vals else ""
                if dni.endswith(".0"): dni = dni[:-2]
                parts = (vals[1] if len(vals) > 1 else "").split(",", 1)
                apellido = parts[0].strip().title()
                nombre = parts[1].strip().title() if len(parts) > 1 else ""
                email = vals[2] if len(vals) > 2 else None
            else:
                def gv(k):
                    idx = col_map.get(k, -1)
                    return vals[idx] if 0 <= idx < len(vals) and vals[idx] else None
                dni = (gv("dni") or (vals[0] if vals else "")).replace(".", "").replace("-", "").replace(" ", "")
                if dni.endswith(".0"): dni = dni[:-2]
                nombre = gv("nombre") or ""; apellido = gv("apellido") or ""; email = gv("email")
            if not dni or len(dni) < 7: continue
            ex = db.query(Docente).filter(Docente.dni == dni).first()
            if ex:
                if nombre: ex.nombre = nombre
                if apellido: ex.apellido = apellido
                if email: ex.email = email
                actualizados += 1
            else:
                if not nombre and not apellido: continue
                db.add(Docente(dni=dni, nombre=nombre, apellido=apellido, email=email)); creados += 1
        db.commit(); wb.close()
        return {"creados": creados, "actualizados": actualizados, "errores": errores[:10]}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error: {str(e)}")

@app.post("/api/importar/catedra-cursos")
async def importar_catedra_cursos(file: UploadFile = File(...), db: Session = Depends(get_db)):
    try:
        content = await file.read()
        wb = load_workbook(filename=io.BytesIO(content), read_only=True)
        ws = wb[wb.sheetnames[0]]
        creados = 0; errores = []
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            vals = [str(c).strip() if c is not None else "" for c in row]
            if len(vals) < 3: continue
            materia_texto = vals[1]; curso_nombre = vals[2]; sede_nombre = vals[3] if len(vals) > 3 else ""
            m = re.match(r'^(c\.\d+)\s+(.+?)(?:\s*-\s*(Mañana|Noche|Virtual|Tarde))?\s*$', materia_texto, re.IGNORECASE)
            codigo = m.group(1) if m else (f"c.{vals[0]}" if vals[0] else None)
            turno = m.group(3) if m else None
            if not codigo or not curso_nombre: continue
            catedra = db.query(Catedra).filter(Catedra.codigo == codigo).first()
            if not catedra: continue
            curso = db.query(Curso).filter(Curso.nombre == curso_nombre).first()
            if not curso:
                sede = db.query(Sede).filter(Sede.nombre == sede_nombre).first() if sede_nombre else None
                curso = Curso(nombre=curso_nombre, sede_id=sede.id if sede else None); db.add(curso); db.flush()
            sede = db.query(Sede).filter(Sede.nombre == sede_nombre).first() if sede_nombre else None
            if not db.query(CatedraCurso).filter(CatedraCurso.catedra_id == catedra.id, CatedraCurso.curso_id == curso.id, CatedraCurso.turno == turno).first():
                db.add(CatedraCurso(catedra_id=catedra.id, curso_id=curso.id, turno=turno, sede_id=sede.id if sede else None)); creados += 1
        db.commit(); wb.close()
        return {"creados": creados, "errores": errores[:20]}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error: {str(e)}")

@app.post("/api/importar/links-meet")
async def importar_links_meet(file: UploadFile = File(...), db: Session = Depends(get_db)):
    try:
        content = await file.read()
        wb = load_workbook(filename=io.BytesIO(content), read_only=True)
        ws = wb[wb.sheetnames[0]]
        actualizados = 0; errores = []
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            vals = [str(c).strip() if c is not None else "" for c in row]
            if len(vals) < 2: continue
            codigo = link = None
            for v in vals:
                if re.match(r'^c\.\d+', v, re.IGNORECASE): codigo = v
                elif 'meet.google.com' in v or 'http' in v: link = v
            if not codigo or not link: continue
            cat = db.query(Catedra).filter(Catedra.codigo == codigo).first()
            if cat: cat.link_meet = link; actualizados += 1
        db.commit(); wb.close()
        return {"actualizados": actualizados, "errores": errores[:10]}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error: {str(e)}")

# ===== v5.0: Importar alumnos con clasificación sede/turno/modalidad =====
@app.post("/api/importar/alumnos")
async def importar_alumnos(file: UploadFile = File(...), cuatrimestre_id: int = 1, db: Session = Depends(get_db)):
    from sqlalchemy import text
    try:
        content = await file.read()
        wb = load_workbook(filename=io.BytesIO(content), read_only=True)
        creados = 0; inscripciones = 0; actualizados = 0; errores = []; edi_total = 0
        stats = {'virtual': 0, 'presencial': 0, 'turnos': {}, 'sedes': {}}
        for ws in wb:
            # v16.0: Collect all rows, find dominant code per sheet for EDI matching
            all_rows = list(ws.iter_rows(min_row=2, values_only=True))
            sheet_codes = {}
            for pre_row in all_rows:
                pv = [str(c).strip() if c is not None else "" for c in pre_row]
                if len(pv) < 4: continue
                pm = re.match(r'^(c\.\d+)', pv[3], re.IGNORECASE)
                if pm:
                    pc = pm.group(1)
                    sheet_codes[pc] = sheet_codes.get(pc, 0) + 1
            dominant_code = max(sheet_codes, key=sheet_codes.get) if sheet_codes else None
            edi_count = 0
            for row_num, row in enumerate(all_rows, start=2):
                vals = [str(c).strip() if c is not None else "" for c in row]
                if len(vals) < 4: continue
                alumno_texto = vals[1]
                dni_raw = str(vals[2]).strip() if vals[2] else ""
                materia_texto = vals[3]
                curso_texto = vals[4] if len(vals) > 4 else ""
                dni = re.sub(r'[.\-\s]', '', dni_raw)
                if '.' in dni:
                    try: dni = str(int(float(dni)))
                    except: pass
                if not dni or len(dni) < 6: continue
                m_nombre = re.match(r'^(.+?)\s*\(\d+\)', alumno_texto)
                nombre_completo = m_nombre.group(1).strip() if m_nombre else alumno_texto
                partes = nombre_completo.strip().split(' ')
                nombre = ' '.join(partes[:-1]) if len(partes) >= 2 else nombre_completo
                apellido = partes[-1] if len(partes) >= 2 else ""
                m_cod = re.match(r'^(c\.\d+)', materia_texto, re.IGNORECASE)
                is_edi = False; edi_mat = None
                if not m_cod:
                    # v16.0: If it says EDI, use the dominant cátedra code of this sheet
                    if 'EDI' in materia_texto.upper() and dominant_code:
                        codigo = dominant_code
                        is_edi = True; edi_mat = materia_texto[:100]
                        edi_count += 1
                    else:
                        continue
                else:
                    codigo = m_cod.group(1)
                catedra = db.query(Catedra).filter(Catedra.codigo == codigo).first()
                if not catedra: continue
                # v5.0: Clasificar por curso
                modalidad_alumno, sede_ref, es_cied = clasificar_alumno_curso(curso_texto)
                turno = extraer_turno_materia(materia_texto)
                alumno = db.query(Alumno).filter(Alumno.dni == dni).first()
                if not alumno:
                    alumno = Alumno(dni=dni, nombre=nombre, apellido=apellido)
                    db.add(alumno); db.flush(); creados += 1
                existe = db.query(Inscripcion).filter(
                    Inscripcion.alumno_id == alumno.id,
                    Inscripcion.catedra_id == catedra.id,
                    Inscripcion.cuatrimestre_id == cuatrimestre_id
                ).first()
                if not existe:
                    insc = Inscripcion(alumno_id=alumno.id, catedra_id=catedra.id, cuatrimestre_id=cuatrimestre_id)
                    db.add(insc); db.flush()
                    try:
                        db.execute(text(
                            "UPDATE inscripciones SET turno = :turno, modalidad_alumno = :mod, sede_referencia = :sede, curso_nombre = :curso, es_edi = :edi, edi_materia = :edim WHERE id = :id"
                        ), {"turno": turno, "mod": modalidad_alumno, "sede": sede_ref, "curso": curso_texto[:200] if curso_texto else None, "edi": is_edi, "edim": edi_mat, "id": insc.id})
                    except Exception:
                        pass
                    inscripciones += 1
                else:
                    try:
                        db.execute(text(
                            "UPDATE inscripciones SET turno = :turno, modalidad_alumno = :mod, sede_referencia = :sede, curso_nombre = :curso, es_edi = :edi, edi_materia = :edim WHERE id = :id"
                        ), {"turno": turno, "mod": modalidad_alumno, "sede": sede_ref, "curso": curso_texto[:200] if curso_texto else None, "edi": is_edi, "edim": edi_mat, "id": existe.id})
                        actualizados += 1
                    except Exception:
                        pass
                # Contar stats siempre
                stats[modalidad_alumno] = stats.get(modalidad_alumno, 0) + 1
                if turno: stats['turnos'][turno] = stats['turnos'].get(turno, 0) + 1
                if sede_ref: stats['sedes'][sede_ref] = stats['sedes'].get(sede_ref, 0) + 1
            edi_total += edi_count
        db.commit(); wb.close()
        return {
            "alumnos_nuevos": creados, "inscripciones_nuevas": inscripciones,
            "inscripciones_actualizadas": actualizados,
            "edi_contabilizados": edi_total,
            "virtuales": stats.get('virtual', 0), "presenciales": stats.get('presencial', 0),
            "por_turno": stats['turnos'], "por_sede": stats['sedes'],
            "errores": errores[:20]
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Error: {str(e)}")


# ===== v16.0: EDI Inscripciones — listar alumnos EDI por cátedra =====
@app.get("/api/edi-inscripciones")
def get_edi_inscripciones(cuatrimestre_id: int = None, db: Session = Depends(get_db)):
    from sqlalchemy import text
    q = "SELECT i.id, i.catedra_id, i.edi_materia, i.curso_nombre, i.sede_referencia, i.turno, a.nombre, a.apellido, a.dni, c.codigo, c.nombre as cat_nombre FROM inscripciones i JOIN alumnos a ON i.alumno_id = a.id JOIN catedras c ON i.catedra_id = c.id WHERE i.es_edi = TRUE"
    if cuatrimestre_id: q += f" AND i.cuatrimestre_id = {cuatrimestre_id}"
    q += " ORDER BY c.codigo, a.apellido, a.nombre"
    try: rows = db.execute(text(q)).fetchall()
    except: return {"por_catedra": {}, "total": 0}
    # Group by cátedra
    por_cat = {}
    for r in rows:
        cod = r[9]; cat_nombre = r[10]
        key = f"{cod} ({cat_nombre})"
        if key not in por_cat: por_cat[key] = {"codigo": cod, "nombre": cat_nombre, "alumnos": [], "total": 0}
        por_cat[key]["alumnos"].append({
            "nombre": f"{r[6]} {r[7]}", "dni": r[8],
            "edi_materia": r[2], "curso": (r[3] or '')[:50],
            "sede": r[4], "turno": r[5],
        })
        por_cat[key]["total"] += 1
    return {"por_catedra": por_cat, "total": len(rows)}


# ==================== SOLAPAMIENTOS ====================

def verificar_solapamiento_catedra(catedra_id, dia, hora_inicio, cuatrimestre_id, excluir_id, db):
    q = db.query(Asignacion).filter(Asignacion.catedra_id == catedra_id, Asignacion.dia == dia, Asignacion.hora_inicio == hora_inicio, Asignacion.cuatrimestre_id == cuatrimestre_id, Asignacion.modalidad != 'asincronica')
    if excluir_id: q = q.filter(Asignacion.id != excluir_id)
    if q.first(): return f"⛔ La cátedra ya tiene clase el {dia} a las {hora_inicio}."
    return None

@app.get("/api/horarios/solapamientos")
def get_solapamientos(cuatrimestre_id: int = None, db: Session = Depends(get_db)):
    """v17.0: compara franjas reales (inicio–fin), no sólo la hora de inicio.
    Así detecta que 18:00-19:30 se pisa con 18:30-20:00."""
    try:
        q = db.query(Asignacion).filter(Asignacion.dia.isnot(None), Asignacion.hora_inicio.isnot(None), Asignacion.modalidad != 'asincronica')
        if cuatrimestre_id: q = q.filter(Asignacion.cuatrimestre_id == cuatrimestre_id)
        asigs = q.all()
        solapamientos = []; checked = set()
        for i, a1 in enumerate(asigs):
            for a2 in asigs[i+1:]:
                if a1.dia != a2.dia: continue
                f1 = getattr(a1, 'hora_fin', None); f2 = getattr(a2, 'hora_fin', None)
                if not rangos_se_pisan(a1.hora_inicio, f1, a2.hora_inicio, f2): continue
                pair = tuple(sorted([a1.id, a2.id]))
                if pair in checked: continue
                checked.add(pair)
                cat1 = a1.catedra; cat2 = a2.catedra
                franja1 = f"{a1.hora_inicio}{'-' + f1 if f1 else ''}"
                franja2 = f"{a2.hora_inicio}{'-' + f2 if f2 else ''}"
                detalle = franja1 if franja1 == franja2 else f"{franja1} vs {franja2}"
                if a1.catedra_id == a2.catedra_id:
                    solapamientos.append({"tipo": "CATEDRA", "severidad": "CRITICO",
                        "mensaje": f"Cátedra {cat1.codigo if cat1 else '?'} tiene dos clases {a1.dia} {detalle}.",
                        "dia": a1.dia, "hora": a1.hora_inicio})
                elif a1.docente_id and a1.docente_id == a2.docente_id:
                    doc = a1.docente
                    solapamientos.append({"tipo": "DOCENTE", "severidad": "ALTO",
                        "mensaje": f"{doc.nombre} {doc.apellido} tiene {cat1.codigo if cat1 else '?'} y {cat2.codigo if cat2 else '?'} el {a1.dia} {detalle}.",
                        "dia": a1.dia, "hora": a1.hora_inicio})
        return solapamientos
    except Exception: return []

@app.get("/api/docentes/estadisticas")
def get_estadisticas_docentes(cuatrimestre_id: int = None, db: Session = Depends(get_db)):
    stats = {"presencial_virtual": 0, "sede_virtual": 0, "remoto": 0, "sin_asignaciones": 0}
    for d in db.query(Docente).all():
        tipo = calcular_tipo_modalidad(d, db)
        if tipo == "PRESENCIAL_VIRTUAL": stats["presencial_virtual"] += 1
        elif tipo == "SEDE_VIRTUAL": stats["sede_virtual"] += 1
        elif tipo == "REMOTO": stats["remoto"] += 1
        else: stats["sin_asignaciones"] += 1
    return stats

# ===== v15.0: Importar alumnos BCE/BEA =====
@app.post("/api/importar/alumnos-bce-bea")
async def importar_alumnos_bce_bea(file: UploadFile = File(...), cuatrimestre_id: int = 1, db: Session = Depends(get_db)):
    """v17.2 — Reescrito con lógica simple.

    BCE y BEA son 100% virtuales: no tienen turno, ni día, ni modalidad. Lo único que
    hace falta es CONTAR cuántos alumnos (DNI) hay en cada cátedra.

    El código de cátedra viene en el NOMBRE DEL ARCHIVO (ej: "c_2028_Lengua_I_-_BCE.xlsx"),
    porque dentro del Excel la columna MATERIA sólo trae el nombre ("Lengua I").
    Si esa cátedra todavía no existe en el sistema, se crea automáticamente: antes el
    importador la descartaba en silencio y por eso el contador daba siempre cero.
    """
    from openpyxl import load_workbook
    from sqlalchemy import text
    import io
    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No se pudo abrir el Excel: {str(e)[:150]}")

    nombre_archivo = file.filename or ""
    # 1) Código de cátedra desde el nombre del archivo
    codigo = None
    m = re.search(r'c[_\.\s-]*(\d+)', nombre_archivo, re.IGNORECASE)
    if m: codigo = f"c.{m.group(1)}"

    # 2) Nombre de la materia: primera fila de datos con contenido en la columna MATERIA
    nombre_materia = None
    filas = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=2, values_only=True):
            vals = list(row)
            if len(vals) < 4: continue
            filas.append(vals)
            if not nombre_materia:
                mat = str(vals[3] or '').strip()
                if mat and mat.upper() not in ('MATERIA', 'NONE'):
                    nombre_materia = mat
    wb.close()

    if not codigo and not nombre_materia:
        raise HTTPException(status_code=400,
            detail="No se pudo identificar la cátedra: el archivo debe llamarse como 'c_2028_Nombre.xlsx' o traer la materia en la columna D.")

    # 3) Buscar la cátedra; si no existe, crearla
    cat = None
    creada = False
    if codigo:
        cat = db.query(Catedra).filter(Catedra.codigo == codigo).first()
    if not cat and nombre_materia:
        for c in db.query(Catedra).all():
            if (c.nombre or '').strip().lower() == nombre_materia.strip().lower():
                cat = c; break
    if not cat:
        cat = Catedra(codigo=codigo or f"c.BCE-{abs(hash(nombre_materia)) % 9999}",
                      nombre=nombre_materia or "Cátedra BCE/BEA")
        db.add(cat)
        try:
            db.commit(); db.refresh(cat); creada = True
        except Exception as e:
            db.rollback()
            raise HTTPException(status_code=400, detail=f"No se pudo crear la cátedra: {str(e)[:150]}")

    # 4) Contar alumnos por DNI (una inscripción por DNI, sin turno ni modalidad)
    dnis = {}          # dni -> nombre limpio
    sin_dni = 0
    for vals in filas:
        alumno_nombre = str(vals[1] or '').strip()
        if not alumno_nombre or alumno_nombre.upper() in ('ALUMNO', 'NONE'): continue
        dni = re.sub(r'[^\d]', '', str(vals[2] or ''))[:10]
        if not dni:
            md = re.search(r'\((\d{6,10})\)', alumno_nombre)
            if md: dni = md.group(1)
        if not dni:
            sin_dni += 1
            continue
        if dni not in dnis:
            dnis[dni] = re.sub(r'\s*\(.*\)', '', alumno_nombre).strip()

    duplicados = max(0, len([1 for v in filas
                             if str(v[1] or '').strip() and str(v[1] or '').strip().upper() != 'ALUMNO'])
                     - len(dnis) - sin_dni)

    # 5) Crear alumnos e inscripciones
    nuevos_alumnos = 0; nuevas_insc = 0; ya_estaban = 0; errores = []
    for dni, nombre_alumno in dnis.items():
        try:
            al = db.query(Alumno).filter(Alumno.dni == dni).first()
            if not al:
                al = Alumno(nombre=nombre_alumno, dni=dni)
                db.add(al); db.flush(); nuevos_alumnos += 1
            insc = db.query(Inscripcion).filter(
                Inscripcion.alumno_id == al.id,
                Inscripcion.catedra_id == cat.id,
                Inscripcion.cuatrimestre_id == cuatrimestre_id).first()
            if insc:
                ya_estaban += 1
            else:
                insc = Inscripcion(alumno_id=al.id, catedra_id=cat.id,
                                   cuatrimestre_id=cuatrimestre_id)
                db.add(insc); db.flush(); nuevas_insc += 1
            # Marcarlas como virtuales (columnas fuera del ORM -> SQL directo)
            db.execute(text(
                "UPDATE inscripciones SET turno = 'Virtual', modalidad_alumno = 'virtual', "
                "sede_referencia = 'Online - Interior', curso_nombre = :curso WHERE id = :id"
            ), {"curso": "BCE/BEA", "id": insc.id})
        except Exception as e:
            db.rollback()
            errores.append(f"DNI {dni}: {str(e)[:80]}")
    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Error guardando: {str(e)[:200]}")

    # 6) Total real en la base para esta cátedra
    try:
        total_en_catedra = db.execute(text(
            "SELECT COUNT(*) FROM inscripciones WHERE catedra_id = :c AND cuatrimestre_id = :cu"
        ), {"c": cat.id, "cu": cuatrimestre_id}).scalar() or 0
    except Exception:
        db.rollback(); total_en_catedra = nuevas_insc + ya_estaban

    return {
        "catedra": f"{cat.codigo} — {cat.nombre}",
        "catedra_creada": creada,
        "inscriptos_en_el_archivo": len(dnis),
        "inscriptos_totales_en_la_catedra": total_en_catedra,
        "alumnos_nuevos": nuevos_alumnos,
        "inscripciones_nuevas": nuevas_insc,
        "ya_estaban_inscriptos": ya_estaban,
        "filas_duplicadas_omitidas": duplicados,
        "filas_sin_dni": sin_dni,
        "errores": errores[:10],
    }

@app.post("/api/cuatrimestres/replicar")
def replicar_cuatrimestre(data: dict, db: Session = Depends(get_db)):
    origen_id = data.get("origen_id"); destino_id = data.get("destino_id")
    if not origen_id or not destino_id: raise HTTPException(status_code=400, detail="Faltan datos")
    asigs = db.query(Asignacion).filter(Asignacion.cuatrimestre_id == origen_id).all()
    if not asigs: raise HTTPException(status_code=400, detail="Origen sin asignaciones")
    replicadas = ya_existentes = 0
    for a in asigs:
        if db.query(Asignacion).filter(Asignacion.catedra_id == a.catedra_id, Asignacion.cuatrimestre_id == destino_id, Asignacion.modalidad == a.modalidad).first():
            ya_existentes += 1; continue
        db.add(Asignacion(catedra_id=a.catedra_id, cuatrimestre_id=destino_id, modalidad=a.modalidad, dia=a.dia, hora_inicio=a.hora_inicio, hora_fin=a.hora_fin, sede_id=a.sede_id, recibe_alumnos_presenciales=a.recibe_alumnos_presenciales))
        replicadas += 1
    db.commit()
    return {"replicadas": replicadas, "ya_existentes": ya_existentes}

# ===== v11.0: Dashboard con flujo guiado =====
@app.get("/api/dashboard")
def get_dashboard(cuatrimestre_id: int = None, db: Session = Depends(get_db)):
    from sqlalchemy import text
    total_cats = db.query(Catedra).count()
    total_docs = db.query(Docente).count()
    total_cursos = db.query(Curso).count()
    # Asignaciones
    q_a = db.query(Asignacion)
    if cuatrimestre_id: q_a = q_a.filter(Asignacion.cuatrimestre_id == cuatrimestre_id)
    asigs = q_a.all()
    cats_abiertas = len(set(a.catedra_id for a in asigs))
    con_docente = len([a for a in asigs if a.docente_id])
    sin_docente = len([a for a in asigs if not a.docente_id])
    # Inscriptos
    q_i = "SELECT COUNT(*) FROM inscripciones"
    if cuatrimestre_id: q_i += f" WHERE cuatrimestre_id = {cuatrimestre_id}"
    total_insc = db.execute(text(q_i)).scalar() or 0
    # Inscriptos clasificados vs sin clasificar
    q_clas = "SELECT COUNT(*) FROM inscripciones WHERE modalidad_alumno IS NOT NULL"
    if cuatrimestre_id: q_clas += f" AND cuatrimestre_id = {cuatrimestre_id}"
    clasificados = db.execute(text(q_clas)).scalar() or 0
    sin_clasificar = total_insc - clasificados
    # Decisiones tomadas
    q_dec = "SELECT COUNT(*) FROM catedras WHERE decision_apertura IS NOT NULL AND decision_apertura != ''"
    decisiones_tomadas = db.execute(text(q_dec)).scalar() or 0
    cats_con_inscriptos_q = "SELECT COUNT(DISTINCT catedra_id) FROM inscripciones"
    if cuatrimestre_id: cats_con_inscriptos_q += f" WHERE cuatrimestre_id = {cuatrimestre_id}"
    cats_con_inscriptos = db.execute(text(cats_con_inscriptos_q)).scalar() or 0
    decisiones_pendientes = max(0, cats_con_inscriptos - decisiones_tomadas)
    # Disponibilidad docentes
    docs_con_dispo = db.execute(text("SELECT COUNT(DISTINCT docente_id) FROM docente_disponibilidad")).scalar() or 0
    # Solapamientos
    solaps = len(get_solapamientos(cuatrimestre_id, db))
    # Criterio
    criterio = get_criterio_apertura(cuatrimestre_id, db)
    # Cobertura
    cobertura = round((con_docente / max(1, con_docente + sin_docente)) * 100) if asigs else 0
    docs_con_asig = len(set(a.docente_id for a in asigs if a.docente_id))
    # v17.0: estado del dictado
    dictado = _mapa_dictado(db, cuatrimestre_id) if cuatrimestre_id else {}
    se_dictan = len([1 for v in dictado.values() if v.get("se_dicta")])
    cats_con_horario = len(set(a.catedra_id for a in asigs if a.dia and a.hora_inicio))
    cats_abiertas_doc = len(set(a.catedra_id for a in asigs if a.docente_id))
    cats_asincronicas = max(0, se_dictan - cats_abiertas_doc)
    # v17.0: Pasos alineados al flujo real de trabajo del equipo
    pasos = [
        {"num": 1, "titulo": "Marcar qué cátedras se dictan",
         "desc": "Selector masivo. Incluye las asincrónicas: se dictan aunque no tengan docente.",
         "completo": se_dictan > 0,
         "parcial": False,
         "detalle": (f"{se_dictan} cátedras marcadas para dictarse"
                     if se_dictan else "Todavía no se marcó ninguna cátedra"),
         "seccion": "dictado"},
        {"num": 2, "titulo": "Importar alumnos inscriptos",
         "desc": "Para saber cuántos alumnos tiene cada cátedra.",
         "completo": total_insc > 0 and sin_clasificar == 0,
         "parcial": total_insc > 0 and sin_clasificar > 0,
         "detalle": f"{total_insc} inscripciones ({clasificados} clasificadas" + (f", {sin_clasificar} sin clasificar)" if sin_clasificar > 0 else ")"),
         "seccion": "importar"},
        {"num": 3, "titulo": "Exportar la planilla de trabajo",
         "desc": "Excel con inscriptos por cátedra. El equipo asigna docentes y horarios ahí.",
         "completo": False, "parcial": False,
         "detalle": (f"Listo para exportar: {se_dictan} cátedras con sus inscriptos"
                     if se_dictan and total_insc else "Completá los pasos 1 y 2 primero"),
         "seccion": "exportar"},
        {"num": 4, "titulo": "Reimportar la planilla completada",
         "desc": "El sistema detecta solapamientos y deja todo asignado.",
         "completo": cats_con_horario > 0 and solaps == 0,
         "parcial": cats_con_horario > 0,
         "detalle": (f"{cats_con_horario} cátedras con horario cargado"
                     + (f" · {solaps} solapamientos a revisar" if solaps else " · sin solapamientos")
                     if cats_con_horario else "Todavía no se importó la planilla"),
         "seccion": "importar"},
        {"num": 5, "titulo": "Revisar y distribuir",
         "desc": "Verificar solapamientos, completar faltantes y exportar el archivo final.",
         "completo": cats_con_horario > 0 and solaps == 0 and sin_docente == 0,
         "parcial": cats_con_horario > 0,
         "detalle": (f"{cats_abiertas_doc} abiertas con docente · {cats_asincronicas} asincrónicas"
                     + (f" · {sin_docente} sin resolver" if sin_docente else "")),
         "seccion": "exportar"},
    ]
    return {
        "cobertura_pct": cobertura, "pasos": pasos,
        "se_dictan": se_dictan, "catedras_abiertas_docente": cats_abiertas_doc,
        "catedras_asincronicas": cats_asincronicas,
        "total_catedras": total_cats, "catedras_abiertas": cats_abiertas,
        "con_docente": con_docente, "sin_docente": sin_docente,
        "total_inscripciones": total_insc, "solapamientos": solaps,
        "abrir": criterio['stats']['total_abrir'],
        "asincronicas": criterio['stats']['total_asincronica'],
        "sin_alumnos": criterio['stats']['total_sin_alumnos'],
        "docs_sugeridos": criterio['stats']['total_docentes_sugeridos'],
        "total_docentes": total_docs, "docentes_con_asignacion": docs_con_asig,
        "total_asignaciones": len(asigs),
        "decisiones_tomadas": decisiones_tomadas, "decisiones_pendientes": decisiones_pendientes,
    }


# ===== v13.0: Plan Carrera - Importar molde de horarios =====
@app.post("/api/importar/plan-carrera")
async def importar_plan_carrera(file: UploadFile = File(...), db: Session = Depends(get_db)):
    from openpyxl import load_workbook
    from sqlalchemy import text
    import io
    content = await file.read()
    wb = load_workbook(io.BytesIO(content), read_only=True)
    total = 0
    # Clear existing plan
    try: db.execute(text("DELETE FROM plan_carrera")); db.commit()
    except: db.rollback()
    # PHASE 1: Collect all records first, then deduplicate
    all_records = []  # [(sede, carrera, anno, codigo, nombre, dtm, htm, dtn, htn)]
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sede = sheet_name.strip()
        raw_rows = [list(r) for r in ws.iter_rows(values_only=True)]
        carrera_act = ''
        current_anno = ''
        pending_cats = []  # cats before first año label
        expect_reset = False  # set after Practica Formativa
        edi_counter = {}  # (sede, carrera, anno) → count
        def flush_pending(anno_to_use):
            nonlocal pending_cats
            for pc in pending_cats:
                all_records.append((sede, carrera_act, anno_to_use, pc[0], pc[1], pc[2], pc[3], pc[4], pc[5]))
            pending_cats = []
        for vals in raw_rows:
            if len(vals) < 5: continue
            b = str(vals[1] or '').strip()
            c = str(vals[2] or '').strip()
            d = str(vals[3] or '').strip()
            e = str(vals[4] or '').strip()
            # Detect carrera in columns B, C, or D
            for txt in [b, c, d]:
                t_up = txt.upper()
                if ('TECNICO' in t_up or 'TECNICATURA' in t_up) and len(txt) > 15:
                    flush_pending(current_anno)
                    carrera_act = txt.strip()
                    current_anno = ''
            c_up = c.upper()
            e_up = e.upper()
            if 'INSCRIPCION' in c_up or 'INSCRIPCIÓN' in c_up:
                current_anno = ''
                continue
            # "Practica Formativa" without code: set flag to reset AFTER next Profesionalizante (if any)
            if 'PRACTICA FORMATIVA' in e_up:
                expect_reset = True
                continue
            # "CARRERA / AÑO / CODIGO" header row = start of new carrera block
            if d.upper().strip() == 'CODIGO' and e_up.strip() == 'MATERIA':
                current_anno = ''
                continue
            if ('1ER' in c_up or '2DO' in c_up or '3ER' in c_up or '4TO' in c_up) and 'AÑO' in c_up:
                new_anno = c.strip()
                flush_pending(new_anno)
                current_anno = new_anno
            # Extract horarios
            dia_tm = str(vals[6] or '').strip() if len(vals) > 6 else ''
            hora_tm = str(vals[7] or '').strip() if len(vals) > 7 else ''
            dia_tn = str(vals[9] or '').strip() if len(vals) > 9 else ''
            hora_tn = str(vals[10] or '').strip() if len(vals) > 10 else ''
            # Detect catedra code
            try:
                cod_num = int(float(d))
                if cod_num > 0 and e and carrera_act:
                    cod = f'c.{cod_num}'
                    e_up_check = e.upper()
                    is_prof = 'PROFESIONALIZANTE' in e_up_check
                    # If we saw "Practica Formativa" and this is NOT a Profesionalizante → reset year first
                    if expect_reset and not is_prof and current_anno:
                        current_anno = ''
                        expect_reset = False
                    if current_anno:
                        all_records.append((sede, carrera_act, current_anno, cod, e.strip(), dia_tm, hora_tm, dia_tn, hora_tn))
                        # Práctica Profesionalizante = last item of a year block → reset
                        if is_prof:
                            current_anno = ''
                            expect_reset = False
                    else:
                        pending_cats.append((cod, e.strip(), dia_tm, hora_tm, dia_tn, hora_tn))
            except:
                # EDI detection: no code, but "EDI" in column E
                if e.strip().upper() == 'EDI' and carrera_act and current_anno:
                    edi_key = (sede, carrera_act)  # per carrera, NOT per anno
                    edi_count = edi_counter.get(edi_key, 0) + 1
                    edi_counter[edi_key] = edi_count
                    all_records.append((sede, carrera_act, current_anno, f'EDI-{edi_count}', f'EDI {edi_count} (Espacio de Definición Institucional)', '', '', '', ''))
        flush_pending(current_anno or 'AÑO')
    # PHASE 2: Deduplicate — one code per carrera per sede
    seen = set()
    unique_records = []
    for rec in all_records:
        key = (rec[0], rec[1], rec[3])  # (sede, carrera, codigo)
        if key in seen: continue
        seen.add(key)
        unique_records.append(rec)
    # PHASE 3: Insert
    for rec in unique_records:
        try:
            db.execute(text("""INSERT INTO plan_carrera (sede,carrera,anno,codigo_catedra,nombre_catedra,dia_tm,hora_tm,dia_tn,hora_tn)
                VALUES (:s,:ca,:an,:co,:no,:dtm,:htm,:dtn,:htn)"""),
                {"s":rec[0],"ca":rec[1],"an":rec[2],"co":rec[3],"no":rec[4],"dtm":rec[5],"htm":rec[6],"dtn":rec[7],"htn":rec[8]})
            total += 1
        except: pass
    db.commit()
    wb.close()
    return {"importados": total, "hojas": wb.sheetnames}

# ===== v15.0: Importar docentes desde archivo CUIT =====
@app.post("/api/importar/docentes-cuit")
async def importar_docentes_cuit(file: UploadFile = File(...), db: Session = Depends(get_db)):
    from openpyxl import load_workbook
    import io
    content = await file.read()
    wb = load_workbook(io.BytesIO(content), read_only=True)
    nuevos = 0; existentes = 0; errores = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            vals = list(row)
            if len(vals) < 2: continue
            cuit = str(vals[0] or '').strip()
            nombre_completo = str(vals[1] or '').strip()
            if not nombre_completo or nombre_completo == 'None' or ',' not in nombre_completo: continue
            # Extract DNI from CUIT
            dni_digits = cuit.replace('-', '')
            dni = dni_digits[2:-1] if len(dni_digits) >= 10 else dni_digits
            # Parse "APELLIDO, NOMBRE"
            parts = nombre_completo.split(',', 1)
            apellido = parts[0].strip()
            nombre = parts[1].strip() if len(parts) > 1 else ''
            # Check if exists
            existing = db.query(Docente).filter(Docente.dni == dni).first()
            if existing:
                # Update name if needed
                if not existing.nombre or existing.nombre != nombre: existing.nombre = nombre
                if not existing.apellido or existing.apellido != apellido: existing.apellido = apellido
                existentes += 1
            else:
                doc = Docente(nombre=nombre, apellido=apellido, dni=dni)
                db.add(doc)
                nuevos += 1
    db.commit()
    wb.close()
    return {"nuevos": nuevos, "actualizados": existentes}

# ===== v16.0: Auto-asignar cátedras de referencia desde asignaciones actuales =====
@app.post("/api/docentes/auto-referencia")
def auto_referencia_docentes(cuatrimestre_id: int = 1, db: Session = Depends(get_db)):
    from sqlalchemy import text
    docentes = db.query(Docente).all()
    actualizados = 0
    for d in docentes:
        asigs = [a for a in (d.asignaciones or []) if a.cuatrimestre_id == cuatrimestre_id and a.catedra]
        if not asigs: continue
        codes = sorted(set(a.catedra.codigo for a in asigs))
        # Merge with existing refs
        existing_refs = set((getattr(d, 'catedras_referencia', '') or '').split(','))
        existing_refs = {r.strip() for r in existing_refs if r.strip()}
        all_refs = sorted(existing_refs | set(codes))
        new_val = ', '.join(all_refs)
        try:
            db.execute(text("UPDATE docentes SET catedras_referencia = :val WHERE id = :id"),
                {"val": new_val, "id": d.id})
            actualizados += 1
        except: pass
    db.commit()
    return {"actualizados": actualizados}

# ===== v16.0: Importar cátedras de referencia desde Excel de designaciones =====
@app.post("/api/importar/catedras-referencia")
async def importar_catedras_referencia(file: UploadFile = File(...), db: Session = Depends(get_db)):
    from openpyxl import load_workbook
    import io
    content = await file.read()
    wb = load_workbook(io.BytesIO(content), read_only=True)
    # Build docente_name → set of codes
    doc_cats = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            vals = list(row)
            if len(vals) < 6: continue
            try: cn = int(float(str(vals[0] or '')))
            except: continue
            if cn <= 0: continue
            cod = f'c.{cn}'
            doc_raw = str(vals[5] or '').strip()
            if not doc_raw or doc_raw.lower().startswith('ver '): continue
            dc = doc_raw.upper().strip()
            dc = DOCENTE_TYPO_MAP.get(dc, dc)
            if dc not in doc_cats: doc_cats[dc] = set()
            doc_cats[dc].add(cod)
    wb.close()
    # Match to docentes in DB
    from sqlalchemy import text
    all_docs = db.query(Docente).all()
    doc_by_apellido = {}
    for d in all_docs:
        ap = (d.apellido or '').upper().strip()
        if ap: doc_by_apellido[ap] = d
        full = f"{(d.apellido or '')} {(d.nombre or '')}".upper().strip()
        if full: doc_by_apellido[full] = d
        full2 = f"{(d.nombre or '')} {(d.apellido or '')}".upper().strip()
        if full2: doc_by_apellido[full2] = d
    actualizados = 0; no_match = []
    for doc_name, codes in doc_cats.items():
        # Find docente
        docente = doc_by_apellido.get(doc_name)
        if not docente:
            for key, d in doc_by_apellido.items():
                if doc_name in key or key in doc_name:
                    docente = d; break
        if not docente:
            no_match.append(doc_name); continue
        # Merge with existing
        existing = set((getattr(docente, 'catedras_referencia', '') or '').split(','))
        existing = {r.strip() for r in existing if r.strip()}
        merged = sorted(existing | codes)
        try:
            db.execute(text("UPDATE docentes SET catedras_referencia = :val WHERE id = :id"),
                {"val": ', '.join(merged), "id": docente.id})
            actualizados += 1
        except: pass
    db.commit()
    return {"actualizados": actualizados, "no_encontrados": no_match}

# ===== v15.0: Typo correction map for docente names =====
DOCENTE_TYPO_MAP = {
    'CAPUCCHETI': 'CAPUCHETTI', 'DATRI': "D'ATRI", "D´ATRI": "D'ATRI",
    'TOMATI': 'TOMATTI', 'MARIA LAURA PAVEL': 'PAVEL',
    'YANELA CAPUCHETTI': 'CAPUCHETTI', 'YANELA CAPUCCHETI': 'CAPUCHETTI',
    'DIEGO MARTINEZ': 'MARTINEZ', 'PEREZ LUCAS': 'PEREZ',
    'ROSCHMAN': 'GONZALEZ LARES ROSCHMAN',
    'GONZALEZ R': 'GONZALEZ', 'GASTON GONZALEZ': 'GONZALEZ',
    'GONZALEZ ARIEL': 'GONZALEZ', 'FERNANDEZ L': 'FERNANDEZ',
    'RODRIGUEZ S': 'RODRIGUEZ', 'HERRERA S.': 'HERRERA', 'HERRERA S': 'HERRERA',
    'LOPEZ G': 'LOPEZ GHIGLIERI', 'AGUSTINA ACOSTA': 'ACOSTA',
    'ACOSTA AGUSTINA': 'ACOSTA',
    # NOT mapping: PALERMO ≠ PALMERO LLANOS (different people)
    # NOT mapping: KAREN PAMELA FLORENTIN → goes to Caren Pamela, not Isaul
}

def _parse_horarios_excel(file_content, db, cuatrimestre_id):
    """Parse horarios Excel and return structured data without applying changes."""
    from openpyxl import load_workbook
    import io
    wb = load_workbook(io.BytesIO(file_content))
    all_cats = {c.codigo: c for c in db.query(Catedra).all()}
    all_docs = db.query(Docente).all()
    doc_by_apellido = {}
    for d in all_docs:
        ap = (d.apellido or '').upper().strip()
        if ap: doc_by_apellido[ap] = d
        full = f"{(d.apellido or '')} {(d.nombre or '')}".upper().strip()
        if full: doc_by_apellido[full] = d
        full2 = f"{(d.nombre or '')} {(d.apellido or '')}".upper().strip()
        if full2: doc_by_apellido[full2] = d
        # Also match by full name as typed (e.g. "Luciano Salinas")
        full3 = f"{(d.nombre or '')} {(d.apellido or '')}".strip()
        if full3: doc_by_apellido[full3.upper()] = d
    all_sedes = {s.nombre: s for s in db.query(Sede).all()}
    dia_map = {'LUNES':'Lunes','MARTES':'Martes','MIERCOLES':'Miércoles','MIÉRCOLES':'Miércoles',
        'JUEVES':'Jueves','VIERNES':'Viernes','SABADO':'Sábado','SÁBADO':'Sábado'}
    def _limpiar_hora(h):
        h = str(h or '').replace('.', ':').replace(' HS', '').replace(' hs', '').replace('HS', '').strip()
        if h and ':' in h:
            partes = h.split(':')
            try: h = f"{int(partes[0]):02d}:{partes[1].strip()[:2]}"
            except Exception: pass
        return h

    def _parece_hora(v):
        return bool(re.match(r'^\s*\d{1,2}[:.]\d{2}', str(v or '').strip()))

    results = []; no_cat = []; no_doc = set(); doc_to_create = set()
    for ws in wb.worksheets:
        if ws.title.strip().lower() in ('instructivo', 'instrucciones'): continue
        filas = list(ws.iter_rows(values_only=True))
        # v17.0: autodetección de formato.
        # Formato CLÁSICO:  A=cod B=materia C=dia D=hora E=sede F=docente G=meet
        # Formato PLANILLA: A=cod B=materia C=dia D=hora_ini E=hora_fin F=sede G=docente H=meet
        formato_planilla = False
        for f in filas[:6]:
            textos = [str(x or '').upper().strip() for x in list(f)[:9]]
            if any('HORA FIN' in t for t in textos):
                formato_planilla = True; break
        if not formato_planilla:
            # Heurística: si la columna E parece una hora en varias filas de datos, es planilla
            hits = 0; muestras = 0
            for f in filas:
                v = list(f)
                if len(v) < 6: continue
                try:
                    if int(float(str(v[0] or ''))) <= 0: continue
                except Exception: continue
                muestras += 1
                if _parece_hora(v[4]): hits += 1
                if muestras >= 12: break
            if muestras and hits >= max(2, muestras // 2): formato_planilla = True

        for row in filas:
            vals = list(row)
            if len(vals) < 5: continue
            try: cod_num = int(float(str(vals[0] or '')))
            except: continue
            if cod_num <= 0: continue
            codigo = f'c.{cod_num}'
            materia = str(vals[1] or '').strip()
            dia_raw = str(vals[2] or '').strip()
            hora_raw = str(vals[3] or '').strip()
            if formato_planilla:
                hora_fin_raw = str(vals[4] or '').strip() if len(vals) > 4 else ''
                sede_raw = str(vals[5] or '').strip() if len(vals) > 5 else ''
                doc_raw = str(vals[6] or '').strip() if len(vals) > 6 else ''
                meet_link = str(vals[7] or '').strip() if len(vals) > 7 else ''
            else:
                hora_fin_raw = ''
                sede_raw = str(vals[4] or '').strip()
                doc_raw = str(vals[5] or '').strip() if len(vals) > 5 else ''
                meet_link = str(vals[6] or '').strip() if len(vals) > 6 else ''
            if doc_raw.lower() in ('none', 'nan'): doc_raw = ''
            hora_fin = _limpiar_hora(hora_fin_raw) if hora_fin_raw else ''
            if not dia_raw or not hora_raw: continue
            dia = dia_map.get(dia_raw.upper().strip(), dia_raw.strip().title())
            hora = _limpiar_hora(hora_raw)
            cat = all_cats.get(codigo)
            if not cat: no_cat.append(f"{codigo} {materia}"); continue
            sede_nombre = normalizar_sede(sede_raw) or sede_raw or ''
            sede_obj = None
            def _strip_accents(s):
                import unicodedata
                return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
            for sn, so in all_sedes.items():
                if _strip_accents(sn.lower()).replace(' ','') == _strip_accents(sede_nombre.lower()).replace(' ',''): sede_obj = so; break
            if not sede_obj:
                for sn, so in all_sedes.items():
                    if _strip_accents(sede_nombre.lower())[:4] in _strip_accents(sn.lower()): sede_obj = so; break
            docente_obj = None; doc_display = ''
            if doc_raw and not doc_raw.lower().startswith('ver '):
                doc_clean = doc_raw.upper().strip()
                if doc_clean.startswith('VER '): doc_clean = doc_clean[4:].strip()
                # Apply typo map
                corrected = DOCENTE_TYPO_MAP.get(doc_clean, doc_clean)
                docente_obj = doc_by_apellido.get(corrected)
                if not docente_obj:
                    for key, d in doc_by_apellido.items():
                        if corrected in key or key in corrected: docente_obj = d; break
                if docente_obj:
                    doc_display = f"{docente_obj.nombre} {docente_obj.apellido}"
                else:
                    doc_display = doc_raw
                    doc_to_create.add(doc_raw)
            modalidad = 'remoto' if sede_nombre in ['Online - Interior', 'ONLINE'] else 'presencial_virtual'
            results.append({
                'cat_id': cat.id, 'cat_codigo': codigo, 'cat_nombre': cat.nombre,
                'dia': dia, 'hora': hora, 'hora_fin': hora_fin,
                'sede_id': sede_obj.id if sede_obj else None,
                'sede_nombre': sede_nombre, 'docente_id': docente_obj.id if docente_obj else None,
                'docente_display': doc_display, 'modalidad': modalidad,
                'doc_raw': doc_raw, 'meet_link': meet_link if meet_link.startswith('http') else '',
            })
    wb.close()
    return results, list(set(no_cat)), sorted(list(doc_to_create))

# ===== v15.0: Preview horarios import =====
@app.post("/api/importar/horarios-preview")
async def horarios_preview(file: UploadFile = File(...), cuatrimestre_id: int = 1, db: Session = Depends(get_db)):
    try:
        content = await file.read()
        results, no_cat, doc_to_create = _parse_horarios_excel(content, db, cuatrimestre_id)
        current_count = db.query(Asignacion).filter(Asignacion.cuatrimestre_id == cuatrimestre_id).count()
        con_doc = len([r for r in results if r['docente_id']])
        sin_doc = len([r for r in results if not r['docente_id'] and not r['doc_raw']])
        doc_new = len([r for r in results if not r['docente_id'] and r['doc_raw']])
        con_meet = len([r for r in results if r.get('meet_link')])
        return {
            "asignaciones_actuales_a_borrar": current_count,
            "asignaciones_nuevas": len(results),
            "con_docente_existente": con_doc,
            "sin_docente": sin_doc,
            "con_docente_nuevo_a_crear": doc_new,
            "links_meet": con_meet,
            "docentes_a_crear": doc_to_create,
            "catedras_no_encontradas": no_cat[:20],
            "_debug": {"total_catedras_db": len(db.query(Catedra).all()), "total_docentes_db": len(db.query(Docente).all()), "no_cat_count": len(no_cat)},
            "preview": [{"cat": r['cat_codigo'], "nombre": r['cat_nombre'][:30], "dia": r['dia'],
                "hora": r['hora'], "sede": r['sede_nombre'], "docente": r['docente_display'] or '—',
                "estado": "✅" if r['docente_id'] else ("🆕 Crear" if r['doc_raw'] else "—")} for r in results[:50]],
        }
    except Exception as e:
        import traceback
        return {"error": str(e), "traceback": traceback.format_exc()[-500:],
            "asignaciones_actuales_a_borrar": 0, "asignaciones_nuevas": 0, "con_docente_existente": 0,
            "docentes_a_crear": [], "catedras_no_encontradas": [], "preview": []}

# ===== v15.0: Apply horarios import (after preview) =====
@app.post("/api/importar/horarios-aplicar")
async def horarios_aplicar(file: UploadFile = File(...), cuatrimestre_id: int = 1, db: Session = Depends(get_db)):
    content = await file.read()
    results, no_cat, doc_to_create = _parse_horarios_excel(content, db, cuatrimestre_id)
    # 1) Create missing docentes
    nuevos_docs = 0
    doc_created_map = {}
    for doc_name in doc_to_create:
        parts = doc_name.strip().split(' ', 1)
        if len(parts) == 2:
            apellido = parts[0].strip().title()
            nombre = parts[1].strip().title()
        else:
            apellido = parts[0].strip().title()
            nombre = ''
        # v17.0: el DNI ya es opcional, se crea sin placeholder
        new_doc = Docente(nombre=nombre, apellido=apellido, dni=None)
        db.add(new_doc); db.flush()
        doc_created_map[doc_name.upper()] = new_doc
        nuevos_docs += 1
    # 2) Delete existing asignaciones for this cuatrimestre
    try:
        deleted = db.query(Asignacion).filter(Asignacion.cuatrimestre_id == cuatrimestre_id).delete()
        db.flush()
    except Exception as e:
        db.rollback()
        return {"error": f"Error borrando asignaciones: {str(e)}"}
    # 3) Create new asignaciones + update meet links
    creados = 0; meet_updated = 0
    from sqlalchemy import text as sql_text
    for r in results:
        doc_id = r['docente_id']
        if not doc_id and r['doc_raw']:
            created = doc_created_map.get(r['doc_raw'].upper())
            if created: doc_id = created.id
        asig = Asignacion(
            catedra_id=r['cat_id'], docente_id=doc_id,
            cuatrimestre_id=cuatrimestre_id, dia=r['dia'], hora_inicio=r['hora'],
            sede_id=r['sede_id'], modalidad=r['modalidad'])
        db.add(asig)
        db.flush()
        # v17.0: hora_fin va por SQL (columna creada por migración, no está en el modelo)
        if r.get('hora_fin'):
            try:
                db.execute(sql_text("UPDATE asignaciones SET hora_fin = :hf WHERE id = :id"),
                           {"hf": r['hora_fin'], "id": asig.id})
            except Exception: pass
        creados += 1
        # v16.0: Update meet link on cátedra if provided
        if r.get('meet_link'):
            try:
                db.execute(sql_text("UPDATE catedras SET link_meet = :link WHERE id = :id"),
                    {"link": r['meet_link'], "id": r['cat_id']})
                meet_updated += 1
            except: pass
    try:
        db.commit()
    except Exception as e:
        db.rollback()
        return {"error": f"Error guardando: {str(e)[:200]}"}

    # v17.0: al importar la planilla ya queda todo resuelto.
    # Toda cátedra que aparece en el archivo SE DICTA; si trae docente además queda ABIERTA,
    # y si no trae docente queda como ASINCRÓNICA (video pregrabado). Ya no queda "pendiente".
    cats_en_archivo = {r['cat_id'] for r in results}
    cats_con_docente = {r['cat_id'] for r in results if r['docente_id'] or r['doc_raw']}
    dictadas = 0; decididas = 0
    for cat_id in cats_en_archivo:
        try:
            db.execute(sql_text("""
                INSERT INTO catedra_dictado (catedra_id, cuatrimestre_id, se_dicta)
                VALUES (:cat, :cuat, TRUE)
                ON CONFLICT (catedra_id, cuatrimestre_id) DO UPDATE SET se_dicta = TRUE
            """), {"cat": cat_id, "cuat": cuatrimestre_id})
            dictadas += 1
        except Exception:
            db.rollback()
        decision = 'ABRIR' if cat_id in cats_con_docente else 'ASINCRONICA'
        if sql_set(db, "catedras", "decision_apertura", decision, cat_id):
            decididas += 1
    try: db.commit()
    except Exception: db.rollback()

    return {
        "asignaciones_borradas": deleted,
        "asignaciones_creadas": creados,
        "links_meet_actualizados": meet_updated,
        "catedras_marcadas_se_dicta": dictadas,
        "decisiones_resueltas": decididas,
        "catedras_abiertas": len(cats_con_docente),
        "catedras_asincronicas": len(cats_en_archivo - cats_con_docente),
        "docentes_creados": nuevos_docs,
        "docentes_nuevos": doc_to_create,
        "catedras_no_encontradas": no_cat[:20],
    }

# ===== v13.0: Sugerencias de horarios cruzando plan + inscriptos =====
@app.get("/api/plan-carrera/sugerencias")
def get_sugerencias_plan(cuatrimestre_id: int = None, sede: str = None, db: Session = Depends(get_db)):
    from sqlalchemy import text
    # Get plan
    q = "SELECT * FROM plan_carrera"
    filters = []
    if sede: filters.append(f"sede = '{sede}'")
    if filters: q += " WHERE " + " AND ".join(filters)
    q += " ORDER BY sede, carrera, anno, codigo_catedra"
    try: rows = db.execute(text(q)).fetchall()
    except: return {"sedes": [], "plan_importado": False}
    if not rows: return {"sedes": [], "plan_importado": False}
    # Get inscriptos totales
    total_q = "SELECT catedra_id, COUNT(*) FROM inscripciones"
    if cuatrimestre_id: total_q += f" WHERE cuatrimestre_id = {cuatrimestre_id}"
    total_q += " GROUP BY catedra_id"
    total_map = {}
    try:
        for r in db.execute(text(total_q)).fetchall(): total_map[r[0]] = r[1]
    except: pass
    # Get catedra code → id mapping
    cat_map = {}
    for c in db.query(Catedra).all():
        cat_map[c.codigo] = {"id": c.id, "nombre": c.nombre}
    # Get current asignaciones
    asig_q = db.query(Asignacion)
    if cuatrimestre_id: asig_q = asig_q.filter(Asignacion.cuatrimestre_id == cuatrimestre_id)
    asigs = asig_q.all()
    asig_map = {}  # catedra_id → [{docente, dia, hora, sede}]
    for a in asigs:
        if a.catedra_id not in asig_map: asig_map[a.catedra_id] = []
        asig_map[a.catedra_id].append({
            "docente": f"{a.docente.nombre} {a.docente.apellido}" if a.docente else None,
            "dia": a.dia, "hora": a.hora_inicio, "sede": a.sede.nombre if a.sede else None,
            "modalidad": a.modalidad,
        })
    # Build result grouped by sede → carrera → anno → catedras
    sedes_result = {}
    for r in rows:
        sede_n = r[1]
        carrera = r[2]
        anno = r[3]
        cod = r[4]
        nombre_plan = r[5]
        dia_tm = r[6] or ''
        hora_tm = r[7] or ''
        dia_tn = r[8] or ''
        hora_tn = r[9] or ''
        cat_info = cat_map.get(cod, None)
        cat_id = cat_info["id"] if cat_info else None
        insc = total_map.get(cat_id, 0) if cat_id else 0
        # Criterio
        if insc >= 10: criterio = "ABRIR"
        elif insc > 0: criterio = "ASINCRÓNICA"
        else: criterio = "SIN ALUMNOS"
        # Current assignment
        asig_actual = asig_map.get(cat_id, []) if cat_id else []
        tiene_docente = any(a['docente'] for a in asig_actual)
        docente_actual = ', '.join([a['docente'] for a in asig_actual if a['docente']]) or None
        horario_actual_tm = next((f"{a['dia']} {a['hora']}" for a in asig_actual if a['hora'] and a['hora'] < '15:00'), None)
        horario_actual_tn = next((f"{a['dia']} {a['hora']}" for a in asig_actual if a['hora'] and a['hora'] >= '15:00'), None)
        if sede_n not in sedes_result: sedes_result[sede_n] = {}
        if carrera not in sedes_result[sede_n]: sedes_result[sede_n][carrera] = {}
        if anno not in sedes_result[sede_n][carrera]: sedes_result[sede_n][carrera][anno] = []
        sedes_result[sede_n][carrera][anno].append({
            "codigo": cod, "nombre": nombre_plan,
            "inscriptos": insc, "criterio": criterio,
            "sugerencia_tm": f"{dia_tm} {hora_tm}".strip() if dia_tm else None,
            "sugerencia_tn": f"{dia_tn} {hora_tn}".strip() if dia_tn else None,
            "actual_tm": horario_actual_tm, "actual_tn": horario_actual_tn,
            "docente": docente_actual, "tiene_docente": tiene_docente,
        })
    return {"sedes": sedes_result, "plan_importado": True, "total_registros": len(rows)}


# ===== v16.0: Control de Inscripciones =====
CARRERA_NORMALIZE = {
    'ACOMPAÑANTE TERAPÉUTICO': 'TECNICO SUPERIOR EN ACOMPAÑANTE  TERAPEUTICO',
    'ADMINISTRACIÓN DE EMPRESAS': 'TECNICO SUPERIOR EN ADMINISTRACION DE EMPRESAS',
    'ADMINISTRACION DE EMPRESAS': 'TECNICO SUPERIOR EN ADMINISTRACION DE EMPRESAS',
    'ADMINISTRACION AGROPECUARIA': 'TECNICO SUPERIOR  EN ADMINISTRACION AGROPECUARIA',
    'ADMINISTRACIÓN AGROPECUARIA': 'TECNICO SUPERIOR  EN ADMINISTRACION AGROPECUARIA',
    'ADMINISTRACIÓN BANCARIA': 'TECNICO SUPERIOR EN ADMINISTRACION BANCARIA',
    'CIENCIA DE DATOS E INTELIGENCIA ARTIFICIAL': 'TECNICO SUPERIOR EN CIENCIA DE DATOS',
    'COMERCIO INTERNACIONAL': 'TECNICO SUPERIOR EN COMERCIO',
    'COUNSELING': 'TECNICO SUPERIOR EN COUSELING',
    'DESARROLLO HUMANO': 'TECNICO SUPERIOR EN DESARROLLO HUMANO',
    'DESPACHANTE DE ADUANAS': 'TECNICO SUPERIOR EN DESPACHO ADUANERO',
    'FINANZAS': 'TECNICO SUPERIOR EN FINANZAS',
    'GESTORÍA': 'TECNICO SUPERIOR EN GESTORIA',
    'GUIA DE TURISMO': 'TECNICO SUPERIOR EN GUIA DE TURISMO',
    'GUÍA DE TURISMO': 'TECNICO SUPERIOR EN GUIA DE TURISMO',
    'HOTELERÍA': 'TECNICO SUPERIOR EN HOTELERIA',
    'LOGÍSTICA': 'TECNICO SUPERIOR EN LOGISTICA',
    'LOGISTICA': 'TECNICO SUPERIOR EN LOGISTICA',
    'MARKETING': 'TECNICO SUPERIOR EN MARKETING',
    'NEGOCIOS DIGITALES': 'TECNICO SUPERIOR EN NEGOCIOS DIGITALES',
    'ORGANIZACIÓN DE EVENTOS': 'TECNICO SUPERIOR EN ORGANIZACION DE EVENTOS',
    'PERIODISMO DEPORTIVO': 'TECNICO SUPERIOR EN PERIODISMO DEPORTIVO',
    'PSICOPEDAGOGÍA': 'TECNICO SUPERIOR EN PSICOPEDAGOGIA',
    'PUBLICIDAD': 'TECNICO SUPERIOR EN PUBLICIDAD',
    'RECURSOS HUMANOS': 'TECNICO SUPERIOR EN RECURSOS HUMANOS',
    'RELACIONES PUBLICAS': 'TECNICO SUPERIOR EN RELACIONES PUBLICAS',
    'RELACIONES PÚBLICAS': 'TECNICO SUPERIOR EN RELACIONES PUBLICAS',
    'RÉGIMEN ADUANERO': 'TECNICO SUPERIOR EN REGIMEN ADUANERO',
    'SEGURIDAD E HIGIENE': 'TECNICO SUPERIOR EN HIGIENE Y SEGURIDAD',
    'SEGUROS': 'TECNICO SUPERIOR EN SEGUROS',
    'TRABAJO SOCIAL': 'TECNICO SUPERIOR EN TRABAJO SOCIAL',
    'TURISMO': 'TECNICO SUPERIOR EN TURISMO',
}

def _extract_carrera(curso):
    import re
    c = curso.upper().strip()
    c = re.split(r'\s*[\(]\s*(AVELLANEDA|CABALLITO|VICENTE|LINIERS|ONLINE|VIRTUAL|PILAR|MONTE|LA PLATA)', c)[0].strip()
    c = re.split(r'\s*[\-]\s*(CIED|CURSADA|CFE|DCFE|RDCFE|RMEDGC|RMEIGC|NO DISP|RD |RM |RES)', c)[0].strip().strip(' -')
    return CARRERA_NORMALIZE.get(c, None)

def _calc_anno(fecha_str, inicio_str):
    import math
    try:
        if '/' in fecha_str:
            parts = fecha_str.split('/'); year = int(parts[2]) if len(parts[2]) == 4 else 2000 + int(parts[2])
        elif '-' in fecha_str:
            year = int(fecha_str[:4])
        else: return '3ER AÑO'
        inicio = (inicio_str or '').strip().lower()
        start = (year - 2020) * 2 + (1 if 'agosto' in inicio or 'ago' in inicio else 0)
        current = (2026 - 2020) * 2  # Marzo 2026
        cuats = max(1, current - start + 1)
        return {1:'1ER AÑO', 2:'2DO AÑO', 3:'3ER AÑO'}[min(3, math.ceil(cuats / 2))]
    except: return '3ER AÑO'

@app.post("/api/control-inscripciones")
async def control_inscripciones(file: UploadFile = File(...), cuatrimestre_id: int = 1, db: Session = Depends(get_db)):
    from openpyxl import load_workbook
    import io, re
    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    # 1) Build plan: (sede, carrera_upper) → {anno → set of codes}
    from sqlalchemy import text
    plan = {}; cat_names_db = {}
    try:
        for r in db.execute(text("SELECT sede, carrera, anno, codigo_catedra, nombre_catedra FROM plan_carrera")).fetchall():
            sede_plan = r[0].upper().strip()
            carrera_up = r[1].upper().strip()
            key = (sede_plan, carrera_up)
            if key not in plan: plan[key] = {}
            if r[2] not in plan[key]: plan[key][r[2]] = set()
            plan[key][r[2]].add(r[3])
            cat_names_db[r[3]] = (r[4] or '')[:30]
    except: pass
    for c in db.query(Catedra).all():
        cat_names_db[c.codigo] = c.nombre[:30]
    # 2) Build inscripcion lookup: DNI → set of codigos inscritos
    insc_q = db.query(Inscripcion)
    if cuatrimestre_id and cuatrimestre_id > 0:
        insc_q = insc_q.filter(Inscripcion.cuatrimestre_id == cuatrimestre_id)
    all_insc = insc_q.all()
    dni_to_codes = {}
    for ins in all_insc:
        if not ins.alumno: continue
        dni_raw = (ins.alumno.dni or '').strip()
        if not dni_raw: continue
        dni = dni_raw.replace('.0','').lstrip('0') or dni_raw
        if dni not in dni_to_codes: dni_to_codes[dni] = set()
        if ins.catedra: dni_to_codes[dni].add(ins.catedra.codigo)
        if dni_raw not in dni_to_codes: dni_to_codes[dni_raw] = dni_to_codes[dni]
    # 3) Sede normalization for plan matching
    def norm_sede_plan(sede_str, curso_str):
        """Determine which plan sede to use based on student's sede and curso."""
        s = sede_str.upper() if sede_str else ''
        c = curso_str.upper() if curso_str else ''
        # CIED/Online students → use CIED plan
        if 'CIED' in c or 'ONLINE' in c or 'VIRTUAL SINCR' in c or 'DIEGEP' in c:
            return 'CIED'
        if 'AVELL' in s: return 'AVELLANEDA'
        if 'CABAL' in s or 'CABALI' in s: return 'CABALLITO'
        if 'VICENTE' in s or 'VTE' in s: return 'VICENTE LOPEZ'
        if 'LINIE' in s: return 'CABALLITO'  # Liniers uses Caballito plan
        return 'CIED'  # Default fallback
    # 4) Find plan key helper
    def find_plan_key(sede_norm, carrera_norm):
        if not carrera_norm: return None
        ck = carrera_norm.upper().strip()
        # Exact match
        key = (sede_norm, ck)
        if key in plan: return key
        # Fuzzy match on carrera
        for (s, c) in plan:
            if s == sede_norm and (ck in c or c in ck or ck[:25] == c[:25]):
                return (s, c)
        # Fallback: try any sede
        for (s, c) in plan:
            if ck in c or c in ck or ck[:25] == c[:25]:
                return (s, c)
        return None
    # 4) Process control file
    results = []; stats = {'total': 0, 'ok': 0, 'faltan': 0, 'sobran': 0, 'sin_plan': 0, 'sin_insc': 0, 'doble': 0}
    for ws in wb.worksheets:
        if ws.title.upper() in ['INSTRUCTIVO', 'BCE Y BEA']: continue
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 2: continue
            vals = list(row)
            if len(vals) < 11: continue
            nombre = f"{str(vals[1] or '')} {str(vals[2] or '')}".strip()
            dni_raw = str(vals[3] or '').strip()
            # Normalize DNI: handle float (29128688.0), int, string
            try: dni = str(int(float(dni_raw)))
            except: dni = dni_raw.replace('.0','').strip()
            fecha = str(vals[4] or '').strip()
            inicio = str(vals[5] or '').strip()
            sede = str(vals[9] or '').strip()
            curso = str(vals[10] or '').strip()
            if not nombre or not curso or curso == 'None': continue
            stats['total'] += 1
            is_doble = 'DOBLE' in curso.upper()
            if is_doble: stats['doble'] += 1
            anno = _calc_anno(fecha, inicio)
            carrera_norm = _extract_carrera(curso)
            sede_norm = norm_sede_plan(sede, curso)
            plan_key = find_plan_key(sede_norm, carrera_norm)
            # Get what they SHOULD take (excluding Prácticas Profesionalizantes)
            should_codes = set()
            if plan_key and anno in plan.get(plan_key, {}):
                for cod in plan[plan_key][anno]:
                    name_upper = (cat_names_db.get(cod, '') or '').upper()
                    if 'PROFESIONALIZANTE' not in name_upper and 'PRACTICA PROFESIONAL' not in name_upper:
                        should_codes.add(cod)
            # Get what they ARE taking
            actual_codes = dni_to_codes.get(dni, set())
            # Compare
            if not plan_key:
                estado = 'SIN_PLAN'
                stats['sin_plan'] += 1
                faltantes = []; sobrantes = []
            elif not actual_codes:
                estado = 'SIN_INSCRIPCIONES'
                stats['sin_insc'] += 1
                faltantes = sorted(should_codes)
                sobrantes = []
            else:
                faltantes = sorted(should_codes - actual_codes)
                sobrantes = sorted(actual_codes - should_codes)
                if not faltantes and not sobrantes:
                    estado = 'CORRECTO'
                    stats['ok'] += 1
                elif faltantes and sobrantes:
                    estado = 'FALTAN_Y_SOBRAN'
                    stats['faltan'] += 1
                elif faltantes:
                    estado = 'FALTAN_MATERIAS'
                    stats['faltan'] += 1
                else:
                    estado = 'MATERIAS_EXTRA'
                    stats['sobran'] += 1
            results.append({
                'dni': dni, 'nombre': nombre, 'sede': sede, 'curso': curso[:55],
                'anno': anno, 'carrera_plan': (f"{plan_key[0]} → {plan_key[1]}" if plan_key else (carrera_norm or '?'))[:55],
                'estado': estado, 'is_doble': is_doble,
                'debe_cursar': [f"{c} ({cat_names_db.get(c, '')})" for c in sorted(should_codes)],
                'inscripto_a': [f"{c} ({cat_names_db.get(c, '')})" for c in sorted(actual_codes)],
                'faltantes': [f"{c} ({cat_names_db.get(c, '')})" for c in faltantes],
                'sobrantes': [f"{c} ({cat_names_db.get(c, '')})" for c in sobrantes],
            })
    wb.close()
    results.sort(key=lambda x: (0 if x['estado']=='CORRECTO' else 1 if x['estado']=='MATERIAS_EXTRA' else 2 if x['estado'].startswith('FALTAN') else 3, x['nombre']))
    return {"results": results, "stats": stats, "total_results": len(results),
        "_debug": {"plan_carreras": len(plan), "inscripciones_db": len(all_insc), "dnis_con_inscripciones": len(dni_to_codes), "cuatrimestre_id": cuatrimestre_id}}


@app.post("/api/control-inscripciones/exportar")
async def control_inscripciones_exportar(file: UploadFile = File(...), cuatrimestre_id: int = 1, db: Session = Depends(get_db)):
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from starlette.responses import Response
    import io, re, math
    from sqlalchemy import text
    content = await file.read()
    wb_in = load_workbook(io.BytesIO(content))
    # Same logic as control_inscripciones
    plan = {}; cat_names_db = {}
    try:
        for r in db.execute(text("SELECT sede, carrera, anno, codigo_catedra, nombre_catedra FROM plan_carrera")).fetchall():
            sede_plan = r[0].upper().strip(); carrera_up = r[1].upper().strip()
            key = (sede_plan, carrera_up)
            if key not in plan: plan[key] = {}
            if r[2] not in plan[key]: plan[key][r[2]] = set()
            plan[key][r[2]].add(r[3])
            cat_names_db[r[3]] = (r[4] or '')[:30]
    except: pass
    for c in db.query(Catedra).all(): cat_names_db[c.codigo] = c.nombre[:30]
    insc_q = db.query(Inscripcion)
    if cuatrimestre_id and cuatrimestre_id > 0: insc_q = insc_q.filter(Inscripcion.cuatrimestre_id == cuatrimestre_id)
    dni_to_codes = {}
    for ins in insc_q.all():
        if not ins.alumno: continue
        dni_raw = (ins.alumno.dni or '').strip()
        if not dni_raw: continue
        dni = dni_raw.replace('.0','').lstrip('0') or dni_raw
        if dni not in dni_to_codes: dni_to_codes[dni] = set()
        if ins.catedra: dni_to_codes[dni].add(ins.catedra.codigo)
        if dni_raw not in dni_to_codes: dni_to_codes[dni_raw] = dni_to_codes[dni]
    def _norm_sede_exp(sede_str, curso_str):
        s = (sede_str or '').upper(); c = (curso_str or '').upper()
        if 'CIED' in c or 'ONLINE' in c or 'VIRTUAL SINCR' in c or 'DIEGEP' in c: return 'CIED'
        if 'AVELL' in s: return 'AVELLANEDA'
        if 'CABAL' in s or 'CABALI' in s: return 'CABALLITO'
        if 'VICENTE' in s or 'VTE' in s: return 'VICENTE LOPEZ'
        return 'CIED'
    def _find_pk(sede_n, cn):
        if not cn: return None
        ck = cn.upper().strip()
        for (s, c) in plan:
            if s == sede_n and (ck in c or c in ck or ck[:25] == c[:25]): return (s, c)
        for (s, c) in plan:
            if ck in c or c in ck or ck[:25] == c[:25]: return (s, c)
        return None
    # Process
    rows_out = []
    for ws in wb_in.worksheets:
        if ws.title.upper() in ['INSTRUCTIVO', 'BCE Y BEA']: continue
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 2: continue
            vals = list(row)
            if len(vals) < 11: continue
            nombre = f"{str(vals[1] or '')} {str(vals[2] or '')}".strip()
            try: dni = str(int(float(str(vals[3] or ''))))
            except: dni = str(vals[3] or '').replace('.0','').strip()
            fecha = str(vals[4] or '').strip(); inicio = str(vals[5] or '').strip()
            sede = str(vals[9] or '').strip(); curso = str(vals[10] or '').strip()
            if not nombre or not curso or curso == 'None': continue
            anno = _calc_anno(fecha, inicio)
            carrera_norm = _extract_carrera(curso)
            sede_n = _norm_sede_exp(sede, curso)
            plan_key = _find_pk(sede_n, carrera_norm)
            should = set()
            if plan_key and anno in plan.get(plan_key, {}):
                for cod in plan[plan_key][anno]:
                    nm = (cat_names_db.get(cod,'') or '').upper()
                    if 'PROFESIONALIZANTE' not in nm and 'PRACTICA PROFESIONAL' not in nm: should.add(cod)
            actual = dni_to_codes.get(dni, set())
            faltan = sorted(should - actual); sobran = sorted(actual - should)
            if not plan_key: estado = 'SIN PLAN'
            elif not actual: estado = 'SIN INSCRIPCIONES'
            elif not faltan and not sobran: estado = 'CORRECTO'
            elif faltan and sobran: estado = 'FALTAN Y SOBRAN'
            elif faltan: estado = 'FALTAN MATERIAS'
            else: estado = 'MATERIAS EXTRA'
            rows_out.append([nombre, dni, sede, curso[:55], anno, estado,
                ', '.join([f"{c} ({cat_names_db.get(c,'')})" for c in sorted(should)]),
                ', '.join([f"{c} ({cat_names_db.get(c,'')})" for c in sorted(actual)]),
                ', '.join([f"{c} ({cat_names_db.get(c,'')})" for c in faltan]),
                ', '.join([f"{c} ({cat_names_db.get(c,'')})" for c in sobran])])
    wb_in.close()
    # Build Excel
    wb_out = Workbook(); ws_out = wb_out.active; ws_out.title = 'Control Inscripciones'
    hdr_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    hdr_font = Font(bold=True, color='FFFFFF', size=10)
    fills = {'CORRECTO': PatternFill(start_color='D5F5E3', fill_type='solid'),
        'MATERIAS EXTRA': PatternFill(start_color='D6EAF8', fill_type='solid'),
        'FALTAN MATERIAS': PatternFill(start_color='FEF9E7', fill_type='solid'),
        'FALTAN Y SOBRAN': PatternFill(start_color='FADBD8', fill_type='solid'),
        'SIN INSCRIPCIONES': PatternFill(start_color='F5B7B1', fill_type='solid'),
        'SIN PLAN': PatternFill(start_color='EAECEE', fill_type='solid')}
    border = Border(left=Side(style='thin',color='CCCCCC'), right=Side(style='thin',color='CCCCCC'),
        top=Side(style='thin',color='CCCCCC'), bottom=Side(style='thin',color='CCCCCC'))
    headers = ['Alumno','DNI','Sede','Carrera','Año','Estado','Debe cursar','Inscripto a','Faltantes','Sobrantes']
    for col, h in enumerate(headers, 1):
        c = ws_out.cell(row=1, column=col, value=h); c.fill = hdr_fill; c.font = hdr_font; c.border = border
    for i, rd in enumerate(rows_out, 2):
        fill = fills.get(rd[5], fills['SIN PLAN'])
        for col, val in enumerate(rd, 1):
            c = ws_out.cell(row=i, column=col, value=val); c.fill = fill; c.border = border
            c.font = Font(size=9); c.alignment = Alignment(wrap_text=True, vertical='top')
    for i, w in enumerate([25,12,15,40,10,18,55,55,45,45], 1): ws_out.column_dimensions[chr(64+i)].width = w
    buf = io.BytesIO(); wb_out.save(buf); buf.seek(0)
    return Response(content=buf.read(), media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=control_inscripciones.xlsx'})


# ===== v16.0: Motor de sugerencias de armado de horarios =====
@app.get("/api/sugerencias-armado")
def get_sugerencias_armado(cuatrimestre_id: int = None, sede: str = None, db: Session = Depends(get_db)):
    from sqlalchemy import text
    # 1) Get plan_carrera entries
    q = "SELECT sede, carrera, anno, codigo_catedra, nombre_catedra FROM plan_carrera"
    if sede: q += f" WHERE sede = '{sede}'"
    q += " ORDER BY sede, carrera, anno"
    try: plan = db.execute(text(q)).fetchall()
    except: return {"sedes": {}, "stats": {}}
    if not plan: return {"sedes": {}, "stats": {}}
    # 2) Get inscriptos count
    total_q = "SELECT catedra_id, COUNT(*) FROM inscripciones"
    if cuatrimestre_id: total_q += f" WHERE cuatrimestre_id = {cuatrimestre_id}"
    total_q += " GROUP BY catedra_id"
    total_map = {}
    try:
        for r in db.execute(text(total_q)).fetchall(): total_map[r[0]] = r[1]
    except: pass
    # 3) Catedra map
    cat_map = {c.codigo: {"id": c.id, "nombre": c.nombre} for c in db.query(Catedra).all()}
    # 4) Current asignaciones
    asig_q = db.query(Asignacion)
    if cuatrimestre_id: asig_q = asig_q.filter(Asignacion.cuatrimestre_id == cuatrimestre_id)
    asigs = asig_q.all()
    asig_map = {}  # catedra_id → [{docente, dia, hora, sede}]
    docente_busy = {}  # docente_id → set of (dia, hora)
    for a in asigs:
        if a.catedra_id not in asig_map: asig_map[a.catedra_id] = []
        asig_map[a.catedra_id].append({
            "docente_id": a.docente_id,
            "docente": f"{a.docente.nombre} {a.docente.apellido}" if a.docente else None,
            "dia": a.dia, "hora": a.hora_inicio,
            "sede": a.sede.nombre if a.sede else None,
        })
        if a.docente_id and a.dia and a.hora_inicio:
            if a.docente_id not in docente_busy: docente_busy[a.docente_id] = set()
            docente_busy[a.docente_id].add((a.dia, a.hora_inicio))
    # 5) Docentes with availability and references
    docentes = db.query(Docente).all()
    disp_map = {}
    try:
        for r in db.execute(text("SELECT docente_id, dia, hora FROM docente_disponibilidad WHERE disponible = TRUE")).fetchall():
            if r[0] not in disp_map: disp_map[r[0]] = set()
            disp_map[r[0]].add((r[1], r[2]))
    except: pass
    # Build docente lookup for suggestions
    doc_info = {}
    for d in docentes:
        refs = (getattr(d, 'catedras_referencia', '') or '').strip()
        ref_codes = [r.strip() for r in refs.split(',') if r.strip()] if refs else []
        avail = disp_map.get(d.id, set())
        busy = docente_busy.get(d.id, set())
        free = avail - busy  # Available and not already assigned
        doc_info[d.id] = {
            "id": d.id, "nombre": f"{d.nombre} {d.apellido}",
            "ref_codes": ref_codes, "free_slots": free,
            "sedes": [ds.sede.nombre for ds in (d.sedes or []) if ds.sede],
        }
    # 6) Build result per sede → carrera → anno → catedras with suggestions
    sedes_result = {}
    stats = {"total": 0, "con_docente": 0, "sugerido": 0, "sin_sugerencia": 0, "asincronica": 0}
    for r in plan:
        sede_n, carrera, anno, cod, nombre_plan = r[0], r[1], r[2], r[3], r[4]
        cat_info = cat_map.get(cod)
        cat_id = cat_info["id"] if cat_info else None
        insc = total_map.get(cat_id, 0) if cat_id else 0
        criterio = "ABRIR" if insc >= 10 else ("ASINCRÓNICA" if insc > 0 else "SIN ALUMNOS")
        # Current assignments
        current_asigs = asig_map.get(cat_id, []) if cat_id else []
        tiene_docente = any(a['docente'] for a in current_asigs)
        docente_actual = ', '.join(set(a['docente'] for a in current_asigs if a['docente'])) or None
        horarios_actuales = [f"{a['dia']} {a['hora']}" for a in current_asigs if a['dia']] if current_asigs else []
        # Determine status and suggestion
        estado = "asignado"  # green
        sugerencia_docente = None
        if criterio == "ABRIR" and tiene_docente:
            estado = "asignado"
            stats["con_docente"] += 1
        elif criterio == "ABRIR" and not tiene_docente:
            # Find suggestion
            candidatos = []
            for did, dinfo in doc_info.items():
                # Check if docente has this cátedra as reference
                if cod in dinfo["ref_codes"]:
                    # Has free slots → valid candidate
                    if dinfo["free_slots"]:
                        score = 10
                        candidatos.append({"id": did, "nombre": dinfo["nombre"], "score": score,
                            "free": len(dinfo["free_slots"]), "slots": list(dinfo["free_slots"])[:3]})
                # If no ref_codes at all, NOT a candidate (we only suggest docentes who can teach this)
            candidatos.sort(key=lambda x: -x["score"])
            if candidatos:
                estado = "sugerido"  # blue
                sugerencia_docente = candidatos[0]["nombre"]
                stats["sugerido"] += 1
            else:
                estado = "sin_sugerencia"  # red
                stats["sin_sugerencia"] += 1
        elif criterio == "ASINCRÓNICA":
            estado = "asincronica"
            stats["asincronica"] += 1
        else:
            estado = "sin_alumnos"
        stats["total"] += 1
        if sede_n not in sedes_result: sedes_result[sede_n] = {}
        if carrera not in sedes_result[sede_n]: sedes_result[sede_n][carrera] = {}
        if anno not in sedes_result[sede_n][carrera]: sedes_result[sede_n][carrera][anno] = []
        sedes_result[sede_n][carrera][anno].append({
            "codigo": cod, "nombre": nombre_plan, "inscriptos": insc, "criterio": criterio,
            "estado": estado, "docente_actual": docente_actual, "sugerencia_docente": sugerencia_docente,
            "horarios": horarios_actuales,
        })
    return {"sedes": sedes_result, "stats": stats}


# ===== v15.0: Solapamientos entre carreras (lógica correcta) =====
@app.get("/api/solapamientos-carreras")
def get_solapamientos_carreras(cuatrimestre_id: int = None, db: Session = Depends(get_db)):
    from sqlalchemy import text
    # 1) Get plan_carrera
    try: plan = db.execute(text("SELECT sede, carrera, anno, codigo_catedra FROM plan_carrera")).fetchall()
    except: return {"presencial": [], "cied": [], "docentes": [], "total": 0}
    if not plan: return {"presencial": [], "cied": [], "docentes": [], "total": 0, "sin_plan": True}
    carrera_cats = {}
    for r in plan:
        key = (r[0], r[1], r[2])
        if key not in carrera_cats: carrera_cats[key] = []
        if r[3] not in carrera_cats[key]: carrera_cats[key].append(r[3])
    # 2) Get asignaciones
    asig_q = db.query(Asignacion).filter(Asignacion.dia.isnot(None), Asignacion.hora_inicio.isnot(None))
    if cuatrimestre_id: asig_q = asig_q.filter(Asignacion.cuatrimestre_id == cuatrimestre_id)
    asigs = asig_q.all()
    # Build: code → {sede_name → set of (dia, hora)} and code → all slots
    code_sede_slots = {}; code_all_slots = {}; code_names = {}; code_docentes = {}
    docente_schedule = {}
    for a in asigs:
        if not a.catedra or not a.dia or a.dia == 'Pend.' or not a.hora_inicio or a.hora_inicio == 'Pend.': continue
        cod = a.catedra.codigo
        sede_n = a.sede.nombre if a.sede else 'Remoto'
        doc_name = f"{a.docente.nombre} {a.docente.apellido}" if a.docente else None
        code_names[cod] = a.catedra.nombre
        slot = (a.dia, a.hora_inicio)
        if cod not in code_sede_slots: code_sede_slots[cod] = {}
        if sede_n not in code_sede_slots[cod]: code_sede_slots[cod][sede_n] = set()
        code_sede_slots[cod][sede_n].add(slot)
        if cod not in code_all_slots: code_all_slots[cod] = set()
        code_all_slots[cod].add(slot)
        # Track docente per (cod, dia, hora)
        if doc_name:
            if cod not in code_docentes: code_docentes[cod] = {}
            code_docentes[cod][(a.dia, a.hora_inicio)] = doc_name
            # Docente schedule for type 2
            if doc_name not in docente_schedule: docente_schedule[doc_name] = []
            docente_schedule[doc_name].append({"dia": a.dia, "hora": a.hora_inicio, "cod": cod, "nombre": a.catedra.nombre, "sede": sede_n})
    # Helper to get docente for a (cod, dia, hora)
    def get_doc(cod, dia, hora):
        return (code_docentes.get(cod) or {}).get((dia, hora))
    # === TIPO 1: PRESENCIALES (por sede) ===
    conf_presencial = []
    for (sede_p, carrera, anno), codigos in carrera_cats.items():
        if not anno or sede_p.upper() == 'CIED': continue
        cat_slots = {}
        for cod in codigos:
            for s_name, slots in (code_sede_slots.get(cod) or {}).items():
                if s_name.upper().replace(' ','')[:4] == sede_p.upper().replace(' ','')[:4]:
                    if cod not in cat_slots: cat_slots[cod] = set()
                    cat_slots[cod].update(slots)
        cods = list(cat_slots.keys())
        for i in range(len(cods)):
            for j in range(i+1, len(cods)):
                overlap = cat_slots[cods[i]] & cat_slots[cods[j]]
                for (dia, hora) in overlap:
                    conf_presencial.append({"sede_plan": sede_p, "carrera": carrera, "anno": anno,
                        "dia": dia, "hora": hora,
                        "catedras_en_conflicto": [
                            {"codigo": cods[i], "nombre": code_names.get(cods[i],''), "docente": get_doc(cods[i], dia, hora)},
                            {"codigo": cods[j], "nombre": code_names.get(cods[j],''), "docente": get_doc(cods[j], dia, hora)}],
                        "tipo": "presencial"})
    # === TIPO 1b: CIED (conflicto solo si NO hay combinación posible) ===
    conf_cied = []
    for (sede_p, carrera, anno), codigos in carrera_cats.items():
        if not anno or sede_p.upper() != 'CIED': continue
        cat_slots = {cod: code_all_slots.get(cod, set()) for cod in codigos if code_all_slots.get(cod)}
        cods = list(cat_slots.keys())
        for i in range(len(cods)):
            for j in range(i+1, len(cods)):
                can_avoid = any(sa != sb for sa in cat_slots[cods[i]] for sb in cat_slots[cods[j]])
                if not can_avoid and cat_slots[cods[i]] & cat_slots[cods[j]]:
                    dia, hora = list(cat_slots[cods[i]] & cat_slots[cods[j]])[0]
                    conf_cied.append({"sede_plan": "CIED", "carrera": carrera, "anno": anno,
                        "dia": dia, "hora": hora,
                        "catedras_en_conflicto": [
                            {"codigo": cods[i], "nombre": code_names.get(cods[i],''), "docente": get_doc(cods[i], dia, hora)},
                            {"codigo": cods[j], "nombre": code_names.get(cods[j],''), "docente": get_doc(cods[j], dia, hora)}],
                        "tipo": "cied"})
    # === TIPO 2: DOCENTES (mismo docente, distinta cátedra, mismo dia+hora) ===
    conf_docentes = []
    for doc_name, schedule in docente_schedule.items():
        slots = {}
        for s in schedule:
            key = (s['dia'], s['hora'])
            if key not in slots: slots[key] = []
            slots[key].append(s)
        for (dia, hora), items in slots.items():
            cods_unicos = list(set(it['cod'] for it in items))
            if len(cods_unicos) < 2: continue
            conf_docentes.append({"docente": doc_name, "dia": dia, "hora": hora,
                "asignaciones": [{"codigo": it['cod'], "nombre": it['nombre'], "sede": it['sede']} for it in items],
                "tipo": "docente"})
    conf_presencial.sort(key=lambda x: (x['sede_plan'], x['carrera'], x['anno']))
    conf_docentes.sort(key=lambda x: (x['docente'], x['dia']))
    total = len(conf_presencial) + len(conf_cied) + len(conf_docentes)
    return {"presencial": conf_presencial, "cied": conf_cied, "docentes": conf_docentes, "total": total,
        "total_presencial": len(conf_presencial), "total_cied": len(conf_cied), "total_docentes": len(conf_docentes)}


# ===== v10.0: Exportar COMPLETO =====
@app.get("/api/exportar/planilla-trabajo")
def exportar_planilla_trabajo(cuatrimestre_id: int = None, solo_dictadas: bool = True,
                              sede: str = None, db: Session = Depends(get_db)):
    """PASO 3 del flujo. Excel con las cátedras que se dictan y sus inscriptos desglosados.
    Las últimas columnas van vacías para que el equipo las complete y reimporte el archivo
    en 'Importar Horarios y Designaciones' (mismo formato de columnas)."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from sqlalchemy import text
    if not cuatrimestre_id:
        cu = db.query(Cuatrimestre).first()
        cuatrimestre_id = cu.id if cu else 1

    dictado = _mapa_dictado(db, cuatrimestre_id)
    # Desglose de inscriptos por turno/sede
    desglose = {}
    try:
        q = ("SELECT catedra_id, turno, modalidad_alumno, sede_referencia, COUNT(*) "
             "FROM inscripciones WHERE cuatrimestre_id = :c GROUP BY catedra_id, turno, modalidad_alumno, sede_referencia")
        for cid, turno, mod, sede, cnt in db.execute(text(q), {"c": cuatrimestre_id}).fetchall():
            d = desglose.setdefault(cid, {"total": 0, "tm": 0, "tn": 0, "cied": 0,
                                          "av": 0, "cab": 0, "vl": 0})
            d["total"] += cnt
            if mod == 'virtual': d["cied"] += cnt
            else:
                sl = (sede or '').lower()
                if 'avellaneda' in sl: d["av"] += cnt
                elif 'caballito' in sl: d["cab"] += cnt
                elif 'vicente' in sl: d["vl"] += cnt
            if turno == 'Mañana': d["tm"] += cnt
            elif turno == 'Noche': d["tn"] += cnt
    except Exception:
        db.rollback()

    # Docentes ya asignados (si ya se importó una vuelta)
    asigs = db.query(Asignacion).filter(Asignacion.cuatrimestre_id == cuatrimestre_id).all()
    asig_por_cat = {}
    for a in asigs:
        if a.dia or a.docente_id:
            asig_por_cat.setdefault(a.catedra_id, []).append(a)

    wb = Workbook(); ws = wb.active; ws.title = "Planilla de trabajo"
    hdr_fill = PatternFill("solid", fgColor="1E3A5F")
    info_fill = PatternFill("solid", fgColor="EAF2F8")
    edit_fill = PatternFill("solid", fgColor="FFF9E6")
    thin = Side(style="thin", color="BBBBBB")
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["CODIGO", "MATERIA", "DIA", "HORA INICIO", "HORA FIN", "SEDE", "DOCENTE",
               "LINK MEET", "— INSCRIPTOS —", "TOTAL", "TM", "TN", "CIED",
               "AVELL", "CABA", "VTE LOPEZ", "SUGERENCIA"]
    ws.append(headers)
    for i, _ in enumerate(headers, 1):
        c = ws.cell(row=1, column=i)
        c.fill = hdr_fill; c.font = Font(bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = borde

    cats = sorted(db.query(Catedra).all(), key=lambda c: sort_key_codigo(c.codigo))
    # v17.0: filtro por sede para que varias personas trabajen en paralelo sin pisarse
    clave_sede = None
    if sede:
        s = sede.lower()
        if 'avell' in s: clave_sede = 'av'
        elif 'cabal' in s: clave_sede = 'cab'
        elif 'vicente' in s or 'vte' in s: clave_sede = 'vl'
        elif 'cied' in s or 'online' in s: clave_sede = 'cied'
    fila = 2; escritas = 0
    for cat in cats:
        info = dictado.get(cat.id, {})
        if solo_dictadas and not info.get("se_dicta"): continue
        d = desglose.get(cat.id, {"total": 0, "tm": 0, "tn": 0, "cied": 0, "av": 0, "cab": 0, "vl": 0})
        if clave_sede and d.get(clave_sede, 0) < 1: continue
        total = d["total"]
        sugerencia = "ABRIR (con docente)" if total >= 10 else ("ASINCRÓNICA (pregrabada)" if total >= 1 else "SIN ALUMNOS")
        existentes = asig_por_cat.get(cat.id, [])
        filas_cat = existentes if existentes else [None]
        for a in filas_cat:
            try: num = int(cat.codigo.replace('c.', ''))
            except Exception: num = cat.codigo
            ws.append([
                num, cat.nombre,
                (a.dia if a else "") or "",
                (a.hora_inicio if a else "") or "",
                (getattr(a, 'hora_fin', None) if a else "") or "",
                (a.sede.nombre if a and a.sede else "") or "",
                (f"{a.docente.nombre} {a.docente.apellido}" if a and a.docente else ""),
                getattr(cat, 'link_meet', None) or "",
                "", total, d["tm"], d["tn"], d["cied"], d["av"], d["cab"], d["vl"], sugerencia,
            ])
            for col in range(1, len(headers) + 1):
                c = ws.cell(row=fila, column=col)
                c.border = borde
                if col <= 8: c.fill = edit_fill
                else: c.fill = info_fill
            fila += 1; escritas += 1

    anchos = [9, 38, 12, 13, 12, 15, 28, 34, 4, 9, 7, 7, 8, 8, 8, 11, 22]
    for i, w in enumerate(anchos, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "C2"

    # Hoja de instrucciones
    ws2 = wb.create_sheet("Instructivo")
    for linea in [
        [f"PLANILLA DE TRABAJO — Armado de horarios{' — ' + sede if sede else ''}"],
        [""],
        ["Columnas AMARILLAS (A-H): las completa el equipo."],
        ["Columnas CELESTES (I-Q): sólo informativas, las calcula el sistema. No hace falta tocarlas."],
        [""],
        ["Cómo completar:"],
        ["CODIGO", "Ya viene cargado. Es el número de cátedra sin el prefijo c."],
        ["DIA", "LUNES / MARTES / MIERCOLES / JUEVES / VIERNES / SABADO"],
        ["HORA INICIO", "Formato 08:00 — se admiten medias horas (08:30, 09:30, etc.)"],
        ["HORA FIN", "Formato 09:30. Si se deja vacío el sistema asume 1h30 de duración."],
        ["SEDE", "Avellaneda / Caballito / Vte Lopez / CIED"],
        ["DOCENTE", "Nombre tal como figura en la sección Docentes."],
        ["", "IMPORTANTE: si se deja VACÍO, la cátedra se dicta igual pero como ASINCRÓNICA"],
        ["", "(video pregrabado, sin docente en vivo). Eso es válido y esperado."],
        ["LINK MEET", "Opcional. Si se completa, queda vinculado a la cátedra."],
        [""],
        ["Una vez completada, subir este mismo archivo en:"],
        ["Importar → Importar Horarios y Designaciones"],
        [""],
        ["El sistema va a mostrar una vista previa con los solapamientos detectados"],
        ["antes de aplicar los cambios."],
    ]:
        ws2.append(linea)
    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 80
    ws2.cell(row=1, column=1).font = Font(bold=True, size=14, color="1E3A5F")

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=planilla_trabajo{'_' + sede.replace(' ','_') if sede else ''}.xlsx"})


@app.get("/api/exportar/horarios")
def exportar_horarios(cuatrimestre_id: int = None, modulos: str = None, db: Session = Depends(get_db)):
    """modulos: lista separada por comas de hojas a incluir. Si viene vacío exporta todo.
    Ver /api/exportar/modulos para los identificadores disponibles."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from sqlalchemy import text
    wb = Workbook()
    q = db.query(Asignacion)
    if cuatrimestre_id: q = q.filter(Asignacion.cuatrimestre_id == cuatrimestre_id)
    asigs = sorted(q.all(), key=lambda a: sort_key_codigo(a.catedra.codigo if a.catedra else 'c.9999'))
    # -- Inscriptos desglosados --
    insc_q = """SELECT catedra_id, turno, modalidad_alumno, sede_referencia, COUNT(*) FROM inscripciones WHERE modalidad_alumno IS NOT NULL"""
    if cuatrimestre_id: insc_q += f" AND cuatrimestre_id = {cuatrimestre_id}"
    insc_q += " GROUP BY catedra_id, turno, modalidad_alumno, sede_referencia"
    insc_desg = {}
    try:
        for r in db.execute(text(insc_q)).fetchall():
            cid, turno, mod, sede, cnt = r
            if cid not in insc_desg:
                insc_desg[cid] = {'total':0,'tm_av':0,'tm_cab':0,'tm_vl':0,'tm_cied':0,'tn_av':0,'tn_cab':0,'tn_vl':0,'tn_cied':0,'virt':0}
            d = insc_desg[cid]; d['total'] += cnt
            es_cied = (mod == 'virtual')
            sk = None
            if not es_cied:
                sl = (sede or '').lower()
                if 'avellaneda' in sl: sk = 'av'
                elif 'caballito' in sl: sk = 'cab'
                elif 'vicente' in sl: sk = 'vl'
            if turno == 'Mañana': d[f'tm_{sk}' if sk else 'tm_cied'] += cnt
            elif turno == 'Noche': d[f'tn_{sk}' if sk else 'tn_cied'] += cnt
            else: d['virt'] += cnt
    except: pass
    total_insc_q = "SELECT catedra_id, COUNT(*) FROM inscripciones"
    if cuatrimestre_id: total_insc_q += f" WHERE cuatrimestre_id = {cuatrimestre_id}"
    total_insc_q += " GROUP BY catedra_id"
    total_insc_map = {}
    try:
        for r in db.execute(text(total_insc_q)).fetchall(): total_insc_map[r[0]] = r[1]
    except: pass
    hf = Font(bold=True, color="FFFFFF", size=10)
    YELLOW = PatternFill("solid", fgColor="FFFFCC")
    def make_hdr(ws, headers, color='1D6F42'):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = hf; cell.fill = PatternFill("solid", fgColor=color); cell.alignment = Alignment(horizontal="center")
    all_cats = sorted(db.query(Catedra).all(), key=lambda c: sort_key_codigo(c.codigo))

    # ========== HOJA 1 (PRINCIPAL): Inscriptos de todas las cátedras unificadas ==========
    ws0 = wb.active; ws0.title = "Inscriptos unificados"
    make_hdr(ws0, ["#","Código","Cátedra","TM Avellaneda","TM Caballito","TM Vicente López","TM CIED","Total TM","TN Avellaneda","TN Caballito","TN Vicente López","TN CIED","Total TN","CIED Virtual","Sede Avellaneda","Sede Caballito","Sede V.López","Sede CIED","TOTAL"], '0F766E')
    for i, cat in enumerate(all_cats, 1):
        d = insc_desg.get(cat.id, {})
        tm_t = d.get('tm_av',0)+d.get('tm_cab',0)+d.get('tm_vl',0)+d.get('tm_cied',0)
        tn_t = d.get('tn_av',0)+d.get('tn_cab',0)+d.get('tn_vl',0)+d.get('tn_cied',0)
        s_av = d.get('tm_av',0)+d.get('tn_av',0)
        s_cab = d.get('tm_cab',0)+d.get('tn_cab',0)
        s_vl = d.get('tm_vl',0)+d.get('tn_vl',0)
        s_cied = d.get('tm_cied',0)+d.get('tn_cied',0)+d.get('virt',0)
        tot = total_insc_map.get(cat.id, 0)
        if tot == 0 and d.get('total',0) == 0: continue
        ws0.append([i, cat.codigo, cat.nombre,
            d.get('tm_av',0) or '', d.get('tm_cab',0) or '', d.get('tm_vl',0) or '', d.get('tm_cied',0) or '', tm_t or '',
            d.get('tn_av',0) or '', d.get('tn_cab',0) or '', d.get('tn_vl',0) or '', d.get('tn_cied',0) or '', tn_t or '',
            d.get('virt',0) or '', s_av or '', s_cab or '', s_vl or '', s_cied or '', tot or ''])
        if tot >= 10:
            tiene = any(1 for a in asigs if a.catedra_id == cat.id)
            if not tiene:
                for cell in ws0[ws0.max_row]: cell.fill = YELLOW
    for col, w in [('A',4),('B',8),('C',28),('D',14),('E',13),('F',16),('G',9),('H',9),('I',14),('J',13),('K',16),('L',9),('M',9),('N',11),('O',14),('P',13),('Q',13),('R',10),('S',7)]:
        ws0.column_dimensions[col].width = w

    # ========== HOJA 2: Totalidad General (antes "Todas las sedes") ==========
    ws1 = wb.create_sheet("Totalidad General")
    ws1.append(["#","Código","Cátedra","TM Av","TM Cab","TM VL","TM CIED","Tot TM","TN Av","TN Cab","TN VL","TN CIED","Tot TN","CIED Virt","TOTAL","Docente","Turno Doc.","Modalidad","Día","Hora","Sede"])
    for cell in ws1[1]:
        cell.font = hf; cell.fill = PatternFill("solid", fgColor="1D6F42"); cell.alignment = Alignment(horizontal="center")
    for i, a in enumerate(asigs, 1):
        d = insc_desg.get(a.catedra_id, {})
        tm_t = d.get('tm_av',0)+d.get('tm_cab',0)+d.get('tm_vl',0)+d.get('tm_cied',0)
        tn_t = d.get('tn_av',0)+d.get('tn_cab',0)+d.get('tn_vl',0)+d.get('tn_cied',0)
        tot = total_insc_map.get(a.catedra_id, 0)
        mod = a.modalidad or ''
        td = 'Mañana' if 'tm' in mod else ('Noche' if 'tn' in mod else ('Asincrónica' if mod == 'asincronica' else 'Presencial'))
        ws1.append([i, a.catedra.codigo if a.catedra else "", a.catedra.nombre if a.catedra else "",
            d.get('tm_av',0) or '', d.get('tm_cab',0) or '', d.get('tm_vl',0) or '', d.get('tm_cied',0) or '', tm_t or '',
            d.get('tn_av',0) or '', d.get('tn_cab',0) or '', d.get('tn_vl',0) or '', d.get('tn_cied',0) or '', tn_t or '',
            d.get('virt',0) or '', tot or '',
            f"{a.docente.nombre} {a.docente.apellido}" if a.docente else "Sin asignar", td,
            mod, a.dia or "Pend.", a.hora_inicio or "Pend.", a.sede.nombre if a.sede else "Remoto"])
        if tot >= 10 and not a.docente_id:
            for cell in ws1[ws1.max_row]: cell.fill = YELLOW
    for col, w in [('A',4),('B',8),('C',28),('D',7),('E',7),('F',7),('G',7),('H',7),('I',7),('J',7),('K',7),('L',7),('M',7),('N',9),('O',7),('P',26),('Q',10),('R',12),('S',10),('T',7),('U',16)]:
        ws1.column_dimensions[col].width = w

    # ========== HOJAS POR SEDE ==========
    for sede_key, sede_name, color in [('av','Avellaneda','3B82F6'),('cab','Caballito','10B981'),('vl','Vicente López','F59E0B'),('cied','CIED','8B5CF6')]:
        ws = wb.create_sheet(sede_name[:31])
        if sede_key == 'cied':
            ws.append(["#","Código","Cátedra","TM CIED","TN CIED","Virtual","TOTAL","Docente","Turno","Día","Hora"])
            sede_asigs = [a for a in asigs if not a.sede_id or (a.modalidad and ('virtual' in a.modalidad or a.modalidad == 'asincronica'))]
        else:
            ws.append(["#","Código","Cátedra",f"TM {sede_name}",f"TN {sede_name}","TOTAL","Docente","Turno","Día","Hora"])
            sede_db = db.query(Sede).filter(Sede.nombre == sede_name).first()
            sede_asigs = [a for a in asigs if a.sede_id == (sede_db.id if sede_db else -1)]
        for cell in ws[1]:
            cell.font = hf; cell.fill = PatternFill("solid", fgColor=color); cell.alignment = Alignment(horizontal="center")
        for i, a in enumerate(sede_asigs, 1):
            d = insc_desg.get(a.catedra_id, {})
            mod = a.modalidad or ''
            td = 'Mañana' if 'tm' in mod else ('Noche' if 'tn' in mod else 'Otro')
            if sede_key == 'cied':
                tm_v=d.get('tm_cied',0); tn_v=d.get('tn_cied',0); vv=d.get('virt',0)
                ws.append([i, a.catedra.codigo if a.catedra else "", a.catedra.nombre if a.catedra else "",
                    tm_v or '', tn_v or '', vv or '', (tm_v+tn_v+vv) or '',
                    f"{a.docente.nombre} {a.docente.apellido}" if a.docente else "Sin asignar", td, a.dia or "Pend.", a.hora_inicio or "Pend."])
            else:
                tm_v=d.get(f'tm_{sede_key}',0); tn_v=d.get(f'tn_{sede_key}',0)
                ws.append([i, a.catedra.codigo if a.catedra else "", a.catedra.nombre if a.catedra else "",
                    tm_v or '', tn_v or '', (tm_v+tn_v) or '',
                    f"{a.docente.nombre} {a.docente.apellido}" if a.docente else "Sin asignar", td, a.dia or "Pend.", a.hora_inicio or "Pend."])
        for col, w in [('A',4),('B',8),('C',28),('D',14),('E',14),('F',10),('G',7),('H',26),('I',10),('J',10),('K',7)]:
            ws.column_dimensions[col].width = w

    # ========== HOJA: Listado Total Docentes (ordenado por sede) ==========
    docentes_db = db.query(Docente).order_by(Docente.apellido).all()
    ws_ld = wb.create_sheet("Listado Docentes")
    ws_ld.append(["#","Docente","DNI","Email","Horas","CFPEA","ISFTEA","Sedes","Cátedras asignadas","Cant.","Horarios",""])
    for cell in ws_ld[1]:
        cell.font = hf; cell.fill = PatternFill("solid", fgColor="334155"); cell.alignment = Alignment(horizontal="center")
    idx = 1
    # Ordenar por sede principal
    def sede_sort_key(doc):
        sds = [s.sede.nombre for s in (doc.sedes or []) if s.sede]
        return sds[0] if sds else 'ZZZ'
    for doc in sorted(docentes_db, key=sede_sort_key):
        sds = ', '.join([s.sede.nombre for s in (doc.sedes or []) if s.sede]) or 'Sin sede'
        doc_asigs = [a for a in asigs if a.docente_id == doc.id]
        cats_list = ', '.join([f"{a.catedra.codigo} {a.catedra.nombre[:20]}" for a in doc_asigs if a.catedra]) or 'SIN MATERIA ASIGNADA'
        horarios = ', '.join([f"{a.dia or '?'} {a.hora_inicio or '?'}" for a in doc_asigs]) or ''
        ws_ld.append([idx, f"{doc.nombre} {doc.apellido}", doc.dni, doc.email or '',
            getattr(doc, 'horas_asignadas', 0) or '',
            "Sí" if getattr(doc, 'sociedad_cfpea', False) else "",
            "Sí" if getattr(doc, 'sociedad_isftea', False) else "",
            sds, cats_list, len(doc_asigs), horarios, ''])
        idx += 1
    for col, w in [('A',4),('B',28),('C',12),('D',24),('E',7),('F',7),('G',8),('H',20),('I',34),('J',13),('K',10),('L',7)]:
        ws_ld.column_dimensions[col].width = w

    # ========== HOJA: Docentes TM ==========
    ws_dtm = wb.create_sheet("Docentes TM")
    ws_dtm.append(["#","Docente","DNI","Email","Horas","CFPEA","ISFTEA","Cátedra","Sede","Día","Hora"])
    for cell in ws_dtm[1]:
        cell.font = hf; cell.fill = PatternFill("solid", fgColor="D97706"); cell.alignment = Alignment(horizontal="center")
    def _is_tm(a):
        try: return a.hora_inicio and int(a.hora_inicio.split(':')[0]) < 14
        except: return False
    idx = 1
    for doc in docentes_db:
        for a in [a for a in asigs if a.docente_id == doc.id and _is_tm(a)]:
            ws_dtm.append([idx, f"{doc.nombre} {doc.apellido}", doc.dni, doc.email or '',
                getattr(doc, 'horas_asignadas', 0) or '',
                "Sí" if getattr(doc, 'sociedad_cfpea', False) else "",
                "Sí" if getattr(doc, 'sociedad_isftea', False) else "",
                f"{a.catedra.codigo} {a.catedra.nombre}" if a.catedra else "", a.sede.nombre if a.sede else "Remoto",
                a.dia or "Pend.", a.hora_inicio or "Pend."])
            idx += 1
    for col, w in [('A',4),('B',28),('C',12),('D',24),('E',7),('F',7),('G',8),('H',32),('I',16),('J',10),('K',7)]:
        ws_dtm.column_dimensions[col].width = w

    # ========== HOJA: Docentes TN ==========
    ws_dtn = wb.create_sheet("Docentes TN")
    ws_dtn.append(["#","Docente","DNI","Email","Horas","CFPEA","ISFTEA","Cátedra","Sede","Día","Hora"])
    for cell in ws_dtn[1]:
        cell.font = hf; cell.fill = PatternFill("solid", fgColor="4338CA"); cell.alignment = Alignment(horizontal="center")
    def _is_tn(a):
        try: return a.hora_inicio and int(a.hora_inicio.split(':')[0]) >= 14
        except: return False
    idx = 1
    for doc in docentes_db:
        for a in [a for a in asigs if a.docente_id == doc.id and _is_tn(a)]:
            ws_dtn.append([idx, f"{doc.nombre} {doc.apellido}", doc.dni, doc.email or '',
                getattr(doc, 'horas_asignadas', 0) or '',
                "Sí" if getattr(doc, 'sociedad_cfpea', False) else "",
                "Sí" if getattr(doc, 'sociedad_isftea', False) else "",
                f"{a.catedra.codigo} {a.catedra.nombre}" if a.catedra else "", a.sede.nombre if a.sede else "Remoto",
                a.dia or "Pend.", a.hora_inicio or "Pend."])
            idx += 1
    for col, w in [('A',4),('B',28),('C',12),('D',24),('E',7),('F',7),('G',8),('H',32),('I',16),('J',10),('K',7)]:
        ws_dtn.column_dimensions[col].width = w

    # ========== HOJA: Inscriptos por Curso ==========
    ws_cur = wb.create_sheet("Inscriptos por Curso")
    ws_cur.append(["#","Curso","Sede","Modalidad","Tipo","Alumnos (DNI)","Inscripciones"])
    for cell in ws_cur[1]:
        cell.font = hf; cell.fill = PatternFill("solid", fgColor="6B21A8"); cell.alignment = Alignment(horizontal="center")
    c_q = """SELECT curso_nombre, COUNT(DISTINCT alumno_id), COUNT(*) FROM inscripciones WHERE curso_nombre IS NOT NULL AND curso_nombre != ''"""
    if cuatrimestre_id: c_q += f" AND cuatrimestre_id = {cuatrimestre_id}"
    c_q += " GROUP BY curso_nombre ORDER BY COUNT(*) DESC"
    try:
        for i, r in enumerate(db.execute(text(c_q)).fetchall(), 1):
            curso_raw, unicos, total = r[0], r[1], r[2]
            es_cied = 'CIED' in curso_raw.upper()
            m_s = re.search(r'\(([^)]+)\)', curso_raw)
            sede_r = normalizar_sede(m_s.group(1).strip()) if m_s else 'Sin sede'
            mod = 'CIED' if (es_cied or sede_r in ['Online - Interior']) else 'Presencial'
            es_bce = 'BCE' in curso_raw.upper() or 'SECUNDARIO' in curso_raw.upper()
            tipo = 'BCE' if es_bce else ('BEA' if 'BEA' in curso_raw.upper() else 'Superior')
            ws_cur.append([i, curso_raw, sede_r, mod, tipo, unicos, total])
    except: pass
    for col, w in [('A',4),('B',60),('C',18),('D',12),('E',10),('F',16),('G',14)]:
        ws_cur.column_dimensions[col].width = w

    # ========== HOJA: Necesitan Docente (con sede+turno) ==========
    ws_nd = wb.create_sheet("Necesitan Docente")
    ws_nd.append(["#","Código","Cátedra","Inscriptos","Doc. necesarios","Doc. actuales","Faltan","Sedes asignadas","Desglose inscriptos"])
    for cell in ws_nd[1]:
        cell.font = hf; cell.fill = PatternFill("solid", fgColor="DC2626"); cell.alignment = Alignment(horizontal="center")
    nd_data = get_catedras_necesitan_docente(cuatrimestre_id, db)
    for i, nd in enumerate(nd_data, 1):
        sedes_asig = ', '.join([f"{sa['turno']} {sa['sede']} ({sa['docente']})" for sa in nd.get('sedes_asignadas', [])]) or 'Sin docente'
        desglose = ', '.join([f"{ap['turno']} {ap['sede']}: {ap['inscriptos']}" for ap in nd.get('aperturas_info', [])])
        ws_nd.append([i, nd['codigo'], nd['nombre'], nd['inscriptos_total'],
            nd['docs_necesarios'], nd['docentes_asignados'], nd['faltan'], sedes_asig, desglose])
        if nd['faltan'] > 0:
            for cell in ws_nd[ws_nd.max_row]: cell.fill = YELLOW
    for col, w in [('A',4),('B',8),('C',30),('D',10),('E',12),('F',10),('G',8),('H',40),('I',50)]:
        ws_nd.column_dimensions[col].width = w

    # ========== v7.0: 3 nuevas solapas de criterio de apertura ==========
    criterio = get_criterio_apertura(cuatrimestre_id, db)

    # --- HOJA: Abrir Cátedra (>=10 inscriptos total) ---
    ws_abrir = wb.create_sheet("Abrir cátedra")
    ws_abrir.append(["#","Código","Cátedra","Total inscriptos","Docentes sugeridos","Docentes actuales","Faltan","Estado"])
    for cell in ws_abrir[1]:
        cell.font = hf; cell.fill = PatternFill("solid", fgColor="059669"); cell.alignment = Alignment(horizontal="center")
    for i, item in enumerate(criterio['abrir'], 1):
        estado = "✅ Cubierto" if item['faltan'] == 0 else "⚠️ Faltan docentes"
        ws_abrir.append([i, item['codigo'], item['nombre'], item['total'],
            item['docentes_sugeridos'], item['docentes_actuales'], item['faltan'], estado])
        if item['faltan'] > 0:
            for cell in ws_abrir[ws_abrir.max_row]: cell.fill = YELLOW
    for col, w in [('A',4),('B',8),('C',30),('D',14),('E',16),('F',16),('G',8),('H',18)]:
        ws_abrir.column_dimensions[col].width = w

    # --- HOJA: Cátedra Asincrónica (1-9 inscriptos total) ---
    ws_asinc = wb.create_sheet("Cátedra asincrónica")
    ws_asinc.append(["#","Código","Cátedra","Total inscriptos","Observación"])
    for cell in ws_asinc[1]:
        cell.font = hf; cell.fill = PatternFill("solid", fgColor="7C3AED"); cell.alignment = Alignment(horizontal="center")
    for i, item in enumerate(criterio['asincronica'], 1):
        ws_asinc.append([i, item['codigo'], item['nombre'], item['total'], "1 a 9 inscriptos → Se dicta de forma asincrónica (pregrabada)"])
    for col, w in [('A',4),('B',8),('C',30),('D',14),('E',50)]:
        ws_asinc.column_dimensions[col].width = w

    # --- HOJA: Cátedra sin alumnos ---
    ws_sin = wb.create_sheet("Cátedra sin alumnos")
    ws_sin.append(["#","Código","Cátedra","Inscriptos"])
    for cell in ws_sin[1]:
        cell.font = hf; cell.fill = PatternFill("solid", fgColor="6B7280"); cell.alignment = Alignment(horizontal="center")
    idx = 1
    for cat in all_cats:
        ic = total_insc_map.get(cat.id, 0)
        if ic == 0:
            ws_sin.append([idx, cat.codigo, cat.nombre, 0])
            idx += 1
    for col, w in [('A',4),('B',8),('C',30),('D',12)]:
        ws_sin.column_dimensions[col].width = w

    # ========== HOJA: Solapamientos ==========
    ws_sol = wb.create_sheet("Solapamientos")
    ws_sol.append(["#","Tipo","Severidad","Mensaje","Día","Hora"])
    for cell in ws_sol[1]:
        cell.font = hf; cell.fill = PatternFill("solid", fgColor="B91C1C"); cell.alignment = Alignment(horizontal="center")
    for i, s in enumerate(get_solapamientos(cuatrimestre_id, db), 1):
        ws_sol.append([i, s.get('tipo',''), s.get('severidad',''), s.get('mensaje',''), s.get('dia',''), s.get('hora','')])
    for col, w in [('A',4),('B',12),('C',12),('D',60),('E',12),('F',7)]:
        ws_sol.column_dimensions[col].width = w

    # ========== HOJA: Horarios por Día y Sede ==========
    DIAS_ORDEN = ['Lunes','Martes','Miércoles','Jueves','Viernes','Sábado']
    HORAS_EXPORT = ['07:00','08:00','09:00','10:00','11:00','12:00','13:00','14:00','17:00','18:00','19:00','20:00','21:00','22:00','23:00']
    for sede_db_obj in [None] + list(db.query(Sede).all()):
        if sede_db_obj:
            sede_asigs_dia = [a for a in asigs if a.sede_id == sede_db_obj.id]
            sheet_name = f"Horario {sede_db_obj.nombre}"[:31]
            color_h = '2563EB'
        else:
            sede_asigs_dia = asigs
            sheet_name = "Horario General"
            color_h = '1E3A5F'
        if not sede_asigs_dia and sede_db_obj: continue
        ws_dia = wb.create_sheet(sheet_name)
        ws_dia.append(["Hora"] + DIAS_ORDEN)
        for cell in ws_dia[1]:
            cell.font = hf; cell.fill = PatternFill("solid", fgColor=color_h); cell.alignment = Alignment(horizontal="center")
        for hora in HORAS_EXPORT:
            row_data = [hora]
            for dia in DIAS_ORDEN:
                celdas = [a for a in sede_asigs_dia if a.dia == dia and a.hora_inicio == hora]
                if celdas:
                    txt = " | ".join([
                        f"{a.catedra.codigo} {(a.docente.apellido if a.docente else 'Sin doc.')}"
                        for a in celdas
                    ])
                else:
                    txt = ""
                row_data.append(txt)
            ws_dia.append(row_data)
        ws_dia.column_dimensions['A'].width = 7
        for col in ['B','C','D','E','F','G']:
            ws_dia.column_dimensions[col].width = 35

    # ========== v12.0: HOJA Criterio de Decisión ==========
    ws_dec = wb.create_sheet("Criterio de Decisión")
    ws_dec.append(["#","Código","Cátedra","Total inscriptos","Criterio sistema","Docentes sugeridos","Decisión tomada","Notas"])
    for cell in ws_dec[1]:
        cell.font = hf; cell.fill = PatternFill("solid", fgColor="1E40AF"); cell.alignment = Alignment(horizontal="center")
    for i, cat in enumerate(all_cats, 1):
        ic = total_insc_map.get(cat.id, 0)
        if ic >= 10:
            criterio_txt = f"ABRIR ({ic} inscriptos)"
            docs_sug = 1 if ic <= 100 else (1 + -(-max(0, ic - 100) // 100))
        elif ic > 0:
            criterio_txt = f"ASINCRÓNICA ({ic} inscriptos)"
            docs_sug = 0
        else:
            criterio_txt = "SIN ALUMNOS"
            docs_sug = 0
        decision = getattr(cat, 'decision_apertura', '') or ''
        notas_c = getattr(cat, 'notas', '') or ''
        ws_dec.append([i, cat.codigo, cat.nombre, ic, criterio_txt, docs_sug or '', decision, notas_c])
        if ic >= 10 and not decision:
            for cell in ws_dec[ws_dec.max_row]: cell.fill = YELLOW
    for col, w in [('A',4),('B',8),('C',30),('D',14),('E',22),('F',16),('G',28),('H',30)]:
        ws_dec.column_dimensions[col].width = w

    # ========== v10.0: BORRADOR HORARIOS TM-TN (formato como usan ellas) ==========
    DIAS_ORDEN2 = ['Lunes','Martes','Miércoles','Jueves','Viernes','Sábado']
    HORAS_TM = ['07:00','08:00','09:00','10:00','11:00','12:00','13:00','14:00']
    HORAS_TN = ['17:00','18:00','19:00','20:00','21:00','22:00','23:00']
    ws_bor = wb.create_sheet("Borrador Horarios")
    ws_bor.append(["Código","Cátedra","Día","Hora","Sede","Docente","Estado","Notas"])
    for cell in ws_bor[1]:
        cell.font = hf; cell.fill = PatternFill("solid", fgColor="1E3A5F"); cell.alignment = Alignment(horizontal="center")
    PENDIENTE_FILL = PatternFill("solid", fgColor="FEF3C7")
    HEADER_DIA = PatternFill("solid", fgColor="FBBF24")
    row_num = 2
    for dia in DIAS_ORDEN2:
        # TM header
        ws_bor.append([f"{dia.upper()} - TM",'','','','','','',''])
        for cell in ws_bor[ws_bor.max_row]:
            cell.font = Font(bold=True, size=11); cell.fill = HEADER_DIA
        row_num += 1
        dia_asigs_tm = sorted([a for a in asigs if a.dia == dia and a.hora_inicio in HORAS_TM],
            key=lambda a: (a.hora_inicio or '', sort_key_codigo(a.catedra.codigo if a.catedra else 'c.9999')))
        for a in dia_asigs_tm:
            cat = a.catedra
            doc_nombre = f"{a.docente.nombre} {a.docente.apellido}" if a.docente else ""
            estado = "✅ Asignado" if a.docente_id else "⚠️ PENDIENTE"
            notas_cat = ""
            try: notas_cat = getattr(cat, 'notas', '') or ''
            except: pass
            ws_bor.append([
                cat.codigo if cat else "", f"{cat.nombre}" if cat else "",
                dia, a.hora_inicio or "", a.sede.nombre if a.sede else "Remoto",
                doc_nombre or "SIN DOCENTE", estado, notas_cat
            ])
            if not a.docente_id:
                for cell in ws_bor[ws_bor.max_row]: cell.fill = PENDIENTE_FILL
            row_num += 1
        ws_bor.append(['','','','','','','',''])  # blank row
        row_num += 1
        # TN header
        ws_bor.append([f"{dia.upper()} - TN",'','','','','','',''])
        for cell in ws_bor[ws_bor.max_row]:
            cell.font = Font(bold=True, size=11); cell.fill = PatternFill("solid", fgColor="6366F1")
            cell.font = Font(bold=True, size=11, color="FFFFFF")
        row_num += 1
        dia_asigs_tn = sorted([a for a in asigs if a.dia == dia and a.hora_inicio in HORAS_TN],
            key=lambda a: (a.hora_inicio or '', sort_key_codigo(a.catedra.codigo if a.catedra else 'c.9999')))
        for a in dia_asigs_tn:
            cat = a.catedra
            doc_nombre = f"{a.docente.nombre} {a.docente.apellido}" if a.docente else ""
            estado = "✅ Asignado" if a.docente_id else "⚠️ PENDIENTE"
            notas_cat = ""
            try: notas_cat = getattr(cat, 'notas', '') or ''
            except: pass
            ws_bor.append([
                cat.codigo if cat else "", f"{cat.nombre}" if cat else "",
                dia, a.hora_inicio or "", a.sede.nombre if a.sede else "Remoto",
                doc_nombre or "SIN DOCENTE", estado, notas_cat
            ])
            if not a.docente_id:
                for cell in ws_bor[ws_bor.max_row]: cell.fill = PENDIENTE_FILL
            row_num += 1
        ws_bor.append(['','','','','','','',''])  # separator
        ws_bor.append(['','','','','','','',''])
        row_num += 2
    for col, w in [('A',8),('B',32),('C',12),('D',10),('E',16),('F',28),('G',14),('H',30)]:
        ws_bor.column_dimensions[col].width = w

    # ========== v13.0: HORARIOS POR CARRERA Y SEDE ==========
    try:
        plan_rows = db.execute(text("SELECT sede, carrera, anno, codigo_catedra, nombre_catedra, dia_tm, hora_tm, dia_tn, hora_tn FROM plan_carrera ORDER BY sede, carrera, anno")).fetchall()
    except: plan_rows = []
    if plan_rows:
        # Group by sede
        plan_by_sede = {}
        for r in plan_rows:
            s = r[0]
            if s not in plan_by_sede: plan_by_sede[s] = []
            plan_by_sede[s].append(r)
        # Asig lookup
        asig_lookup_exp = {}
        for a in asigs:
            cid = a.catedra_id
            if cid not in asig_lookup_exp: asig_lookup_exp[cid] = []
            asig_lookup_exp[cid].append(a)
        cat_code_to_id = {c.codigo: c.id for c in all_cats}
        for sede_name_p, sede_rows in plan_by_sede.items():
            ws_pc = wb.create_sheet(f"Plan {sede_name_p}"[:31])
            ws_pc.append(["Carrera","Año","Código","Cátedra","Inscr.","Criterio",
                "Sugerencia TM","Sugerencia TN","Docente actual","Horario actual TM","Horario actual TN","Estado"])
            for cell in ws_pc[1]:
                cell.font = hf; cell.fill = PatternFill("solid", fgColor="1E40AF"); cell.alignment = Alignment(horizontal="center")
            prev_carrera = ''
            for r in sede_rows:
                sede_p, carrera_p, anno_p, cod_p, nombre_p, dtm, htm, dtn, htn = r
                cat_id_p = cat_code_to_id.get(cod_p)
                insc_p = total_insc_map.get(cat_id_p, 0) if cat_id_p else 0
                crit = "ABRIR" if insc_p >= 10 else ("ASINCRÓNICA" if insc_p > 0 else "SIN ALUMNOS")
                # Current assignment
                cat_asigs = asig_lookup_exp.get(cat_id_p, [])
                doc_act = ', '.join([f"{a.docente.nombre} {a.docente.apellido}" for a in cat_asigs if a.docente]) or ''
                h_act_tm = next((f"{a.dia} {a.hora_inicio}" for a in cat_asigs if a.hora_inicio and a.hora_inicio < '15:00'), '')
                h_act_tn = next((f"{a.dia} {a.hora_inicio}" for a in cat_asigs if a.hora_inicio and a.hora_inicio >= '15:00'), '')
                estado = "✅ Con docente" if doc_act else ("🎥 Asincrónica" if crit == "ASINCRÓNICA" else ("⚠️ PENDIENTE" if crit == "ABRIR" else "—"))
                show_carrera = carrera_p if carrera_p != prev_carrera else ''
                prev_carrera = carrera_p
                ws_pc.append([show_carrera, anno_p, cod_p, nombre_p, insc_p or '', crit,
                    f"{dtm} {htm}".strip() if dtm else '', f"{dtn} {htn}".strip() if dtn else '',
                    doc_act, h_act_tm, h_act_tn, estado])
                if crit == "ABRIR" and not doc_act:
                    for cell in ws_pc[ws_pc.max_row]: cell.fill = YELLOW
                elif crit == "ASINCRÓNICA":
                    for cell in ws_pc[ws_pc.max_row]: cell.fill = PatternFill("solid", fgColor="E9D5FF")
            for col, w in [('A',38),('B',14),('C',8),('D',34),('E',8),('F',14),('G',18),('H',18),('I',28),('J',18),('K',18),('L',16)]:
                ws_pc.column_dimensions[col].width = w

    # ========== v14.0: Solapamientos entre carreras ==========
    solap_carr = get_solapamientos_carreras(cuatrimestre_id, db)
    if solap_carr.get('conflictos'):
        ws_sc = wb.create_sheet("Solapamientos Carreras")
        ws_sc.append(["#","Sede","Carrera","Año","Día","Hora","Cátedras en conflicto","Docentes","Sugerencia"])
        for cell in ws_sc[1]:
            cell.font = hf; cell.fill = PatternFill("solid", fgColor="B91C1C"); cell.alignment = Alignment(horizontal="center")
        for i, conf in enumerate(solap_carr['conflictos'], 1):
            cats_txt = ' vs '.join([f"{c['codigo']} {c['nombre']}" for c in conf['catedras_en_conflicto']])
            docs_txt = ', '.join([c['docente'] or 'Sin doc.' for c in conf['catedras_en_conflicto']])
            ws_sc.append([i, conf['sede_plan'], conf['carrera'][:40], conf['anno'],
                conf['dia'], conf['hora'], cats_txt, docs_txt, conf['sugerencia']])
            for cell in ws_sc[ws_sc.max_row]: cell.fill = YELLOW
        for col, w in [('A',4),('B',16),('C',40),('D',12),('E',12),('F',10),('G',50),('H',30),('I',60)]:
            ws_sc.column_dimensions[col].width = w

    # ========== v17.0: exportar sólo los módulos pedidos ==========
    if modulos:
        pedidos = {m.strip().lower() for m in modulos.split(',') if m.strip()}
        # Cada módulo se corresponde con una o varias hojas (por nombre o prefijo)
        mapa_modulos = {
            'general': ['Totalidad General'],
            'sedes': ['__SEDE__'],
            'docentes': ['Listado Docentes'],
            'docentes_tm': ['Docentes TM'],
            'docentes_tn': ['Docentes TN'],
            'cursos': ['Inscriptos por Curso'],
            'necesitan_docente': ['Necesitan Docente'],
            'apertura': ['Abrir cátedra', 'Cátedra asincrónica', 'Cátedra sin alumnos'],
            'solapamientos': ['Solapamientos', 'Solapamientos Carreras'],
            'decision': ['Criterio de Decisión'],
            'borrador': ['Borrador Horarios'],
            'horarios_dia': ['__DIA__'],
            'plan_carrera': ['__PLAN__'],
        }
        nombres_sede = {s.nombre[:31] for s in db.query(Sede).all()}
        conservar = set()
        for m in pedidos:
            for etiqueta in mapa_modulos.get(m, []):
                if etiqueta == '__SEDE__': conservar |= nombres_sede
                elif etiqueta == '__DIA__':
                    conservar |= {d for d in ['Lunes','Martes','Miércoles','Jueves','Viernes','Sábado']}
                elif etiqueta == '__PLAN__':
                    conservar |= {n for n in wb.sheetnames if n.startswith('Plan ')}
                else: conservar.add(etiqueta)
        for nombre_hoja in list(wb.sheetnames):
            if nombre_hoja not in conservar:
                del wb[nombre_hoja]
        if not wb.sheetnames:
            ws_vacio = wb.create_sheet("Sin datos")
            ws_vacio.append(["No hay hojas para los módulos seleccionados."])

    output = io.BytesIO(); wb.save(output); output.seek(0)
    from datetime import datetime
    ahora = datetime.now().strftime("%d-%m-%Y_%H_%M_hs")
    sufijo = f"_{modulos.replace(',', '-')[:40]}" if modulos else ""
    nombre = f"IEA_Horarios{'_Cuat' + str(cuatrimestre_id) if cuatrimestre_id else ''}{sufijo}_{ahora}.xlsx"
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={nombre}"})


@app.get("/api/exportar/modulos")
def listar_modulos_exportables():
    """Catálogo para armar el selector de exportación parcial."""
    return [
        {"id": "general", "nombre": "Totalidad General", "icono": "📊"},
        {"id": "sedes", "nombre": "Una hoja por sede", "icono": "🏫"},
        {"id": "docentes", "nombre": "Listado de Docentes", "icono": "👨‍🏫"},
        {"id": "docentes_tm", "nombre": "Docentes Turno Mañana", "icono": "🌅"},
        {"id": "docentes_tn", "nombre": "Docentes Turno Noche", "icono": "🌙"},
        {"id": "cursos", "nombre": "Inscriptos por Curso", "icono": "📋"},
        {"id": "necesitan_docente", "nombre": "Necesitan Docente", "icono": "🔴"},
        {"id": "apertura", "nombre": "Apertura (abrir / asincrónica / sin alumnos)", "icono": "🚦"},
        {"id": "solapamientos", "nombre": "Solapamientos", "icono": "⚠️"},
        {"id": "decision", "nombre": "Criterio de Decisión", "icono": "🎯"},
        {"id": "borrador", "nombre": "Borrador de Horarios", "icono": "📝"},
        {"id": "horarios_dia", "nombre": "Horarios por Día", "icono": "📅"},
        {"id": "plan_carrera", "nombre": "Plan por Carrera y Sede", "icono": "🎓"},
    ]

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
